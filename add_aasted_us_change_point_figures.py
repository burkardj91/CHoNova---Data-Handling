from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw, ImageFont

from analyze_aasted_runs_zone_patterns import (
    COMPARISON_CSV,
    COMPARISON_RUN_NAME,
    FINAL_REPORT_WORKBOOK,
    OUTPUT_DIR,
    REFERENCE_CSV,
    REFERENCE_RUN_NAME,
    read_run,
)


FIGURE_DIR = OUTPUT_DIR / "us_change_point_figures"
COLORS = {
    "Rx1Tx1": (20, 83, 45),
    "Rx2Tx2": (146, 64, 14),
    "threshold": (37, 99, 235),
    "reference": (107, 114, 128),
    "detachment_onset": (220, 38, 38),
    "change_point": (156, 163, 175),
    "selected": (22, 163, 74),
}


def font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        r"C:\Windows\Fonts\arialbd.ttf" if bold else r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\calibrib.ttf" if bold else r"C:\Windows\Fonts\calibri.ttf",
    ]
    for path in candidates:
        if Path(path).exists():
            return ImageFont.truetype(path, size=size)
    return ImageFont.load_default()


def downsample(df: pd.DataFrame, max_points: int = 2200) -> pd.DataFrame:
    if len(df) <= max_points:
        return df
    return df.iloc[:: int(np.ceil(len(df) / max_points))].copy()


def draw_polyline(draw: ImageDraw.ImageDraw, points: list[tuple[float, float]], color: tuple[int, int, int], width: int = 3) -> None:
    clean = [(float(x), float(y)) for x, y in points if np.isfinite(x) and np.isfinite(y)]
    if len(clean) > 1:
        draw.line(clean, fill=color, width=width, joint="curve")


def make_us_change_point_figure(
    run_name: str,
    df: pd.DataFrame,
    summary_row: pd.Series,
    change_points: pd.DataFrame,
    out_path: Path,
) -> Path:
    channel = str(summary_row["channel"])
    signal_col = str(summary_row["signal_column"])
    onset = float(summary_row["detachment_onset_s"])
    end = float(summary_row["analysis_window_end_s"])
    offset = pd.to_numeric(pd.Series([summary_row.get("detachment_offset_s")]), errors="coerce").iloc[0]
    threshold = float(summary_row["threshold_10pct_lower"])
    reference = float(summary_row["reference_us_level"])
    zone = str(summary_row.get("detachment_offset_zone", ""))
    seg = df[(df["elapsed_s"] >= onset) & (df["elapsed_s"] <= end)][["elapsed_s", signal_col]].dropna().copy()
    seg = downsample(seg)
    cp = change_points[(change_points["run"].eq(run_name)) & (change_points["channel"].eq(channel))].copy()

    x_min = onset
    x_max = end
    values = pd.concat([seg[signal_col], pd.Series([threshold, reference])], ignore_index=True)
    y_min = float(np.nanquantile(values, 0.01))
    y_max = float(np.nanquantile(values, 0.99))
    pad = max((y_max - y_min) * 0.12, 0.05)
    y_min -= pad
    y_max += pad

    width, height = 1280, 720
    left, right, top, bottom = 92, 36, 82, 138
    plot_w = width - left - right
    plot_h = height - top - bottom

    def sx(t: float) -> float:
        return left + (t - x_min) / max(x_max - x_min, 1e-9) * plot_w

    def sy(v: float) -> float:
        return top + (y_max - v) / max(y_max - y_min, 1e-9) * plot_h

    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img, "RGBA")
    title = f"{run_name} {channel}: T-corrected US detachment change points"
    draw.text((left, 24), title, fill=(17, 24, 39), font=font(22, True))
    subtitle = "Search domain: detachment onset to run end. Complete detachment = first positive change point with US >= 10% lower pre-deposition threshold."
    draw.text((left, 52), subtitle, fill=(71, 85, 105), font=font(13))

    # grid
    for frac in np.linspace(0, 1, 6):
        y = top + frac * plot_h
        val = y_max - frac * (y_max - y_min)
        draw.line((left, y, left + plot_w, y), fill=(226, 232, 240), width=1)
        draw.text((left - 68, y - 8), f"{val:.2f}", fill=(51, 65, 85), font=font(11))
    for frac in np.linspace(0, 1, 7):
        x = left + frac * plot_w
        val = x_min + frac * (x_max - x_min)
        draw.line((x, top, x, top + plot_h), fill=(241, 245, 249), width=1)
        draw.text((x - 24, top + plot_h + 18), f"{val:.0f}", fill=(51, 65, 85), font=font(11))

    draw.rectangle((left, top, left + plot_w, top + plot_h), outline=(51, 65, 85), width=2)
    points = [(sx(t), sy(v)) for t, v in zip(seg["elapsed_s"], seg[signal_col])]
    draw_polyline(draw, points, COLORS.get(channel, (31, 41, 55)), width=3)

    for value, label, color in [
        (reference, "pre-deposition reference", COLORS["reference"]),
        (threshold, "10% lower threshold", COLORS["threshold"]),
    ]:
        y = sy(value)
        draw.line((left, y, left + plot_w, y), fill=color, width=2)
        draw.text((left + plot_w - 230, y - 18), f"{label}: {value:.3f}", fill=color, font=font(11, True))

    x_onset = sx(onset)
    draw.line((x_onset, top, x_onset, top + plot_h), fill=COLORS["detachment_onset"], width=3)
    draw.text((x_onset + 5, top + 8), "detachment onset", fill=COLORS["detachment_onset"], font=font(11, True))

    for _, row in cp.iterrows():
        t = float(row["change_point_s"])
        x = sx(t)
        met = bool(row.get("threshold_met", False))
        color = COLORS["selected"] if met and pd.notna(offset) and abs(t - float(offset)) < 1e-6 else COLORS["change_point"]
        draw.line((x, top, x, top + plot_h), fill=color, width=3 if color == COLORS["selected"] else 2)
        idx = int(row["change_point_index"])
        label = f"CP {idx}"
        if color == COLORS["selected"]:
            label += " selected"
        draw.text((x + 4, top + plot_h - 22 - (idx % 4) * 16), label, fill=color, font=font(10, True))

    draw.text((left, height - 96), f"First CP after onset: {summary_row.get('first_change_point_after_onset_s', np.nan):.1f} s", fill=(17, 24, 39), font=font(13, True))
    draw.text((left, height - 72), f"Selected complete offset: {offset:.1f} s absolute; {summary_row.get('detachment_onset_to_offset_s', np.nan):.1f} s after onset; zone/domain: {zone}", fill=(17, 24, 39), font=font(13, True))
    draw.text((left, height - 48), f"Detected CPs until decision: {summary_row.get('change_points_until_detachment_or_last', np.nan)} | total CPs in domain: {summary_row.get('change_points_detected_total', np.nan)}", fill=(51, 65, 85), font=font(12))
    draw.text((left + plot_w / 2 - 55, height - 16), "elapsed time in run [s]", fill=(17, 24, 39), font=font(13))
    draw.text((12, top + plot_h // 2), "T-corrected US", fill=(17, 24, 39), font=font(13))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    img.save(out_path)
    return out_path


def add_us_change_point_figures() -> Path:
    runs = {
        REFERENCE_RUN_NAME: read_run(REFERENCE_CSV),
        COMPARISON_RUN_NAME: read_run(COMPARISON_CSV),
    }
    summary = pd.read_excel(FINAL_REPORT_WORKBOOK, sheet_name="Aasted Detachment Summary")
    change_points = pd.read_excel(FINAL_REPORT_WORKBOOK, sheet_name="Aasted US Change Points")
    figure_rows = []
    for _, row in summary.iterrows():
        run_name = str(row["run"])
        if run_name not in runs:
            continue
        filename = f"{run_name}_{row['channel']}_us_change_points.png".replace("/", "_")
        path = FIGURE_DIR / filename
        make_us_change_point_figure(run_name, runs[run_name], row, change_points, path)
        figure_rows.append(
            {
                "run": run_name,
                "channel": row["channel"],
                "figure_file": str(path),
                "first_change_point_after_onset_s": row.get("first_change_point_after_onset_s"),
                "selected_offset_after_onset_s": row.get("detachment_onset_to_offset_s"),
                "detachment_offset_zone": row.get("detachment_offset_zone"),
                "analysis_domain": row.get("analysis_domain"),
            }
        )

    wb = load_workbook(FINAL_REPORT_WORKBOOK)
    for sheet_name in ["US Change Point Figures", "US Pattern Figures"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
    ws = wb.create_sheet("US Change Point Figures", 2)
    ws["A1"] = "Ultrasound detachment change-point decision figures"
    ws["A1"].font = Font(bold=True, size=15, color="1F4E78")
    ws["A1"].fill = PatternFill("solid", fgColor="D9EAF7")
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 28
    row_anchor = 3
    for item in figure_rows:
        ws.cell(row_anchor, 1, f"{item['run']} | {item['channel']}")
        ws.cell(row_anchor, 1).font = Font(bold=True, color="1F4E78")
        ws.cell(row_anchor + 1, 1, f"Domain: {item['analysis_domain']}; first CP after onset: {item['first_change_point_after_onset_s']:.1f} s; selected offset after onset: {item['selected_offset_after_onset_s']:.1f} s; zone: {item['detachment_offset_zone']}")
        ws.cell(row_anchor + 1, 1).alignment = Alignment(wrap_text=True)
        img = XLImage(item["figure_file"])
        img.width = 960
        img.height = 540
        ws.add_image(img, f"A{row_anchor + 2}")
        row_anchor += 31
    ws.column_dimensions["A"].width = 120
    wb.save(FINAL_REPORT_WORKBOOK)
    return FINAL_REPORT_WORKBOOK


def main() -> None:
    print(add_us_change_point_figures())


if __name__ == "__main__":
    main()
