from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

from analyze_lab_trials import load_setup


WORKSPACE = Path(r"C:\Users\BurkardJohannes\Documents\CHoNova _ Data Understanding")
LAB_INPUT = WORKSPACE / "inputs" / "lab_trials"
AAK_INPUT = WORKSPACE / "inputs" / "AAK_cooling_tunnel"
AAK_OUTPUT = WORKSPACE / "outputs" / "AAK_cooling_tunnel"

RAW_XLSX = LAB_INPUT / "lab_trials_raw.xlsx"
SUMMARY_XLSX = LAB_INPUT / "lab_trials_summary.xlsx"
SETUP_XLSX = LAB_INPUT / "lab_trials_experimental setup.xlsx"

AAK_RAW = AAK_INPUT / "AAK_cooling_tunnel_raw.xlsx"
AAK_SUMMARY = AAK_INPUT / "AAK_cooling_tunnel_summary.xlsx"
AAK_SETUP = AAK_INPUT / "AAK_cooling_tunnel_experimental setup.xlsx"

SHEETS = ["at2455_a_test", "at2455_a_test3", "at2455_a_test4"]
REP_IDS = ["2455_a_test", "2455_a_test3", "2455_a_test4"]
CSV_RAW_SOURCES = {
    "at2455_a_test": Path(r"C:\Users\BurkardJohannes\Desktop\Temp Files\AAK\a_test.csv"),
    "at2455_a_test3": Path(r"C:\Users\BurkardJohannes\Desktop\Temp Files\AAK\a_test_3.csv"),
    "at2455_a_test4": Path(r"C:\Users\BurkardJohannes\Desktop\Temp Files\AAK\a_test_4.csv"),
}
TIME_COL = "Absolut[s]"
PRODUCT_TEMPS = ["T1", "T3", "T4", "T5"]
MOULD_TEMP = "T2"
AMBIENT_TEMP = "T6"
PRIMARY_US = ["Rx1Tx1", "Rx2Tx2"]
ALL_US = ["Rx1Tx1", "Rx2Tx2", "Rx1Tx2", "Rx2Tx1"]

REFERENCE_SHEET = "at2455_a_test"
REFERENCE_BOUNDARIES = {
    "cooling_zone_1_end_t6_drop": 990.0,
    "cooling_zone_2_end_t6_change": 1500.0,
    "cooling_zone_3_end_warming": 2195.0,
}
PROCESS_START_TIMES = {
    "2455_a_test": {"deposition_s": 155.0, "cooling_start_s": 210.0},
    "2455_a_test3": {"deposition_s": 106.0, "cooling_start_s": 155.0},
    "2455_a_test4": {"deposition_s": 676.0, "cooling_start_s": 728.0},
}


def norm_id(value: object) -> str:
    s = str(value or "").strip().lower()
    if s.startswith("at"):
        s = s[2:]
    return s.replace(" ", "_")


def write_df(ws, df: pd.DataFrame) -> None:
    if df.empty:
        ws["A1"] = "No rows."
        return
    for col_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(1, col_idx, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            if isinstance(value, (np.floating, float)) and pd.notna(value):
                value = float(value)
            elif isinstance(value, (np.integer, int)) and pd.notna(value):
                value = int(value)
            elif pd.isna(value):
                value = None
            ws.cell(row_idx, col_idx, value)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, col in enumerate(df.columns, 1):
        samples = [str(col), *[str(v) for v in df[col].head(80).fillna("").tolist()]]
        ws.column_dimensions[get_column_letter(idx)].width = min(max(len(s) for s in samples) + 2, 42)


def copy_worksheet_values(src_ws, dst_ws) -> None:
    for r_idx, row in enumerate(src_ws.iter_rows(values_only=True), 1):
        for c_idx, value in enumerate(row, 1):
            dst_ws.cell(r_idx, c_idx, value)


def extract_subgroup_files() -> pd.DataFrame:
    AAK_INPUT.mkdir(parents=True, exist_ok=True)

    raw_dst = Workbook()
    raw_dst.remove(raw_dst.active)
    for sheet in SHEETS:
        dst = raw_dst.create_sheet(sheet)
        df = read_raw(sheet)
        for c_idx, col in enumerate(df.columns, 1):
            dst.cell(1, c_idx, col)
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                if pd.isna(value):
                    value = None
                dst.cell(r_idx, c_idx, value)
    raw_dst.save(AAK_RAW)

    setup = pd.read_excel(SETUP_XLSX)
    setup_sub = setup[setup["Experimental Name"].map(norm_id).isin(REP_IDS)].copy()
    with pd.ExcelWriter(AAK_SETUP, engine="openpyxl") as writer:
        setup_sub.to_excel(writer, index=False, sheet_name="experimental_setup")

    # Copy only the AAK cooling-tunnel summary block as values.
    sum_src = load_workbook(SUMMARY_XLSX, read_only=True, data_only=True).active
    rows = list(sum_src.iter_rows(values_only=True))
    dataset_rows = [idx for idx, row in enumerate(rows) if row and row[0] == "Dataset"]
    dataset_rows.append(len(rows))
    block_start = None
    block_end = None
    for i, start in enumerate(dataset_rows[:-1]):
        if rows[start][1] == "2455_a_test/test3/test4":
            block_start = start
            block_end = dataset_rows[i + 1]
            break
    sum_dst = Workbook()
    ws = sum_dst.active
    ws.title = "summary"
    if block_start is not None:
        for r_idx, row in enumerate(rows[block_start:block_end], 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(r_idx, c_idx, value)
    sum_dst.save(AAK_SUMMARY)

    return pd.DataFrame(
        [
            {"created_file": AAK_RAW.name, "source": "a_test.csv, a_test_3.csv, a_test_4.csv", "content": "Raw sheets at2455_a_test, at2455_a_test3, at2455_a_test4"},
            {"created_file": AAK_SETUP.name, "source": SETUP_XLSX.name, "content": "Experimental setup rows for the cooling-tunnel subgroup"},
            {"created_file": AAK_SUMMARY.name, "source": SUMMARY_XLSX.name, "content": "Summary block 2455_a_test/test3/test4"},
        ]
    )


def load_aak_landmarks() -> pd.DataFrame:
    setup = load_setup()
    ws = load_workbook(SUMMARY_XLSX, read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    dataset_rows = [idx for idx, row in enumerate(rows) if row and row[0] == "Dataset"]
    dataset_rows.append(len(rows))
    block = None
    dataset = None
    for i, start in enumerate(dataset_rows[:-1]):
        if rows[start][1] == "2455_a_test/test3/test4":
            block = rows[start : dataset_rows[i + 1]]
            dataset = rows[start][1]
            break
    if block is None:
        return pd.DataFrame()
    rep_row_idx = next(i for i, row in enumerate(block) if row and row[0] == "Repetitions")
    sensor_row_idx = next(i for i, row in enumerate(block) if row and row[0] == "Sensors")
    rep_row = block[rep_row_idx]
    sensor_row = block[sensor_row_idx]
    col_map = {}
    current_rep = None
    for col_idx in range(1, min(len(sensor_row), 9)):
        if col_idx < len(rep_row) and rep_row[col_idx] is not None:
            current_rep = norm_id(rep_row[col_idx])
        sensor = sensor_row[col_idx] if col_idx < len(sensor_row) else None
        if current_rep and sensor is not None:
            if isinstance(sensor, (int, float)):
                sensor = str(int(sensor))
            else:
                sensor = str(sensor)
            if sensor in {"11", "22"}:
                col_map[col_idx] = (current_rep, sensor)
    metric_rows = []
    wanted = {
        "Viscosity Damping",
        "Cryst. Onset (rel.)",
        "Cryst. Efficiency",
        "Detachment Onset (rel.)",
        "Detachment Completion (rel.)",
        "Detachment Duration",
        "Cryst. Length",
    }
    for row in block[sensor_row_idx + 1 :]:
        label = row[0] if row else None
        if label not in wanted:
            continue
        for col_idx, (rep_id, sensor_code) in col_map.items():
            value = row[col_idx] if col_idx < len(row) else None
            if value is None:
                continue
            metric_rows.append({"dataset": dataset, "rep_id": rep_id, "sensor_code": sensor_code, "metric": label, "value": value})
    long = pd.DataFrame(metric_rows)
    pivot = long.pivot_table(index=["dataset", "rep_id", "sensor_code"], columns="metric", values="value", aggfunc="first").reset_index()
    pivot.columns.name = None
    pivot = pivot.merge(setup, on="rep_id", how="left")
    pivot["raw_sheet"] = "at" + pivot["rep_id"]
    pivot["deposition_s"] = pivot["rep_id"].map(lambda r: PROCESS_START_TIMES.get(r, {}).get("deposition_s", np.nan))
    pivot["cooling_start_s"] = pivot["rep_id"].map(lambda r: PROCESS_START_TIMES.get(r, {}).get("cooling_start_s", np.nan))
    for col in [
        "Cryst. Onset (rel.)",
        "Detachment Onset (rel.)",
        "Detachment Completion (rel.)",
    ]:
        if col in pivot.columns:
            abs_col = col.replace(" (rel.)", " absolute_s")
            pivot[abs_col] = pivot[col] + pivot["cooling_start_s"]
    return pivot


def read_raw(sheet: str) -> pd.DataFrame:
    cols = ["time", *PRODUCT_TEMPS, MOULD_TEMP, AMBIENT_TEMP, "T7", "T8", "T9", *ALL_US]
    csv_path = CSV_RAW_SOURCES.get(sheet)
    if csv_path is not None and csv_path.exists():
        df = pd.read_csv(csv_path, usecols=lambda c: c in cols)
        df = df.rename(columns={"time": TIME_COL})
    else:
        df = pd.read_excel(RAW_XLSX, sheet_name=sheet, usecols=lambda c: c in [TIME_COL, *PRODUCT_TEMPS, MOULD_TEMP, AMBIENT_TEMP, *ALL_US])
    df = df.dropna(subset=[TIME_COL]).copy()
    return df


def value_at(df: pd.DataFrame, col: str, t: float) -> float:
    valid = df[[TIME_COL, col]].dropna()
    idx = (valid[TIME_COL] - t).abs().idxmin()
    return float(valid.loc[idx, col])


def first_crossing_after(df: pd.DataFrame, col: str, threshold: float, after: float, direction: str) -> float | None:
    valid = df[[TIME_COL, col]].dropna().sort_values(TIME_COL)
    valid = valid[valid[TIME_COL] >= after]
    if direction == "below":
        hit = valid[valid[col] <= threshold]
    else:
        hit = valid[valid[col] >= threshold]
    if hit.empty:
        return None
    return float(hit.iloc[0][TIME_COL])


def t6_rise_boundary(df: pd.DataFrame, after: float, rise_C: float = 1.0, search_end: float | None = None) -> float | None:
    valid = df[[TIME_COL, AMBIENT_TEMP]].dropna().sort_values(TIME_COL).copy()
    if search_end is None:
        search_end = float(valid[TIME_COL].max())
    valid = valid[(valid[TIME_COL] >= after) & (valid[TIME_COL] <= search_end)].copy()
    if len(valid) < 8:
        return None
    valid["smooth_t6"] = smooth_series(valid[AMBIENT_TEMP], 21)
    low_idx = valid["smooth_t6"].idxmin()
    low_time = float(valid.loc[low_idx, TIME_COL])
    low_value = float(valid.loc[low_idx, "smooth_t6"])
    after_low = valid[valid[TIME_COL] > low_time].copy()
    hit = after_low[after_low["smooth_t6"] >= low_value + rise_C]
    if hit.empty:
        return None
    return float(hit.iloc[0][TIME_COL])


def smooth_series(s: pd.Series, window: int = 17) -> pd.Series:
    return s.rolling(window, center=True, min_periods=max(3, window // 4)).median()


def detect_zones(raw: dict[str, pd.DataFrame], landmarks: pd.DataFrame) -> pd.DataFrame:
    ref = raw[REFERENCE_SHEET]
    ref_t6_b1 = value_at(ref, AMBIENT_TEMP, REFERENCE_BOUNDARIES["cooling_zone_1_end_t6_drop"])
    ref_t6_b2 = value_at(ref, AMBIENT_TEMP, REFERENCE_BOUNDARIES["cooling_zone_2_end_t6_change"])
    ref_t6_b3 = value_at(ref, AMBIENT_TEMP, REFERENCE_BOUNDARIES["cooling_zone_3_end_warming"])

    rows = []
    for sheet, df in raw.items():
        rep_id = norm_id(sheet)
        start_meta = PROCESS_START_TIMES.get(rep_id, {})
        deposition_s = start_meta.get("deposition_s", np.nan)
        cooling_start_s = start_meta.get("cooling_start_s", 0.0)
        tmax = float(df[TIME_COL].max())
        if sheet == REFERENCE_SHEET:
            b1 = REFERENCE_BOUNDARIES["cooling_zone_1_end_t6_drop"]
            b2_detected = t6_rise_boundary(df, b1, rise_C=1.0, search_end=REFERENCE_BOUNDARIES["cooling_zone_3_end_warming"])
            b2 = b2_detected or REFERENCE_BOUNDARIES["cooling_zone_2_end_t6_change"]
            b3 = min(REFERENCE_BOUNDARIES["cooling_zone_3_end_warming"], tmax)
            basis = "reference pattern; zone 2 to 3 boundary verified by smoothed T6 rise >1 C"
        else:
            b1 = first_crossing_after(df, AMBIENT_TEMP, ref_t6_b1, cooling_start_s + 100.0, "below")
            product_mean = df[PRODUCT_TEMPS].mean(axis=1)
            tmp = df[[TIME_COL, AMBIENT_TEMP]].copy()
            tmp["product_mean"] = product_mean
            tmp["smooth_product"] = smooth_series(tmp["product_mean"], 21)
            tmp["slope_product"] = tmp["smooth_product"].diff() / tmp[TIME_COL].diff()
            b2 = t6_rise_boundary(df, (b1 or (cooling_start_s + 780.0)) + 60.0, rise_C=1.0)
            after_b2 = tmp[tmp[TIME_COL] > ((b2 or 1500.0) + 250.0)].copy()
            warm_hits = after_b2[(after_b2[AMBIENT_TEMP] >= ref_t6_b3) & (after_b2["slope_product"] > 0.001)]
            if warm_hits.empty:
                warm_hits = after_b2.sort_values("slope_product", ascending=False).head(1)
            b3 = float(warm_hits.iloc[0][TIME_COL]) if not warm_hits.empty else min(2195.0, tmax)
            basis = "pattern-derived; zone 2 to 3 boundary detected by smoothed T6 rise >1 C"
            if b1 is None:
                b1 = min(cooling_start_s + 780.0, tmax)
            if b2 is None:
                b2 = min(max(b1 + 400.0, cooling_start_s + 1290.0), tmax)
            b3 = min(max(b3, b2 + 60.0), tmax)

        zones = [
            ("deposition_to_cooling_start", deposition_s, cooling_start_s, "user-provided deposition and cooling-start times"),
            ("cooling_zone_1_t6_drop", cooling_start_s, b1, basis),
            ("cooling_zone_2_t6_change", b1, b2, basis),
            ("cooling_zone_3_until_warming", b2, b3, basis),
            ("cooling_zone_4_demoulding", b3, tmax, "after strong warming/demoulding transition"),
        ]
        for zone, start, end, note in zones:
            rows.append(
                {
                    "rep_id": rep_id,
                    "raw_sheet": sheet,
                    "zone": zone,
                    "start_s": start,
                    "end_s": end,
                    "duration_s": end - start if pd.notna(start) else np.nan,
                    "basis": note,
                    "T6_start_C": value_at(df, AMBIENT_TEMP, start) if pd.notna(start) else np.nan,
                    "T6_end_C": value_at(df, AMBIENT_TEMP, end) if pd.notna(end) else np.nan,
                    "product_mean_start_C": float(df.loc[(df[TIME_COL] - start).abs().idxmin(), PRODUCT_TEMPS].mean()) if pd.notna(start) else np.nan,
                    "product_mean_end_C": float(df.loc[(df[TIME_COL] - end).abs().idxmin(), PRODUCT_TEMPS].mean()) if pd.notna(end) else np.nan,
                }
            )
    return pd.DataFrame(rows)


def product_summaries(raw: dict[str, pd.DataFrame], zones: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    delta_rows = []
    hot_rows = []
    for _, z in zones.iterrows():
        if not str(z["zone"]).startswith("cooling_zone") or pd.isna(z["start_s"]):
            continue
        df = raw[z["raw_sheet"]]
        seg = df[(df[TIME_COL] >= z["start_s"]) & (df[TIME_COL] <= z["end_s"])]
        if seg.empty:
            continue
        sensor_means = seg[PRODUCT_TEMPS].mean()
        sensor_ranges = seg[PRODUCT_TEMPS].max() - seg[PRODUCT_TEMPS].min()
        start_vals = df.loc[(df[TIME_COL] - z["start_s"]).abs().idxmin(), PRODUCT_TEMPS]
        end_vals = df.loc[(df[TIME_COL] - z["end_s"]).abs().idxmin(), PRODUCT_TEMPS]
        for sensor in PRODUCT_TEMPS:
            delta_rows.append(
                {
                    "rep_id": z["rep_id"],
                    "zone": z["zone"],
                    "sensor": sensor,
                    "start_C": float(start_vals[sensor]),
                    "end_C": float(end_vals[sensor]),
                    "delta_C": float(end_vals[sensor] - start_vals[sensor]),
                    "mean_C": float(sensor_means[sensor]),
                    "range_C": float(sensor_ranges[sensor]),
                }
            )
        hot_sensor = str(sensor_means.idxmax())
        cold_sensor = str(sensor_means.idxmin())
        hot_rows.append(
            {
                "rep_id": z["rep_id"],
                "zone": z["zone"],
                "hottest_product_sensor": hot_sensor,
                "hottest_mean_C": float(sensor_means[hot_sensor]),
                "coldest_product_sensor": cold_sensor,
                "coldest_mean_C": float(sensor_means[cold_sensor]),
                "within_mould_spread_C": float(sensor_means.max() - sensor_means.min()),
                "largest_range_sensor": str(sensor_ranges.idxmax()),
                "largest_sensor_range_C": float(sensor_ranges.max()),
            }
        )
    return pd.DataFrame(delta_rows), pd.DataFrame(hot_rows)


def change_scores(t: np.ndarray, y: np.ndarray, window_points: int) -> np.ndarray:
    scores = np.full(len(y), np.nan)
    for i in range(window_points, len(y) - window_points):
        left = y[i - window_points : i]
        right = y[i : i + window_points]
        denom = np.nanstd(np.r_[left, right])
        if not np.isfinite(denom) or denom <= 1e-9:
            denom = 1e-9
        scores[i] = (np.nanmean(right) - np.nanmean(left)) / denom
    return scores


def detect_change_points(seg: pd.DataFrame, channel: str) -> list[dict[str, float]]:
    valid = seg[[TIME_COL, channel]].dropna().sort_values(TIME_COL)
    if len(valid) < 12:
        return []
    t = valid[TIME_COL].to_numpy(float)
    y = valid[channel].rolling(5, center=True, min_periods=2).median().to_numpy(float)
    dt = np.nanmedian(np.diff(t))
    window_points = max(4, int(round(45.0 / max(dt, 1.0))))
    scores = change_scores(t, y, window_points)
    positive_scores = scores[np.isfinite(scores) & (scores > 0)]
    threshold = np.nanpercentile(positive_scores, 80) if len(positive_scores) else np.inf
    candidates = []
    min_gap = 60.0
    for idx in np.argsort(np.nan_to_num(scores, nan=-np.inf))[::-1]:
        if not np.isfinite(scores[idx]) or scores[idx] <= 0 or scores[idx] < max(0.8, threshold):
            continue
        tt = float(t[idx])
        if any(abs(tt - c["time_s"]) < min_gap for c in candidates):
            continue
        before = np.nanmedian(y[max(0, idx - window_points) : idx])
        after = np.nanmedian(y[idx : min(len(y), idx + window_points)])
        local_delta = float(after - before)
        if local_delta < 0.025:
            continue
        candidates.append(
            {
                "time_s": tt,
                "score": float(scores[idx]),
                "before": float(before),
                "after": float(after),
                "delta": local_delta,
                "direction": "positive_upward",
            }
        )
        if len(candidates) >= 30:
            break
    return sorted(candidates, key=lambda c: c["time_s"])


def median_after(seg: pd.DataFrame, channel: str, t0: float, width: float = 60.0) -> float:
    w = seg[(seg[TIME_COL] >= t0) & (seg[TIME_COL] <= t0 + width)][channel].dropna()
    if w.empty:
        return np.nan
    return float(w.median())


def median_window(df: pd.DataFrame, channel: str, start: float, end: float) -> float:
    w = df[(df[TIME_COL] >= start) & (df[TIME_COL] <= end)][channel].dropna()
    if w.empty:
        idx = (df[TIME_COL] - end).abs().idxmin()
        return float(df.loc[idx, channel])
    return float(w.median())


def threshold_from_reference(reference: float) -> float:
    return float(reference - 0.10 * abs(reference))


def threshold_met(value: float, threshold: float) -> bool:
    return bool(pd.notna(value) and pd.notna(threshold) and value >= threshold)


def ultrasound_detachment(raw: dict[str, pd.DataFrame], zones: pd.DataFrame, landmarks: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    rows = []
    cp_rows = []
    lm = landmarks.set_index(["rep_id", "sensor_code"])
    for sheet, df in raw.items():
        rep_id = norm_id(sheet)
        z = zones[(zones["raw_sheet"] == sheet) & (zones["zone"] == "cooling_zone_4_demoulding")]
        demould_start = float(z.iloc[0]["start_s"]) if not z.empty else float(df[TIME_COL].max())
        deposition_s = PROCESS_START_TIMES.get(rep_id, {}).get("deposition_s", np.nan)
        for sensor_code in ["11", "22"]:
            if (rep_id, sensor_code) not in lm.index:
                continue
            row_lm = lm.loc[(rep_id, sensor_code)]
            if "Detachment Onset absolute_s" in row_lm and pd.notna(row_lm["Detachment Onset absolute_s"]):
                onset = float(row_lm["Detachment Onset absolute_s"])
            else:
                onset = float(row_lm["Detachment Onset (rel.)"])
            window_end = min(demould_start, float(df[TIME_COL].max()))
            seg = df[(df[TIME_COL] >= onset) & (df[TIME_COL] <= window_end)].copy()
            if seg.empty:
                continue
            reference_levels = {
                ch: median_window(df, ch, deposition_s - 5.0, deposition_s) if pd.notna(deposition_s) else median_after(seg, ch, onset, 5.0)
                for ch in PRIMARY_US
            }
            thresholds = {ch: threshold_from_reference(reference_levels[ch]) for ch in PRIMARY_US}
            for ch in PRIMARY_US:
                cps = detect_change_points(seg, ch)
                complete_cp = None
                last_cp = cps[-1] if cps else {"time_s": np.nan, "score": np.nan, "before": np.nan, "after": np.nan}
                cp_count_until_decision = 0
                for cp_idx, cp in enumerate(cps, 1):
                    value_at_cp = median_after(seg, ch, cp["time_s"], 5.0)
                    met = threshold_met(value_at_cp, thresholds[ch])
                    if complete_cp is None:
                        cp_count_until_decision = cp_idx
                    cp_rows.append(
                        {
                            "rep_id": rep_id,
                            "sensor_code": sensor_code,
                            "channel": ch,
                            "change_point_index": cp_idx,
                            "change_point_s": cp["time_s"],
                            "change_score": cp["score"],
                            "change_direction": cp.get("direction", "positive_upward"),
                            "local_before_median": cp.get("before", np.nan),
                            "local_after_median": cp.get("after", np.nan),
                            "local_delta_after_minus_before": cp.get("delta", np.nan),
                            "reference_window_s": f"{deposition_s - 5.0:.1f}-{deposition_s:.1f}",
                            "reference_us_level": reference_levels[ch],
                            "threshold_10pct_lower": thresholds[ch],
                            "us_level_at_change_point": value_at_cp,
                            "threshold_met": met,
                            "note": (
                                "first passing change point used as complete detachment"
                                if met and complete_cp is None
                                else ("later passing change point after selected decision" if met else "")
                            ),
                        }
                    )
                    if met and complete_cp is None:
                        complete_cp = cp
                if complete_cp is not None:
                    selected_cp = complete_cp
                    selected_value = median_after(seg, ch, selected_cp["time_s"], 5.0)
                    status = "complete_detachment_for_channel"
                elif cps:
                    selected_cp = last_cp
                    selected_value = median_after(seg, ch, selected_cp["time_s"], 5.0)
                    status = "partial_detachment_no_change_point_met_threshold"
                    cp_count_until_decision = len(cps)
                else:
                    selected_cp = {"time_s": np.nan, "score": np.nan}
                    selected_value = np.nan
                    status = "no_change_point_detected"
                    cp_count_until_decision = 0
                rows.append(
                    {
                        "rep_id": rep_id,
                        "sensor_code": sensor_code,
                        "channel": ch,
                        "deposition_s": deposition_s,
                        "reference_window_s": f"{deposition_s - 5.0:.1f}-{deposition_s:.1f}",
                        "reference_us_level": reference_levels[ch],
                        "threshold_10pct_lower": thresholds[ch],
                        "detachment_onset_s": onset,
                        "analysis_window_end_s": window_end,
                        "window_duration_s": window_end - onset,
                        "change_points_detected_total": len(cps),
                        "change_points_until_decision": cp_count_until_decision,
                        "selected_change_point_s": selected_cp["time_s"],
                        "selected_change_score": selected_cp["score"],
                        "us_level_at_selected_cp": selected_value,
                        "threshold_met_at_selected_cp": threshold_met(selected_value, thresholds[ch]),
                        "detachment_status": status,
                        "method": "positive/upward local mean-shift change points only; complete detachment per channel requires US level at a change point >= the 10%-below-reference threshold",
                    }
                )
    return pd.DataFrame(rows), pd.DataFrame(cp_rows)


def rupture_decision_summary(us_detach: pd.DataFrame) -> pd.DataFrame:
    if us_detach.empty:
        return pd.DataFrame()
    rows = []
    for _, row in us_detach.iterrows():
        rows.append(
            {
                "rep_id": row["rep_id"],
                "sensor_code": row["sensor_code"],
                "channel": row["channel"],
                "deposition_s": row["deposition_s"],
                "reference_window_s": row["reference_window_s"],
                "reference_us_level": row["reference_us_level"],
                "threshold_10pct_lower": row["threshold_10pct_lower"],
                "detachment_onset_s_absolute": row["detachment_onset_s"],
                "selected_change_point_s": row["selected_change_point_s"],
                "us_level_at_selected_cp": row["us_level_at_selected_cp"],
                "threshold_met_at_selected_cp": row["threshold_met_at_selected_cp"],
                "change_points_detected_total": row["change_points_detected_total"],
                "change_points_until_decision": row["change_points_until_decision"],
                "complete_detachment_detected": row["detachment_status"] == "complete_detachment_for_channel",
                "written_decision": (
                    f"{row['rep_id']} sensor {row['sensor_code']} {row['channel']}: pre-deposition reference "
                    f"{row['reference_us_level']:.4f}, threshold {row['threshold_10pct_lower']:.4f}. "
                    f"Selected change point {row['selected_change_point_s']:.1f} s has US {row['us_level_at_selected_cp']:.4f}. "
                    f"Change points until decision: {row['change_points_until_decision']}. "
                    f"Decision: {row['detachment_status']}."
                ),
            }
        )
    return pd.DataFrame(rows)


def plot_series(draw: ImageDraw.ImageDraw, x, y, box, color, width=2, ymin=None, ymax=None):
    x0, y0, x1, y1 = box
    vals = pd.DataFrame({"x": x, "y": y}).dropna()
    if vals.empty:
        return
    xmin, xmax = vals["x"].min(), vals["x"].max()
    ymin = vals["y"].min() if ymin is None else ymin
    ymax = vals["y"].max() if ymax is None else ymax
    if xmax == xmin or ymax == ymin:
        return
    pts = []
    for _, r in vals.iterrows():
        px = x0 + (r["x"] - xmin) / (xmax - xmin) * (x1 - x0)
        py = y1 - (r["y"] - ymin) / (ymax - ymin) * (y1 - y0)
        pts.append((px, py))
    if len(pts) > 1:
        draw.line(pts, fill=color, width=width)


def make_figures(raw: dict[str, pd.DataFrame], zones: pd.DataFrame, us_detach: pd.DataFrame) -> list[Path]:
    fig_dir = AAK_OUTPUT / "figures"
    fig_dir.mkdir(parents=True, exist_ok=True)
    paths = []
    colors = {"T1": "#cc3311", "T3": "#0077bb", "T4": "#009988", "T5": "#ee7733", "T2": "#aa4499", "T6": "#000000", "Rx1Tx1": "#3344cc", "Rx2Tx2": "#cc3377"}
    try:
        font = ImageFont.truetype("arial.ttf", 16)
        small = ImageFont.truetype("arial.ttf", 12)
    except Exception:
        font = ImageFont.load_default()
        small = ImageFont.load_default()
    for sheet, df in raw.items():
        rep_id = norm_id(sheet)
        img = Image.new("RGB", (1400, 820), "white")
        draw = ImageDraw.Draw(img)
        draw.text((40, 18), f"{sheet}: AAK cooling tunnel zones and ultrasound detachment checks", fill="black", font=font)
        temp_box = (80, 70, 1320, 430)
        us_box = (80, 500, 1320, 760)
        for box, label in [(temp_box, "Temperature / T6 ambient"), (us_box, "Primary ultrasound")]:
            draw.rectangle(box, outline="#444444", width=1)
            draw.text((box[0], box[1] - 22), label, fill="black", font=small)
        tmin, tmax = float(df[TIME_COL].min()), float(df[TIME_COL].max())
        temp_vals = df[[*PRODUCT_TEMPS, MOULD_TEMP, AMBIENT_TEMP]].stack().dropna()
        temp_ymin, temp_ymax = float(temp_vals.min()) - 0.5, float(temp_vals.max()) + 0.5
        us_vals = df[PRIMARY_US].stack().dropna()
        us_ymin, us_ymax = float(us_vals.min()) - 0.02, float(us_vals.max()) + 0.02
        # zone shading and labels
        rz = zones[zones["raw_sheet"] == sheet]
        shade_colors = ["#e8f2ff", "#eef8ea", "#fff3d6", "#f9e6e8"]
        cooling_zones = rz[rz["zone"].str.startswith("cooling_zone")]
        for idx, (_, z) in enumerate(cooling_zones.iterrows()):
            x0 = temp_box[0] + (z["start_s"] - tmin) / (tmax - tmin) * (temp_box[2] - temp_box[0])
            x1 = temp_box[0] + (z["end_s"] - tmin) / (tmax - tmin) * (temp_box[2] - temp_box[0])
            draw.rectangle((x0, temp_box[1], x1, temp_box[3]), fill=shade_colors[idx % len(shade_colors)])
            draw.rectangle((x0, us_box[1], x1, us_box[3]), fill=shade_colors[idx % len(shade_colors)])
            draw.text((x0 + 4, temp_box[1] + 4), z["zone"].replace("cooling_", "C "), fill="#333333", font=small)
        for s in [*PRODUCT_TEMPS, MOULD_TEMP, AMBIENT_TEMP]:
            plot_series(draw, df[TIME_COL], df[s], temp_box, colors.get(s, "#666666"), width=2 if s != AMBIENT_TEMP else 3, ymin=temp_ymin, ymax=temp_ymax)
        for ch in PRIMARY_US:
            plot_series(draw, df[TIME_COL], df[ch], us_box, colors[ch], width=3, ymin=us_ymin, ymax=us_ymax)
        # redraw box outlines
        draw.rectangle(temp_box, outline="#444444", width=1)
        draw.rectangle(us_box, outline="#444444", width=1)
        # selected change points
        det = us_detach[us_detach["rep_id"] == rep_id]
        for _, row in det.iterrows():
            if row.get("detachment_status") != "complete_detachment_for_channel":
                continue
            cp = row["selected_change_point_s"]
            if pd.notna(cp):
                x = us_box[0] + (cp - tmin) / (tmax - tmin) * (us_box[2] - us_box[0])
                draw.line((x, us_box[1], x, us_box[3]), fill="#d62728", width=2)
                draw.text((x + 3, us_box[1] + 12), f"full detach {row['sensor_code']} {row['channel']}", fill="#d62728", font=small)
        legend_x, legend_y = 90, 440
        for i, s in enumerate([*PRODUCT_TEMPS, MOULD_TEMP, AMBIENT_TEMP, *PRIMARY_US]):
            draw.line((legend_x + i * 105, legend_y, legend_x + i * 105 + 22, legend_y), fill=colors.get(s, "#666666"), width=3)
            draw.text((legend_x + i * 105 + 26, legend_y - 7), s, fill="black", font=small)
        out = fig_dir / f"{sheet}_zones_us_detachment.png"
        img.save(out)
        paths.append(out)
    return paths


def make_rupture_figures(raw: dict[str, pd.DataFrame], us_detach: pd.DataFrame, cp_details: pd.DataFrame) -> list[Path]:
    fig_dir = AAK_OUTPUT / "figures"
    fig_dir.mkdir(parents=True, exist_ok=True)
    paths = []
    try:
        font = ImageFont.truetype("arial.ttf", 16)
        small = ImageFont.truetype("arial.ttf", 12)
    except Exception:
        font = ImageFont.load_default()
        small = ImageFont.load_default()
    colors = {"Rx1Tx1": "#3344cc", "Rx2Tx2": "#cc3377"}
    for _, row in us_detach.iterrows():
        sheet = "at" + row["rep_id"]
        if sheet not in raw:
            continue
        ch = row["channel"]
        df = raw[sheet]
        start = float(row["detachment_onset_s"])
        end = float(row["analysis_window_end_s"])
        seg = df[(df[TIME_COL] >= start) & (df[TIME_COL] <= end)].copy()
        if seg.empty:
            continue
        img = Image.new("RGB", (1300, 680), "white")
        draw = ImageDraw.Draw(img)
        title = f"{row['rep_id']} sensor {row['sensor_code']} {ch}: rupture/detachment decision"
        draw.text((40, 20), title, fill="black", font=font)
        box = (80, 95, 1220, 500)
        draw.rectangle(box, outline="#444444", width=1)
        us_vals = seg[ch].dropna()
        ymin, ymax = float(us_vals.min()) - 0.02, float(us_vals.max()) + 0.02
        plot_series(draw, seg[TIME_COL], seg[ch], box, colors[ch], width=3, ymin=ymin, ymax=ymax)
        for level, label, line_color in [
            (row["reference_us_level"], "pre-deposition reference", "#666666"),
            (row["threshold_10pct_lower"], "-10% threshold", colors[ch]),
        ]:
            if pd.notna(level):
                ypix = box[3] - (level - ymin) / (ymax - ymin) * (box[3] - box[1])
                draw.line((box[0], ypix, box[2], ypix), fill=line_color, width=1)
                draw.text((box[2] - 260, ypix - 14), f"{ch} {label}", fill=line_color, font=small)
        # onset. Change points are drawn below with clearer pass/fail coding.
        for tt, label, color in [
            (start, "detachment onset", "#222222"),
        ]:
            if pd.notna(tt):
                x = box[0] + (tt - start) / (end - start) * (box[2] - box[0])
                draw.line((x, box[1], x, box[3]), fill=color, width=2)
                draw.text((x + 4, box[1] + 8), label, fill=color, font=small)
        cps = cp_details[
            (cp_details["rep_id"] == row["rep_id"])
            & (cp_details["sensor_code"].astype(str) == str(row["sensor_code"]))
            & (cp_details["channel"] == ch)
        ]
        for _, cp in cps.iterrows():
            tt = cp["change_point_s"]
            if pd.notna(tt):
                x = box[0] + (tt - start) / (end - start) * (box[2] - box[0])
                cp_color = "#2ca02c" if cp["threshold_met"] else "#999999"
                line_top = box[1] if cp["threshold_met"] else box[3] - 70
                draw.line((x, line_top, x, box[3]), fill=cp_color, width=2 if cp["threshold_met"] else 1)
                draw.text((x + 2, box[3] - 18), str(int(cp["change_point_index"])), fill=cp_color, font=small)
        selected_tt = row["selected_change_point_s"]
        is_complete = row["detachment_status"] == "complete_detachment_for_channel"
        if is_complete and pd.notna(selected_tt):
            x = box[0] + (selected_tt - start) / (end - start) * (box[2] - box[0])
            draw.line((x, box[1], x, box[3]), fill="#d62728", width=3)
            draw.text((x + 4, box[1] + 8), "selected first passing CP", fill="#d62728", font=small)
        draw.text((85, 520), f"Reference window: {row['reference_window_s']} s | reference {row['reference_us_level']:.4f} | threshold {row['threshold_10pct_lower']:.4f}", fill="#000000", font=font)
        if is_complete:
            decision_line = f"Selected passing CP US: {row['us_level_at_selected_cp']:.4f} | pass if US >= threshold: {row['threshold_met_at_selected_cp']} | CPs until decision: {row['change_points_until_decision']}"
        else:
            decision_line = f"No passing CP found. Last diagnostic CP US: {row['us_level_at_selected_cp']:.4f} | pass if US >= threshold: {row['threshold_met_at_selected_cp']} | total CPs: {row['change_points_detected_total']}"
        draw.text((85, 550), decision_line, fill=colors[ch], font=font)
        draw.text((85, 590), f"Decision: {row['detachment_status']}", fill="#000000", font=font)
        draw.text((760, 590), "Grey = all CPs; green = CP passes threshold; red = selected only if full detachment", fill="#333333", font=small)
        out = fig_dir / f"{row['rep_id']}_sensor_{row['sensor_code']}_{ch}_rupture_decision.png"
        img.save(out)
        paths.append(out)
    return paths


def build_report() -> Path:
    AAK_OUTPUT.mkdir(parents=True, exist_ok=True)
    fig_dir = AAK_OUTPUT / "figures"
    if fig_dir.exists():
        for png in fig_dir.glob("*.png"):
            png.unlink()
    created = extract_subgroup_files()
    landmarks = load_aak_landmarks()
    raw = {sheet: read_raw(sheet) for sheet in SHEETS}
    zones = detect_zones(raw, landmarks)
    product_delta, hotspots = product_summaries(raw, zones)
    us_detach, us_change_points = ultrasound_detachment(raw, zones, landmarks)
    us_decisions = rupture_decision_summary(us_detach)
    figures = make_figures(raw, zones, us_detach)
    rupture_figures = make_rupture_figures(raw, us_detach, us_change_points)

    readme = pd.DataFrame(
        [
            {"item": "source subgroup", "note": "Only at2455_a_test, at2455_a_test3, and at2455_a_test4 are included."},
            {"item": "user-provided deposition/cooling starts", "note": "a_test: deposition 155 s, cooling start 210 s; a_test3: deposition 106 s, cooling start 155 s; a_test4: deposition 676 s, cooling start 728 s."},
            {"item": "summary time conversion", "note": "Crystallization and detachment landmarks from the summary are relative to cooling start, so cooling_start_s is added before analyzing raw curves."},
            {"item": "zone reference", "note": "at2455_a_test uses user-provided approximate boundaries: 990 s, 1500 s, 2195 s."},
            {"item": "zone 2 to 3 boundary", "note": "The transition from cooling zone 2 to zone 3 is detected/verified by a smoothed T6 increase of more than 1 C after the local T6 low."},
            {"item": "zone transfer", "note": "Other repetitions use T6 ambient levels, the T6 +1 C rise rule, and product-temperature warming to transfer the reference cooling-zone pattern."},
            {"item": "temperature scope", "note": "Product deltas and hotspots use product sensors T1, T3, T4, T5 only. T2 is mould and T6 is ambient."},
            {"item": "ultrasound rupture method", "note": "Change points are detected as positive/upward local mean shifts because the ruptures package is unavailable. Downward or flat-regime mean shifts are ignored. Complete detachment is evaluated separately per channel and requires the US level at a change point to be beyond/above the threshold defined as 10% below the pre-deposition reference."},
            {"item": "ultrasound reference", "note": "The reference US signal is the median signal from approximately 5 s before deposition until deposition."},
            {"item": "window restriction", "note": "Ultrasound detachment analysis starts at detachment onset from the summary, shifted by cooling_start_s, and ends before the detected demoulding/warming zone."},
            {"item": "deposition zone", "note": "Zone 1 is now computed as user-provided deposition time to user-provided cooling-start time."},
        ]
    )
    fig_table = pd.DataFrame([{"figure_type": "zone_overview", "figure": p.name, "path": str(p)} for p in figures] + [{"figure_type": "rupture_decision", "figure": p.name, "path": str(p)} for p in rupture_figures])

    out = AAK_OUTPUT / "AAK_cooling_tunnel_report.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    for name, df in [
        ("Read Me", readme),
        ("Created Subgroup Files", created),
        ("Detected Zones", zones),
        ("Product Delta By Zone", product_delta),
        ("Hotspot Summary", hotspots),
        ("US Detachment Analysis", us_detach),
        ("US Change Points", us_change_points),
        ("US Rupture Decisions", us_decisions),
        ("Figures Index", fig_table),
    ]:
        ws = wb.create_sheet(name[:31])
        write_df(ws, df)
    ws = wb.create_sheet("Zone And US Figures")
    row = 1
    for p in figures:
        ws.cell(row, 1, p.name)
        img = ExcelImage(str(p))
        img.width = 980
        img.height = 574
        ws.add_image(img, f"A{row + 1}")
        row += 34
    ws = wb.create_sheet("Rupture Decision Figures")
    row = 1
    for p in rupture_figures:
        ws.cell(row, 1, p.name)
        img = ExcelImage(str(p))
        img.width = 980
        img.height = 512
        ws.add_image(img, f"A{row + 1}")
        row += 31
    try:
        wb.save(out)
    except PermissionError:
        out = AAK_OUTPUT / "AAK_cooling_tunnel_report_updated.xlsx"
        wb.save(out)
    return out


def main() -> None:
    print(build_report())


if __name__ == "__main__":
    main()
