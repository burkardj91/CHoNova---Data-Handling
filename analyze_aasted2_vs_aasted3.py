from __future__ import annotations

import math
import re
import shutil
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


WORKSPACE = Path(r"C:\Users\BurkardJohannes\Documents\CHoNova _ Data Understanding")
AA2_INPUT = WORKSPACE / "inputs" / "aasted2"
AA3_INPUT = WORKSPACE / "inputs" / "aasted3"
AA2_OUTPUT = WORKSPACE / "outputs" / "aasted2_comparison"
AA2_RUN_NAME = "260604_aasted2_Run 1_1"
AA2_RUN_SAFE = "260604_aasted2_Run_1_1"
AA2_EMPTY_NAME = "260604_aasted2_Run 2_0"
AA3_REFERENCE_NAME = "reference_2291-4160"
RUN_FOLDER = AA2_OUTPUT / "runs" / f"{AA2_RUN_SAFE}_vs_aasted3_reference"
REPORT_WORKBOOK = RUN_FOLDER / "reports" / f"{AA2_RUN_SAFE}_vs_aasted3_reference_report.xlsx"
FIGURE_DIR = RUN_FOLDER / "figures"

AA2_RUN_CSV = AA2_INPUT / "260604_aasted2_Run 1_1.csv"
AA2_EMPTY_CSV = AA2_INPUT / "260604_aasted2_Run 2_0.csv"
AA2_SETUP_XLSX = AA2_INPUT / "aa2_trials_experimental_setup.xlsx"
AA2_SUMMARY_XLSX = AA2_INPUT / "aa2_trials_experimental_summary.xlsx"
AA2_CONFIG_YAML = AA2_INPUT / "aasted2_mould_profile.yaml"
AA2_SKETCH = AA2_INPUT / "Form_Aa2.png"

AA3_REFERENCE_CSV = AA3_INPUT / "reference_2291_4160.csv"
AA3_SETUP_XLSX = AA3_INPUT / "aa3_trials_experimental_setup.xlsx"
AA3_SUMMARY_XLSX = AA3_INPUT / "aa3_trials_experimental_summary.xlsx"
AA3_CONFIG_YAML = AA3_INPUT / "aasted3_mould_profile.yaml"
AA3_CONFIG_IMAGE = AA3_INPUT / "aasted3_config.png"
AA3_RUN_FOLDER = WORKSPACE / "outputs" / "aasted3_run_comparison" / "runs" / "17_3_f_m_aa3_25_06_26_2160_4374s"
AA3_RAW_TABLES = AA3_RUN_FOLDER / "reports" / "aasted3_raw_detail_tables.xlsx"
AA3_REFERENCE_FIGURE = AA3_RUN_FOLDER / "figures" / "zone_overview" / "reference_2291_4160_temperature_accz_zones.png"

TEMP_SENSORS = [f"T{i}" for i in range(1, 10)]
IMU_SENSORS = ["acc x", "acc y", "acc z", "gyro x", "gyro y", "gyro z"]
PRIMARY_US = ["Rx1Tx1", "Rx2Tx2"]

AA2_GROUPS = {
    "product": ["T3", "T4", "T6"],
    "humidity": ["T1"],
    "mould": ["T2", "T9"],
    "ambient": ["T8"],
    "process_markers": ["T5", "T7", "T8"],
}
AA3_GROUPS = {
    "product": ["T2", "T3", "T4", "T5", "T7"],
    "humidity": ["T1"],
    "mould": ["T6"],
    "ambient": ["T8", "T9"],
}

AA2_ZONES = [
    (0.0, 354.0, "movement_to_conditioning", "Mould conditioning before deposition; T1-T9, especially T5/T8/T7, rise."),
    (354.0, 376.0, "deposition_transfer", "Deposition at 354 s; primary US channels drop substantially."),
    (376.0, 590.0, "vibration", "Strong acc z/x/y movement while product and ambient temperatures are nearly stable."),
    (590.0, 650.0, "transition_to_cooling", "Transition from vibration end through the first T8/T5 cooling drop."),
    (650.0, 875.0, "cooling_1", "First cooling section with T8/T5 minima."),
    (875.0, 1152.0, "cooling_2", "Second cooling section with T8/T5 minima."),
    (1152.0, 1481.0, "cooling_3", "Third cooling section with T8/T5 minima."),
    (1481.0, 2088.0, "cooling_4_irregular", "Longer and more irregular cooling section."),
    (2088.0, 2452.0, "cooling_5", "Later cooling section."),
    (2452.0, 2725.0, "cooling_6", "Final cooling section before T5/T8 increase."),
    (2725.0, 2740.0, "cooling_exit_transition", "T5/T8 start increasing after cooling."),
    (2740.0, 2780.0, "post_cooling_mechanical_transfer", "Gyro-y and acc x/y activity; likely transfer or repositioning before final demoulding transition."),
    (2780.0, 2980.0, "transition_2", "Post-cooling transition before clear demoulding."),
    (2980.0, 3000.0, "demoulding_twisting", "Twisting subphase similar to Aasted 3."),
    (3000.0, 3047.0, "demoulding_vibration", "Vibration/shaking visible in accelerometer channels."),
    (3047.0, math.inf, "final_demoulding", "Remaining demoulding/run-out."),
]

AA3_REFERENCE_ZONES = [
    (0.0, 50.0, "movement_to_conditioning"),
    (50.0, 210.0, "mould_conditioning"),
    (210.0, 318.0, "deposition_transfer"),
    (318.0, 472.0, "vibration_warming"),
    (472.0, 524.0, "transition_to_cooling"),
    (524.0, 608.0, "cooling_1"),
    (608.0, 756.0, "cooling_2"),
    (756.0, 898.0, "cooling_3"),
    (898.0, 1032.0, "cooling_4"),
    (1032.0, 1638.0, "less_periodic_cooling_2"),
    (1638.0, 1786.0, "transition_2"),
    (1786.0, 1802.0, "demoulding_twisting"),
    (1802.0, 1850.0, "demoulding_vibration"),
    (1850.0, math.inf, "final_demoulding"),
]


def font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        r"C:\Windows\Fonts\arialbd.ttf" if bold else r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\calibrib.ttf" if bold else r"C:\Windows\Fonts\calibri.ttf",
    ]
    for path in candidates:
        if Path(path).exists():
            return ImageFont.truetype(path, size=size)
    return ImageFont.load_default()


def parse_float(value: object) -> float:
    if pd.isna(value):
        return np.nan
    text = str(value).strip().replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else np.nan


def read_run(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    df["elapsed_s"] = df["time"] - df["time"].min()
    return df


def run_end(df: pd.DataFrame) -> float:
    return float(df["elapsed_s"].max())


def effective_end(end: float, df: pd.DataFrame) -> float:
    return run_end(df) if not np.isfinite(end) else float(end)


def slice_window(df: pd.DataFrame, start: float, end: float) -> pd.DataFrame:
    end = effective_end(end, df)
    return df[(df["elapsed_s"] >= start) & (df["elapsed_s"] < end)].copy()


def rowwise_values(window: pd.DataFrame, sensors: list[str]) -> pd.DataFrame:
    available = [sensor for sensor in sensors if sensor in window.columns]
    if not available:
        return pd.DataFrame({"mean": [], "spread": []})
    vals = window[available].dropna(how="all")
    if vals.empty:
        return pd.DataFrame({"mean": [], "spread": []})
    return pd.DataFrame(
        {
            "mean": vals.mean(axis=1),
            "spread": vals.max(axis=1) - vals.min(axis=1),
        }
    ).dropna(how="all")


def summarize_values(values: pd.Series) -> dict[str, object]:
    values = pd.to_numeric(values, errors="coerce").dropna()
    return {
        "mean_C": float(values.mean()) if not values.empty else np.nan,
        "min_C": float(values.min()) if not values.empty else np.nan,
        "max_C": float(values.max()) if not values.empty else np.nan,
        "std_C": float(values.std()) if len(values) > 1 else np.nan,
        "range_C": float(values.max() - values.min()) if not values.empty else np.nan,
        "n_points": int(values.count()),
    }


def temperature_by_zone(
    df: pd.DataFrame,
    zones: list[tuple[float, float, str, str]] | list[tuple[float, float, str]],
    groups: dict[str, list[str]],
    run: str,
    profile: str,
) -> pd.DataFrame:
    rows = []
    for zone_tuple in zones:
        start, end, zone = zone_tuple[:3]
        window = slice_window(df, start, end)
        end_eff = effective_end(end, df)
        for group_name, sensors in groups.items():
            vals = rowwise_values(window, sensors)
            stats = summarize_values(vals["mean"] if "mean" in vals else pd.Series(dtype=float))
            spread = pd.to_numeric(vals["spread"], errors="coerce").dropna() if "spread" in vals else pd.Series(dtype=float)
            rows.append(
                {
                    "profile": profile,
                    "run": run,
                    "zone": zone,
                    "start_s": start,
                    "end_s": end_eff,
                    "duration_s": end_eff - start,
                    "temperature_group": group_name,
                    "sensors": ", ".join([s for s in sensors if s in df.columns]),
                    **stats,
                    "mean_spatial_spread_C": float(spread.mean()) if not spread.empty else np.nan,
                    "max_spatial_spread_C": float(spread.max()) if not spread.empty else np.nan,
                }
            )
    return pd.DataFrame(rows)


def hotspot_summary(
    df: pd.DataFrame,
    zones: list[tuple[float, float, str, str]] | list[tuple[float, float, str]],
    product_sensors: list[str],
    run: str,
    profile: str,
) -> pd.DataFrame:
    rows = []
    available = [sensor for sensor in product_sensors if sensor in df.columns]
    if not available:
        return pd.DataFrame()
    for zone_tuple in zones:
        start, end, zone = zone_tuple[:3]
        window = slice_window(df, start, end)
        sensor_means = window[available].mean(numeric_only=True).dropna()
        if sensor_means.empty:
            continue
        product_mean = float(sensor_means.mean())
        deltas = sensor_means - product_mean
        hottest = str(sensor_means.idxmax())
        coolest = str(sensor_means.idxmin())
        spread = float(sensor_means.max() - sensor_means.min())
        severity = "high" if spread >= 2.0 else "medium" if spread >= 1.0 else "low"
        color = "#F8696B" if severity == "high" else "#FFEB84" if severity == "medium" else "#63BE7B"
        row = {
            "profile": profile,
            "run": run,
            "zone": zone,
            "start_s": start,
            "end_s": effective_end(end, df),
            "product_sensors": ", ".join(available),
            "product_mean_C": product_mean,
            "hottest_sensor": hottest,
            "hottest_mean_C": float(sensor_means[hottest]),
            "coolest_sensor": coolest,
            "coolest_mean_C": float(sensor_means[coolest]),
            "spread_hottest_minus_coolest_C": spread,
            "spread_severity": severity,
            "spread_color": color,
        }
        for sensor in available:
            row[f"{sensor}_mean_C"] = float(sensor_means[sensor])
            row[f"{sensor}_delta_vs_product_mean_C"] = float(deltas[sensor])
        rows.append(row)
    return pd.DataFrame(rows)


def mechanical_metrics(window: pd.DataFrame) -> dict[str, object]:
    out: dict[str, object] = {"n_imu_points": int(window[IMU_SENSORS].dropna(how="all").shape[0]) if set(IMU_SENSORS).issubset(window.columns) else 0}
    for col in IMU_SENSORS:
        if col not in window.columns:
            out[f"{col}_std"] = np.nan
            out[f"{col}_abs_mean"] = np.nan
            continue
        s = pd.to_numeric(window[col], errors="coerce").dropna()
        out[f"{col}_std"] = float(s.std()) if len(s) > 1 else np.nan
        out[f"{col}_abs_mean"] = float(s.abs().mean()) if not s.empty else np.nan
    if all(col in window.columns for col in ["acc x", "acc y", "acc z"]):
        acc = window[["acc x", "acc y", "acc z"]].dropna(how="any")
        acc_mag = np.sqrt((acc**2).sum(axis=1)) if not acc.empty else pd.Series(dtype=float)
        out["acc_vector_magnitude_mean"] = float(acc_mag.mean()) if len(acc_mag) else np.nan
        out["acc_vector_magnitude_std"] = float(acc_mag.std()) if len(acc_mag) > 1 else np.nan
    if all(col in window.columns for col in ["gyro x", "gyro y", "gyro z"]):
        gyro = window[["gyro x", "gyro y", "gyro z"]].dropna(how="any")
        gyro_mag = np.sqrt((gyro**2).sum(axis=1)) if not gyro.empty else pd.Series(dtype=float)
        out["gyro_vector_magnitude_mean"] = float(gyro_mag.mean()) if len(gyro_mag) else np.nan
        out["gyro_vector_magnitude_std"] = float(gyro_mag.std()) if len(gyro_mag) > 1 else np.nan
    return out


def mechanical_by_zone(
    df: pd.DataFrame,
    zones: list[tuple[float, float, str, str]] | list[tuple[float, float, str]],
    run: str,
    profile: str,
) -> pd.DataFrame:
    rows = []
    for zone_tuple in zones:
        start, end, zone = zone_tuple[:3]
        window = slice_window(df, start, end)
        end_eff = effective_end(end, df)
        rows.append(
            {
                "profile": profile,
                "run": run,
                "zone": zone,
                "start_s": start,
                "end_s": end_eff,
                "duration_s": end_eff - start,
                **mechanical_metrics(window),
            }
        )
    return pd.DataFrame(rows)


def load_aa2_setup() -> pd.DataFrame:
    if not AA2_SETUP_XLSX.exists():
        return pd.DataFrame()
    raw = pd.read_excel(AA2_SETUP_XLSX)
    raw = raw.dropna(how="all").copy()
    return raw.rename(
        columns={
            "Experimental Name": "experimental_name",
            "0 = Chocolate Test, 1 = Empty Run": "chocolate_test_or_empty_flag",
            "Runs within a trial": "runs_within_trial",
            "Duration (sec)": "duration_s",
            "Quality (0 = not good, 1 = good) (when 1 = Empty Run then no quality was assessed)": "quality",
        }
    )


def load_setup_summary(path: Path, profile: str, run_name: str) -> dict[str, object]:
    if not path.exists():
        return {}
    raw = pd.read_excel(path).dropna(how="all")
    if raw.empty:
        return {}
    columns = {str(col).lower(): col for col in raw.columns}
    run_col = next((col for key, col in columns.items() if "experimental" in key and "name" in key), None)
    selected = raw
    if run_col:
        if profile == "aasted3_reference":
            hit = raw[raw[run_col].astype(str).str.contains("2291|reference", case=False, na=False)]
        else:
            hit = raw[raw[run_col].astype(str).eq(run_name)]
        if not hit.empty:
            selected = hit
    row = selected.iloc[0]
    def find(*tokens: str) -> object:
        for col in raw.columns:
            text = str(col).lower()
            if all(token in text for token in tokens):
                return row.get(col, "")
        return ""
    return {
        "setup_run_name": row.get(run_col, run_name) if run_col else run_name,
        "quality_0_not_good_1_good": find("quality"),
        "chocolate": find("chocolate"),
        "tempering": find("tempering"),
        "remarks": find("remarks"),
        "protocol_duration_s": parse_float(find("duration")),
    }


def load_aa2_landmarks() -> pd.DataFrame:
    raw = pd.read_excel(AA2_SUMMARY_XLSX).dropna(how="all")
    rows = []
    for _, row in raw.iterrows():
        run = str(row.get("Experimental Name", AA2_RUN_NAME))
        dep = parse_float(row.get("Deposition (sec)"))
        specs = {
            "Rx1Tx1": {
                "crystallization": "Crystallization Start T1R1 (sec)",
                "onset": "Detachment Onset T1R1 (sec)",
                "offset": "Detachment Offset T1R1 (sec)",
            },
            "Rx2Tx2": {
                "crystallization": "Crystallization Start T2R2 (sec)",
                "onset": "Detachment Onset T2R2 (sec)",
                "offset": "Detachment Offset T2R2 (sec)",
            },
        }
        for channel, cols in specs.items():
            rows.append(
                {
                    "profile": "aasted2",
                    "run": run,
                    "channel": channel,
                    "deposition_s": dep,
                    "crystallization_onset_s": parse_float(row.get(cols["crystallization"])),
                    "detachment_onset_s": parse_float(row.get(cols["onset"])),
                    "detachment_offset_s": parse_float(row.get(cols["offset"])),
                    "detachment_offset_source": "manual_experimental_summary",
                }
            )
    out = pd.DataFrame(rows)
    out["crystallization_onset_minus_deposition_s"] = out["crystallization_onset_s"] - out["deposition_s"]
    out["detachment_onset_minus_deposition_s"] = out["detachment_onset_s"] - out["deposition_s"]
    out["detachment_offset_minus_deposition_s"] = out["detachment_offset_s"] - out["deposition_s"]
    out["detachment_onset_to_offset_s"] = out["detachment_offset_s"] - out["detachment_onset_s"]
    return out


def load_aa3_reference_landmarks() -> pd.DataFrame:
    if AA3_RAW_TABLES.exists():
        try:
            det = pd.read_excel(AA3_RAW_TABLES, sheet_name="Aasted Detachment Summary")
            det = det[det["run"].eq(AA3_REFERENCE_NAME)].copy()
            if not det.empty:
                keep = [
                    "run",
                    "channel",
                    "deposition_s",
                    "crystallization_onset_s",
                    "detachment_onset_s",
                    "detachment_offset_s",
                    "detachment_offset_status",
                    "detachment_onset_to_offset_s",
                    "detachment_offset_minus_deposition_s",
                    "detachment_onset_minus_deposition_s",
                    "crystallization_onset_minus_deposition_s",
                ]
                out = det[[col for col in keep if col in det.columns]].copy()
                out.insert(0, "profile", "aasted3_reference")
                out["detachment_offset_source"] = "T7_corrected_US_change_point_from_aasted3_report"
                return out
        except Exception:
            pass
    if not AA3_SUMMARY_XLSX.exists():
        return pd.DataFrame()
    raw = pd.read_excel(AA3_SUMMARY_XLSX).dropna(how="all")
    raw = raw[raw["Experimental Name"].astype(str).str.contains("2291|reference", case=False, na=False)]
    rows = []
    for _, row in raw.iterrows():
        dep = parse_float(row.get("Deposition (sec)"))
        for channel, cryst_col, onset_col in [
            ("Rx1Tx1", "Crystallization Start T1R1 (sec)", "Detachment Onset T1R1 (sec)"),
            ("Rx2Tx2", "Crystallization Start T2R2 (sec)", "Detachment Onset T2R2 (sec)"),
        ]:
            rows.append(
                {
                    "profile": "aasted3_reference",
                    "run": AA3_REFERENCE_NAME,
                    "channel": channel,
                    "deposition_s": dep,
                    "crystallization_onset_s": parse_float(row.get(cryst_col)),
                    "detachment_onset_s": parse_float(row.get(onset_col)),
                    "detachment_offset_s": np.nan,
                    "detachment_offset_source": "not_available",
                }
            )
    return pd.DataFrame(rows)


def median_window(df: pd.DataFrame, channel: str, start: float, end: float) -> float:
    if channel not in df.columns:
        return np.nan
    w = df[(df["elapsed_s"] >= start) & (df["elapsed_s"] <= end)][channel].dropna()
    if w.empty:
        idx = (df["elapsed_s"] - end).abs().idxmin()
        return float(df.loc[idx, channel]) if pd.notna(df.loc[idx, channel]) else np.nan
    return float(w.median())


def viscosity_proxy(df: pd.DataFrame, landmarks: pd.DataFrame, run: str, profile: str) -> pd.DataFrame:
    rows = []
    dep_values = pd.to_numeric(landmarks["deposition_s"], errors="coerce").dropna()
    if dep_values.empty:
        return pd.DataFrame()
    dep = float(dep_values.iloc[0])
    ratios = []
    for channel in PRIMARY_US:
        pre = median_window(df, channel, dep - 5.0, dep)
        post = median_window(df, channel, dep + 50.0, dep + 55.0)
        ratio = post / pre if pd.notna(pre) and abs(pre) > 1e-12 else np.nan
        ratios.append(ratio)
        rows.append(
            {
                "profile": profile,
                "run": run,
                "channel": channel,
                "signal_basis": "raw ultrasound; no temperature correction applied",
                "deposition_s": dep,
                "pre_window_s": f"{dep - 5.0:.1f}-{dep:.1f}",
                "post_window_s": f"{dep + 50.0:.1f}-{dep + 55.0:.1f}",
                "pre_us_level": pre,
                "post_50s_us_level": post,
                "raw_viscosity_proxy_channel": ratio,
                "raw_viscosity_proxy_mean_rx1rx2": np.nan,
            }
        )
    mean_ratio = float(np.nanmean(ratios)) if ratios else np.nan
    for row in rows:
        row["raw_viscosity_proxy_mean_rx1rx2"] = mean_ratio
    return pd.DataFrame(rows)


def temp_until_detachment(df: pd.DataFrame, landmarks: pd.DataFrame, groups: dict[str, list[str]], profile: str) -> pd.DataFrame:
    rows = []
    for _, landmark in landmarks.iterrows():
        dep = parse_float(landmark.get("deposition_s"))
        onset = parse_float(landmark.get("detachment_onset_s"))
        if pd.isna(dep) or pd.isna(onset):
            continue
        window = slice_window(df, dep, onset)
        for group_name, sensors in groups.items():
            vals = rowwise_values(window, sensors)
            mean_stats = summarize_values(vals["mean"] if "mean" in vals else pd.Series(dtype=float))
            spread = pd.to_numeric(vals["spread"], errors="coerce").dropna() if "spread" in vals else pd.Series(dtype=float)
            rows.append(
                {
                    "profile": profile,
                    "run": landmark["run"],
                    "channel": landmark["channel"],
                    "temperature_group": group_name,
                    "start_deposition_s": dep,
                    "end_detachment_onset_s": onset,
                    "duration_s": onset - dep,
                    "sensors": ", ".join([s for s in sensors if s in df.columns]),
                    **mean_stats,
                    "mean_spatial_spread_C": float(spread.mean()) if not spread.empty else np.nan,
                    "max_spatial_spread_C": float(spread.max()) if not spread.empty else np.nan,
                }
            )
    return pd.DataFrame(rows)


def quality_comparison(
    landmarks: pd.DataFrame,
    viscosity: pd.DataFrame,
    cooling_summary: pd.DataFrame,
    hotspot: pd.DataFrame,
    mechanical: pd.DataFrame,
) -> pd.DataFrame:
    rows = []
    for profile, run, setup_path in [
        ("aasted2", AA2_RUN_NAME, AA2_SETUP_XLSX),
        ("aasted3_reference", AA3_REFERENCE_NAME, AA3_SETUP_XLSX),
    ]:
        setup = load_setup_summary(setup_path, profile, run)
        run_landmarks = landmarks[landmarks["profile"].eq(profile)]
        run_visc = viscosity[viscosity["profile"].eq(profile)]
        product_cooling = cooling_summary[
            (cooling_summary["profile"].eq(profile))
            & (cooling_summary["aggregate_zone"].eq("all_cooling"))
            & (cooling_summary["temperature_group"].eq("product"))
        ]
        run_hotspots = hotspot[hotspot["profile"].eq(profile)] if not hotspot.empty else pd.DataFrame()
        run_mech = mechanical[mechanical["profile"].eq(profile)] if not mechanical.empty else pd.DataFrame()
        vibration = run_mech[run_mech["comparison_domain"].eq("process_vibration")]
        twisting = run_mech[run_mech["comparison_domain"].eq("demoulding_twisting")]
        rows.append(
            {
                "profile": profile,
                "run": run,
                **setup,
                "deposition_s": run_landmarks["deposition_s"].dropna().iloc[0] if run_landmarks["deposition_s"].notna().any() else np.nan,
                "crystallization_onset_minus_deposition_mean_s": run_landmarks["crystallization_onset_minus_deposition_s"].mean()
                if "crystallization_onset_minus_deposition_s" in run_landmarks
                else np.nan,
                "detachment_onset_minus_deposition_mean_s": run_landmarks["detachment_onset_minus_deposition_s"].mean()
                if "detachment_onset_minus_deposition_s" in run_landmarks
                else np.nan,
                "detachment_offset_minus_deposition_mean_s": run_landmarks["detachment_offset_minus_deposition_s"].mean()
                if "detachment_offset_minus_deposition_s" in run_landmarks
                else np.nan,
                "detachment_onset_to_offset_mean_s": run_landmarks["detachment_onset_to_offset_s"].mean()
                if "detachment_onset_to_offset_s" in run_landmarks
                else np.nan,
                "viscosity_proxy_mean_rx1rx2": run_visc["raw_viscosity_proxy_mean_rx1rx2"].dropna().iloc[0]
                if not run_visc.empty and run_visc["raw_viscosity_proxy_mean_rx1rx2"].notna().any()
                else np.nan,
                "all_cooling_product_mean_C": product_cooling["mean_C"].iloc[0] if not product_cooling.empty else np.nan,
                "all_cooling_product_spread_C": product_cooling["mean_spatial_spread_C"].iloc[0] if not product_cooling.empty else np.nan,
                "max_hotspot_spread_C": run_hotspots["spread_hottest_minus_coolest_C"].max() if not run_hotspots.empty else np.nan,
                "vibration_gyro_magnitude_mean": vibration["gyro_vector_magnitude_mean"].iloc[0] if not vibration.empty else np.nan,
                "twisting_gyro_magnitude_mean": twisting["gyro_vector_magnitude_mean"].iloc[0] if not twisting.empty else np.nan,
                "quality_note": (
                    "quality missing or not assessed"
                    if str(setup.get("quality_0_not_good_1_good", "")).strip().lower() in ["", "nan", "no information"]
                    else "quality value present"
                ),
            }
        )
    return pd.DataFrame(rows)


def aggregate_windows(
    df: pd.DataFrame,
    zone_groups: dict[str, list[str]],
    zone_table: list[tuple[float, float, str, str]] | list[tuple[float, float, str]],
    temp_groups: dict[str, list[str]],
    profile: str,
    run: str,
) -> pd.DataFrame:
    by_zone = {zone_tuple[2]: zone_tuple for zone_tuple in zone_table}
    rows = []
    for aggregate, zones in zone_groups.items():
        frames = []
        spans = []
        for zone in zones:
            if zone not in by_zone:
                continue
            start, end = by_zone[zone][:2]
            spans.append(f"{zone}:{start:.0f}-{effective_end(end, df):.0f}s")
            frames.append(slice_window(df, start, end))
        if not frames:
            continue
        window = pd.concat(frames, ignore_index=True)
        for group_name, sensors in temp_groups.items():
            vals = rowwise_values(window, sensors)
            stats = summarize_values(vals["mean"] if "mean" in vals else pd.Series(dtype=float))
            spread = pd.to_numeric(vals["spread"], errors="coerce").dropna() if "spread" in vals else pd.Series(dtype=float)
            rows.append(
                {
                    "profile": profile,
                    "run": run,
                    "aggregate_zone": aggregate,
                    "included_zones": "; ".join(spans),
                    "temperature_group": group_name,
                    "sensors": ", ".join([s for s in sensors if s in df.columns]),
                    **stats,
                    "mean_spatial_spread_C": float(spread.mean()) if not spread.empty else np.nan,
                    "max_spatial_spread_C": float(spread.max()) if not spread.empty else np.nan,
                }
            )
    return pd.DataFrame(rows)


def cooling_comparison(aa2_df: pd.DataFrame, aa3_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    aa2_groups = {
        "early_cooling": ["cooling_1", "cooling_2", "cooling_3"],
        "irregular_late_cooling": ["cooling_4_irregular", "cooling_5", "cooling_6"],
        "all_cooling": ["cooling_1", "cooling_2", "cooling_3", "cooling_4_irregular", "cooling_5", "cooling_6"],
    }
    aa3_groups = {
        "early_cooling": ["cooling_1", "cooling_2", "cooling_3", "cooling_4"],
        "irregular_late_cooling": ["less_periodic_cooling_2"],
        "all_cooling": ["cooling_1", "cooling_2", "cooling_3", "cooling_4", "less_periodic_cooling_2"],
    }
    summary = pd.concat(
        [
            aggregate_windows(aa2_df, aa2_groups, AA2_ZONES, AA2_GROUPS, "aasted2", AA2_RUN_NAME),
            aggregate_windows(aa3_df, aa3_groups, AA3_REFERENCE_ZONES, AA3_GROUPS, "aasted3_reference", AA3_REFERENCE_NAME),
        ],
        ignore_index=True,
    )
    rows = []
    for (aggregate, temp_group), group in summary.groupby(["aggregate_zone", "temperature_group"], dropna=False):
        aa2 = group[group["profile"].eq("aasted2")]
        aa3 = group[group["profile"].eq("aasted3_reference")]
        if aa2.empty or aa3.empty:
            continue
        rows.append(
            {
                "aggregate_zone": aggregate,
                "temperature_group": temp_group,
                "aa2_mean_C": aa2["mean_C"].iloc[0],
                "aa3_reference_mean_C": aa3["mean_C"].iloc[0],
                "delta_aa2_minus_aa3_C": aa2["mean_C"].iloc[0] - aa3["mean_C"].iloc[0],
                "aa2_mean_spatial_spread_C": aa2["mean_spatial_spread_C"].iloc[0],
                "aa3_reference_mean_spatial_spread_C": aa3["mean_spatial_spread_C"].iloc[0],
                "delta_spread_aa2_minus_aa3_C": aa2["mean_spatial_spread_C"].iloc[0] - aa3["mean_spatial_spread_C"].iloc[0],
                "aa2_sensors": aa2["sensors"].iloc[0],
                "aa3_reference_sensors": aa3["sensors"].iloc[0],
            }
        )
    return summary, pd.DataFrame(rows)


def selected_mechanical_comparison(aa2_df: pd.DataFrame, aa3_df: pd.DataFrame) -> pd.DataFrame:
    intervals = [
        ("aasted2", AA2_RUN_NAME, aa2_df, "process_vibration", "vibration", 376.0, 590.0),
        ("aasted3_reference", AA3_REFERENCE_NAME, aa3_df, "process_vibration", "vibration_warming", 318.0, 472.0),
        ("aasted2", AA2_RUN_NAME, aa2_df, "post_cooling_mechanical_transfer", "post_cooling_mechanical_transfer", 2740.0, 2780.0),
        ("aasted2", AA2_RUN_NAME, aa2_df, "demoulding_twisting", "demoulding_twisting", 2980.0, 3000.0),
        ("aasted3_reference", AA3_REFERENCE_NAME, aa3_df, "demoulding_twisting", "demoulding_twisting", 1786.0, 1802.0),
        ("aasted2", AA2_RUN_NAME, aa2_df, "demoulding_vibration", "demoulding_vibration", 3000.0, 3047.0),
        ("aasted3_reference", AA3_REFERENCE_NAME, aa3_df, "demoulding_vibration", "demoulding_vibration", 1802.0, 1850.0),
    ]
    rows = []
    for profile, run, df, comparison_domain, source_zone, start, end in intervals:
        window = slice_window(df, start, end)
        row = {
            "profile": profile,
            "run": run,
            "comparison_domain": comparison_domain,
            "source_zone": source_zone,
            "start_s": start,
            "end_s": end,
            "duration_s": end - start,
            **mechanical_metrics(window),
        }
        if comparison_domain == "post_cooling_mechanical_transfer":
            row["interpretation"] = "Likely transfer/repositioning or pre-demoulding handling because gyro-y and acc x/y activity occurs before the clear twisting and vibration demoulding zones."
        else:
            row["interpretation"] = ""
        rows.append(row)
    return pd.DataFrame(rows)


def mechanical_delta_summary(mechanical: pd.DataFrame) -> pd.DataFrame:
    if mechanical.empty:
        return pd.DataFrame()
    metrics = [
        ("acc_vector_magnitude_std", "acc vibration variability"),
        ("gyro_vector_magnitude_mean", "mean rotational intensity"),
        ("gyro_vector_magnitude_std", "rotational variability"),
        ("gyro y_std", "gyro-y variability"),
        ("acc z_std", "acc-z variability"),
    ]
    rows = []
    for domain in ["process_vibration", "demoulding_twisting", "demoulding_vibration"]:
        aa2 = mechanical[(mechanical["profile"] == "aasted2") & (mechanical["comparison_domain"] == domain)]
        aa3 = mechanical[(mechanical["profile"] == "aasted3_reference") & (mechanical["comparison_domain"] == domain)]
        if aa2.empty or aa3.empty:
            continue
        for metric, label in metrics:
            aa2_val = pd.to_numeric(aa2[metric], errors="coerce").dropna().iloc[0] if metric in aa2 and aa2[metric].notna().any() else np.nan
            aa3_val = pd.to_numeric(aa3[metric], errors="coerce").dropna().iloc[0] if metric in aa3 and aa3[metric].notna().any() else np.nan
            delta = aa2_val - aa3_val if pd.notna(aa2_val) and pd.notna(aa3_val) else np.nan
            ratio = aa2_val / aa3_val if pd.notna(aa2_val) and pd.notna(aa3_val) and abs(aa3_val) > 1e-12 else np.nan
            if pd.isna(delta):
                reading = ""
            elif ratio >= 1.25:
                reading = "AA2 clearly stronger"
            elif ratio <= 0.75:
                reading = "AA2 clearly weaker"
            else:
                reading = "similar magnitude"
            rows.append(
                {
                    "comparison_domain": domain,
                    "metric": metric,
                    "plain_language_metric": label,
                    "aa2_value": aa2_val,
                    "aa3_reference_value": aa3_val,
                    "delta_aa2_minus_aa3": delta,
                    "ratio_aa2_over_aa3": ratio,
                    "reading": reading,
                }
            )
    transfer = mechanical[
        (mechanical["profile"] == "aasted2") & (mechanical["comparison_domain"] == "post_cooling_mechanical_transfer")
    ]
    if not transfer.empty:
        row = transfer.iloc[0]
        rows.append(
            {
                "comparison_domain": "post_cooling_mechanical_transfer",
                "metric": "process_interpretation",
                "plain_language_metric": "AA2-only mechanical event",
                "aa2_value": np.nan,
                "aa3_reference_value": np.nan,
                "delta_aa2_minus_aa3": np.nan,
                "ratio_aa2_over_aa3": np.nan,
                "reading": row.get("interpretation", ""),
            }
        )
    return pd.DataFrame(rows)


def spread_reading(delta: float) -> str:
    if pd.isna(delta):
        return ""
    if delta >= 1.0:
        return "AA2 more heterogeneous"
    if delta <= -1.0:
        return "AA2 more homogeneous"
    return "similar spread"


def temperature_spread_comparison(cooling_delta: pd.DataFrame, temp_until: pd.DataFrame) -> pd.DataFrame:
    rows = []
    if not cooling_delta.empty:
        for _, row in cooling_delta.iterrows():
            delta = row.get("delta_spread_aa2_minus_aa3_C", np.nan)
            rows.append(
                {
                    "comparison_domain": row.get("aggregate_zone"),
                    "channel": "aggregate",
                    "temperature_group": row.get("temperature_group"),
                    "aa2_mean_spatial_spread_C": row.get("aa2_mean_spatial_spread_C"),
                    "aa3_reference_mean_spatial_spread_C": row.get("aa3_reference_mean_spatial_spread_C"),
                    "delta_spread_aa2_minus_aa3_C": delta,
                    "aa2_mean_C": row.get("aa2_mean_C"),
                    "aa3_reference_mean_C": row.get("aa3_reference_mean_C"),
                    "delta_mean_aa2_minus_aa3_C": row.get("delta_aa2_minus_aa3_C"),
                    "reading": spread_reading(delta),
                }
            )
    if not temp_until.empty:
        for (channel, temp_group), group in temp_until.groupby(["channel", "temperature_group"], dropna=False):
            aa2 = group[group["profile"].eq("aasted2")]
            aa3 = group[group["profile"].eq("aasted3_reference")]
            if aa2.empty or aa3.empty:
                continue
            delta = aa2["mean_spatial_spread_C"].iloc[0] - aa3["mean_spatial_spread_C"].iloc[0]
            rows.append(
                {
                    "comparison_domain": "deposition_to_detachment_onset",
                    "channel": channel,
                    "temperature_group": temp_group,
                    "aa2_mean_spatial_spread_C": aa2["mean_spatial_spread_C"].iloc[0],
                    "aa3_reference_mean_spatial_spread_C": aa3["mean_spatial_spread_C"].iloc[0],
                    "delta_spread_aa2_minus_aa3_C": delta,
                    "aa2_mean_C": aa2["mean_C"].iloc[0],
                    "aa3_reference_mean_C": aa3["mean_C"].iloc[0],
                    "delta_mean_aa2_minus_aa3_C": aa2["mean_C"].iloc[0] - aa3["mean_C"].iloc[0],
                    "reading": spread_reading(delta),
                }
            )
    return pd.DataFrame(rows)


def zone_definitions() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "profile": "aasted2",
                "run": AA2_RUN_NAME,
                "zone": zone,
                "start_s": start,
                "end_s": "run_end" if not np.isfinite(end) else end,
                "interpretation": desc,
            }
            for start, end, zone, desc in AA2_ZONES
        ]
    )


def write_df(ws, df: pd.DataFrame) -> None:
    if df.empty:
        ws.cell(1, 1, "No data")
        return
    for col_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(1, col_idx, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            if isinstance(val, (float, np.floating)) and not pd.isna(val):
                val = float(val)
            elif isinstance(val, (int, np.integer)) and not pd.isna(val):
                val = int(val)
            elif pd.isna(val):
                val = None
            ws.cell(row_idx, col_idx, val)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, col in enumerate(df.columns, 1):
        samples = [str(col), *[str(v) for v in df[col].head(120).fillna("").tolist()]]
        ws.column_dimensions[get_column_letter(idx)].width = min(max(len(s) for s in samples) + 2, 42)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                if isinstance(cell.value, float):
                    cell.number_format = "0.000"
        if ws.title == "Hotspot Summary" and ws.max_row > 1:
            headers = {cell.value: cell.column for cell in ws[1]}
            color_col = headers.get("spread_color")
            severity_col = headers.get("spread_severity")
            spread_col = headers.get("spread_hottest_minus_coolest_C")
            for row_idx in range(2, ws.max_row + 1):
                color = str(ws.cell(row_idx, color_col).value or "").replace("#", "") if color_col else ""
                if len(color) == 6:
                    for col_idx in [severity_col, color_col, spread_col]:
                        if col_idx:
                            ws.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor=color)
        if ws.title in {"Mechanical Delta", "Spread Comparison"} and ws.max_row > 1:
            headers = {cell.value: cell.column for cell in ws[1]}
            reading_col = headers.get("reading")
            if reading_col:
                for row_idx in range(2, ws.max_row + 1):
                    reading = str(ws.cell(row_idx, reading_col).value or "").lower()
                    color = ""
                    if "stronger" in reading or "heterogeneous" in reading:
                        color = "F8696B"
                    elif "weaker" in reading or "homogeneous" in reading:
                        color = "63BE7B"
                    elif "similar" in reading:
                        color = "FFEB84"
                    if color:
                        ws.cell(row_idx, reading_col).fill = PatternFill("solid", fgColor=color)
    wb.save(path)


def scale_points(df: pd.DataFrame, col: str, x0: int, y0: int, w: int, h: int, xmax: float, ymin: float, ymax: float) -> list[tuple[int, int]]:
    series = df[["elapsed_s", col]].dropna()
    if series.empty or not np.isfinite(ymin) or not np.isfinite(ymax) or ymin == ymax:
        return []
    if len(series) > 1600:
        series = series.iloc[:: int(np.ceil(len(series) / 1600))]
    xs = x0 + (series["elapsed_s"].to_numpy() / xmax) * w
    ys = y0 + h - ((series[col].to_numpy() - ymin) / (ymax - ymin)) * h
    return [(int(x), int(y)) for x, y in zip(xs, ys) if np.isfinite(x) and np.isfinite(y)]


def robust_normalized(series: pd.Series) -> pd.Series:
    s = pd.to_numeric(series, errors="coerce")
    clean = s.dropna()
    if clean.empty:
        return s * np.nan
    center = clean.median()
    lo = clean.quantile(0.02)
    hi = clean.quantile(0.98)
    scale = max(abs(hi - center), abs(center - lo), 1e-9)
    return ((s - center) / scale).clip(-1.15, 1.15)


def draw_polyline(draw: ImageDraw.ImageDraw, pts: list[tuple[int, int]], color: str, width: int = 2) -> None:
    if len(pts) >= 2:
        draw.line(pts, fill=color, width=width, joint="curve")


def make_aa2_zone_figure(df: pd.DataFrame, landmarks: pd.DataFrame) -> Path:
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)
    out = FIGURE_DIR / f"{AA2_RUN_SAFE}_zone_overview.png"
    width, height = 1650, 1000
    margin_l, margin_r = 92, 42
    top, gap = 90, 58
    panel_h = 235
    panel_w = width - margin_l - margin_r
    y_temp = top
    y_imu = y_temp + panel_h + gap
    y_us = y_imu + panel_h + gap
    xmax = run_end(df)
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    title_font = font(24, True)
    label_font = font(15, True)
    small_font = font(12)
    draw.text((margin_l, 28), f"{AA2_RUN_NAME}: zone overview", fill="#111827", font=title_font)
    zone_colors = ["#EFF6FF", "#F0FDF4", "#FFF7ED", "#FEF2F2", "#F8FAFC", "#ECFEFF", "#F5F3FF"]
    for idx, (start, end, zone, _) in enumerate(AA2_ZONES):
        end_eff = effective_end(end, df)
        x_start = margin_l + int((start / xmax) * panel_w)
        x_end = margin_l + int((end_eff / xmax) * panel_w)
        color = zone_colors[idx % len(zone_colors)]
        for y0 in [y_temp, y_imu, y_us]:
            draw.rectangle([x_start, y0, x_end, y0 + panel_h], fill=color)
        label = zone.replace("_", " ")
        if x_end - x_start > 50:
            draw.text((x_start + 3, y_temp - 19), label[:24], fill="#334155", font=small_font)
    for y0, title in [(y_temp, "Temperature + RH T1"), (y_imu, "IMU robust-normalized"), (y_us, "Ultrasound raw (a.u.)")]:
        draw.rectangle([margin_l, y0, margin_l + panel_w, y0 + panel_h], outline="#334155", width=1)
        draw.text((12, y0 + 8), title, fill="#111827", font=label_font)
    temp_vals = df[TEMP_SENSORS].stack().dropna()
    tmin = float(temp_vals.quantile(0.01)) if not temp_vals.empty else 0
    tmax = float(temp_vals.quantile(0.99)) if not temp_vals.empty else 1
    temp_colors = {
        "T1": "#7F1D1D",
        "T2": "#DC2626",
        "T3": "#2563EB",
        "T4": "#16A34A",
        "T5": "#F97316",
        "T6": "#7C3AED",
        "T7": "#0891B2",
        "T8": "#111827",
        "T9": "#A16207",
    }
    for sensor in TEMP_SENSORS:
        draw_polyline(draw, scale_points(df, sensor, margin_l, y_temp, panel_w, panel_h, xmax, tmin, tmax), temp_colors[sensor], 2)
    draw.line([margin_l, y_imu + panel_h // 2, margin_l + panel_w, y_imu + panel_h // 2], fill="#CBD5E1", width=1)
    imu_plot = df[["elapsed_s"]].copy()
    for sensor, color in [("acc x", "#2563EB"), ("acc y", "#16A34A"), ("acc z", "#111827"), ("gyro y", "#DC2626")]:
        plot_col = f"{sensor}_robust"
        imu_plot[plot_col] = robust_normalized(df[sensor])
        draw_polyline(draw, scale_points(imu_plot, plot_col, margin_l, y_imu, panel_w, panel_h, xmax, -1.2, 1.2), color, 2)
    us_vals = df[PRIMARY_US].stack().dropna()
    umin = float(us_vals.quantile(0.01)) if not us_vals.empty else -1
    umax = float(us_vals.quantile(0.99)) if not us_vals.empty else 1
    for sensor, color in [("Rx1Tx1", "#2563EB"), ("Rx2Tx2", "#D0268A")]:
        draw_polyline(draw, scale_points(df, sensor, margin_l, y_us, panel_w, panel_h, xmax, umin, umax), color, 3)
    x_axis_y = y_us + panel_h + 28
    draw.line([margin_l, x_axis_y, margin_l + panel_w, x_axis_y], fill="#111827", width=1)
    for tick in np.linspace(0, math.ceil(xmax / 500) * 500, 8):
        x = margin_l + int((tick / xmax) * panel_w)
        draw.line([x, x_axis_y - 5, x, x_axis_y + 5], fill="#111827")
        draw.text((x - 18, x_axis_y + 8), f"{tick:.0f}", fill="#111827", font=small_font)
    draw.text((margin_l + panel_w // 2 - 40, x_axis_y + 32), "elapsed time (s)", fill="#111827", font=label_font)
    verticals = []
    dep = pd.to_numeric(landmarks["deposition_s"], errors="coerce").dropna()
    if not dep.empty:
        verticals.append((float(dep.iloc[0]), "deposition", "#111827"))
    for _, row in landmarks.iterrows():
        verticals.append((row["crystallization_onset_s"], f"cryst {row['channel']}", "#7C3AED"))
        verticals.append((row["detachment_onset_s"], f"onset {row['channel']}", "#F97316"))
        verticals.append((row["detachment_offset_s"], f"offset {row['channel']}", "#16A34A"))
    for value, label, color in verticals:
        if pd.isna(value):
            continue
        x = margin_l + int((float(value) / xmax) * panel_w)
        draw.line([x, y_temp, x, y_us + panel_h], fill=color, width=2)
        draw.text((x + 3, y_us + panel_h - 18), label[:18], fill=color, font=small_font)
    legend_x = width - 340
    legend_y = 30
    legend_items = [("T sensors + RH T1", "#2563EB"), ("robust IMU", "#111827"), ("Rx1Tx1/Rx2Tx2", "#D0268A"), ("manual landmarks", "#16A34A")]
    for i, (label, color) in enumerate(legend_items):
        y = legend_y + i * 18
        draw.line([legend_x, y + 7, legend_x + 26, y + 7], fill=color, width=3)
        draw.text((legend_x + 32, y), label, fill="#111827", font=small_font)
    img.save(out)
    return out


def copy_inputs_to_run_folder() -> None:
    dst = RUN_FOLDER / "inputs"
    dst.mkdir(parents=True, exist_ok=True)
    for path in [
        AA2_RUN_CSV,
        AA2_EMPTY_CSV,
        AA2_SETUP_XLSX,
        AA2_SUMMARY_XLSX,
        AA2_CONFIG_YAML,
        AA2_SKETCH,
        AA2_INPUT / "aasted2_update_request_checklist.md",
    ]:
        if path.exists():
            shutil.copy2(path, dst / path.name)
    aa3_dst = dst / "aasted3_reference_inputs"
    aa3_dst.mkdir(parents=True, exist_ok=True)
    for path in [AA3_REFERENCE_CSV, AA3_SETUP_XLSX, AA3_SUMMARY_XLSX, AA3_CONFIG_YAML, AA3_CONFIG_IMAGE]:
        if path.exists():
            shutil.copy2(path, aa3_dst / path.name)


def input_inventory() -> pd.DataFrame:
    rows = []
    for role, path in [
        ("aa2_chocolate_raw_csv", AA2_RUN_CSV),
        ("aa2_empty_raw_csv", AA2_EMPTY_CSV),
        ("aa2_experimental_setup", AA2_SETUP_XLSX),
        ("aa2_experimental_summary", AA2_SUMMARY_XLSX),
        ("aa2_config_yaml", AA2_CONFIG_YAML),
        ("aa2_mould_sketch", AA2_SKETCH),
        ("aa3_reference_raw_csv", AA3_REFERENCE_CSV),
        ("aa3_experimental_setup", AA3_SETUP_XLSX),
        ("aa3_experimental_summary", AA3_SUMMARY_XLSX),
        ("aa3_config_yaml", AA3_CONFIG_YAML),
        ("aa3_config_image", AA3_CONFIG_IMAGE),
    ]:
        rows.append(
            {
                "role": role,
                "path": str(path.relative_to(WORKSPACE)) if path.exists() else str(path),
                "exists": path.exists(),
                "bytes": path.stat().st_size if path.exists() else np.nan,
            }
        )
    return pd.DataFrame(rows)


def build_report() -> None:
    RUN_FOLDER.mkdir(parents=True, exist_ok=True)
    (RUN_FOLDER / "reports").mkdir(parents=True, exist_ok=True)
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)
    copy_inputs_to_run_folder()

    aa2 = read_run(AA2_RUN_CSV)
    aa3 = read_run(AA3_REFERENCE_CSV)
    aa2_landmarks = load_aa2_landmarks()
    aa3_landmarks = load_aa3_reference_landmarks()
    all_landmarks = pd.concat([aa2_landmarks, aa3_landmarks], ignore_index=True)

    aa2_temp_zone = temperature_by_zone(aa2, AA2_ZONES, AA2_GROUPS, AA2_RUN_NAME, "aasted2")
    aa2_mech_zone = mechanical_by_zone(aa2, AA2_ZONES, AA2_RUN_NAME, "aasted2")
    aa3_temp_zone = temperature_by_zone(aa3, AA3_REFERENCE_ZONES, AA3_GROUPS, AA3_REFERENCE_NAME, "aasted3_reference")
    aa3_mech_zone = mechanical_by_zone(aa3, AA3_REFERENCE_ZONES, AA3_REFERENCE_NAME, "aasted3_reference")
    cooling_summary, cooling_delta = cooling_comparison(aa2, aa3)
    mech_compare = selected_mechanical_comparison(aa2, aa3)
    mech_delta = mechanical_delta_summary(mech_compare)
    viscosity = pd.concat(
        [
            viscosity_proxy(aa2, aa2_landmarks, AA2_RUN_NAME, "aasted2"),
            viscosity_proxy(aa3, aa3_landmarks, AA3_REFERENCE_NAME, "aasted3_reference"),
        ],
        ignore_index=True,
    )
    temp_until = pd.concat(
        [
            temp_until_detachment(aa2, aa2_landmarks, AA2_GROUPS, "aasted2"),
            temp_until_detachment(aa3, aa3_landmarks, AA3_GROUPS, "aasted3_reference"),
        ],
        ignore_index=True,
    )
    spread_compare = temperature_spread_comparison(cooling_delta, temp_until)
    hotspots = pd.concat(
        [
            hotspot_summary(aa2, AA2_ZONES, AA2_GROUPS["product"], AA2_RUN_NAME, "aasted2"),
            hotspot_summary(aa3, AA3_REFERENCE_ZONES, AA3_GROUPS["product"], AA3_REFERENCE_NAME, "aasted3_reference"),
        ],
        ignore_index=True,
    )
    quality = quality_comparison(all_landmarks, viscosity, cooling_summary, hotspots, mech_compare)
    setup = load_aa2_setup()
    figure_path = make_aa2_zone_figure(aa2, aa2_landmarks)

    readme = pd.DataFrame(
        [
            ["analysis_scope", "First Aasted-2 profile and comparison against the Aasted-3 reference run."],
            ["aa2_zone_source", "Aasted-2 zones use the user-provided manual timings for 260604_aasted2_Run 1_1."],
            ["aa2_temperature_groups", "T1 is humidity, not product. AA2 product = T3/T4/T6; humidity is shown separately as T1; mould = T2/T9; ambient = T8."],
            ["detachment_offset_source", "Aasted-2 detachment offsets are manual values from aa2_trials_experimental_summary.xlsx. No AA2 T-corrected US change-point detection is applied yet."],
            ["raw_viscosity_proxy", "Median raw US 50-55 s after deposition divided by median raw US in the 5 s before deposition; raw method used because AA2 temperature correction is not available yet."],
            ["config_location", "Sensor placement and mapping are stored in inputs/aasted2/aasted2_mould_profile.yaml and inputs/aasted3/aasted3_mould_profile.yaml, not repeated as Excel output sheets."],
            ["post_cooling_mechanical_transfer", "The 2740-2780 s AA2 segment is probably transfer/repositioning or pre-demoulding handling because it precedes clear twisting and vibration demoulding."],
            ["cooling_aggregate_mapping", "AA3 cooling_1-4 are compared to AA2 cooling_1-3; AA3 less_periodic_cooling_2 is compared to AA2 cooling_4_irregular/cooling_5/cooling_6; all_cooling combines those sections."],
            ["imu_figure_scaling", "The overview figure uses robust per-signal IMU normalization for visualization only; mechanical comparison tables retain raw IMU values."],
        ],
        columns=["topic", "explanation"],
    )

    parameter_comparison = all_landmarks.copy()
    for col in ["crystallization_onset_minus_deposition_s", "detachment_onset_minus_deposition_s", "detachment_offset_minus_deposition_s", "detachment_onset_to_offset_s"]:
        if col not in parameter_comparison.columns:
            parameter_comparison[col] = np.nan

    wb = Workbook()
    wb.remove(wb.active)
    sheets = {
        "Read Me": readme,
        "Input Inventory": input_inventory(),
        "AA2 Experimental Setup": setup,
        "AA2 Zone Definitions": zone_definitions(),
        "Manual Landmarks": all_landmarks,
        "Parameter Comparison": parameter_comparison,
        "AA2 Temp By Zone": aa2_temp_zone,
        "AA2 Mechanical By Zone": aa2_mech_zone,
        "AA3 Ref Temp By Zone": aa3_temp_zone,
        "AA3 Ref Mechanical": aa3_mech_zone,
        "Cooling Aggregates": cooling_summary,
        "Cooling Comparison": cooling_delta,
        "Spread Comparison": spread_compare,
        "Hotspot Summary": hotspots,
        "Mechanical Comparison": mech_compare,
        "Mechanical Delta": mech_delta,
        "Viscosity Proxy": viscosity,
        "Temp Until Detach Onset": temp_until,
        "Quality Comparison": quality,
    }
    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(sheet_name[:31])
        write_df(ws, df)

    ws = wb.create_sheet("Zone Figures")
    ws["A1"] = "Aasted-2 zone overview"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    img = XLImage(str(figure_path))
    img.width = 990
    img.height = 600
    ws.add_image(img, "A3")
    if AA3_REFERENCE_FIGURE.exists():
        ws["A38"] = "Aasted-3 reference zone overview"
        ws["A38"].font = Font(bold=True, size=14, color="1F4E78")
        ref_img = XLImage(str(AA3_REFERENCE_FIGURE))
        ref_img.width = 990
        ref_img.height = 600
        ws.add_image(ref_img, "A40")

    wb.save(REPORT_WORKBOOK)
    style_workbook(REPORT_WORKBOOK)

    readme_path = RUN_FOLDER / "README.txt"
    readme_path.write_text(
        "\n".join(
            [
                f"Aasted-2 consolidated output for {AA2_RUN_NAME} compared with {AA3_REFERENCE_NAME}.",
                "",
                "reports/",
                f"  {REPORT_WORKBOOK.name}",
                "figures/",
                f"  {figure_path.name}",
                "inputs/",
                "  AA2 raw CSVs, setup/summary, profile YAML, mould sketch, checklist, and AA3 reference inputs.",
                "",
                "Note: AA2 detachment offset is manual from the experimental summary until temperature-corrected ultrasound coefficients are available.",
            ]
        ),
        encoding="utf-8",
    )
    print(REPORT_WORKBOOK)


if __name__ == "__main__":
    build_report()
