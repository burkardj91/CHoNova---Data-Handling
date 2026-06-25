from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


WORKSPACE = Path(r"C:\Users\BurkardJohannes\Documents\CHoNova _ Data Understanding")
OUTPUT_DIR = WORKSPACE / "outputs" / "aasted3_run_comparison"
COMPARISON_CSV = WORKSPACE / "inputs" / "aasted3" / "old_configuration.csv"
IRREGULAR_CSV = COMPARISON_CSV  # Backward-compatible import name for helper scripts.
REFERENCE_CSV = WORKSPACE / "inputs" / "aasted3" / "reference_2291_4160.csv"
REFERENCE_RUN_NAME = "reference_2291-4160"
COMPARISON_RUN_NAME = "old-configuration"

# Product temperature sensors. User correction listed "T2, T4, T4, T5, T7";
# treat the duplicated T4 as T3 to keep five unique product locations.
PRODUCT_SENSORS = ["T2", "T3", "T4", "T5", "T7"]
TEMP_SENSORS = [f"T{i}" for i in range(1, 10)]
FEATURES = ["T8", "accz_mean", "accz_std", "gyroy_mean", "gyroy_std"]
IMU_SENSOR_PREFIXES = {
    "acc x": "accx",
    "acc y": "accy",
    "acc z": "accz",
    "gyro x": "gyrox",
    "gyro y": "gyroy",
    "gyro z": "gyroz",
    "mag x": "magx",
    "mag y": "magy",
    "mag z": "magz",
}
COORDS = {
    "T1": (53.0, 23.0),
    "T2": (70.0, 13.0),
    "T3": (5.0, 5.0),
    "T4": (5.0, 31.0),
    "T5": (108.0, 5.0),
    "T6": (76.0, 24.0),
    "T7": (99.0, 30.0),
    "T8": (93.0, 23.0),
    "T9": (55.0, 10.0),
}

ZONES = [
    (0, 50, "movement_to_conditioning", "Mould movement to conditioning zone; all product sensors warm up."),
    (50, 210, "mould_conditioning", "Conditioning zone; acc z pattern changes before/after and product temperatures drop."),
    (210, 318, "deposition_transfer", "Deposition around 256 s and transfer toward vibration/warming section."),
    (318, 472, "vibration_warming", "Vibration in non-cooled/warmed zone; T8 ambient signal and product temperatures plateau high."),
    (472, 524, "transition_to_cooling", "Transition after vibration before the first T8 cooling minimum near ventilator."),
    (524, 608, "cooling_1", "First cooling section from T8 minimum to next section."),
    (608, 756, "cooling_2", "Second periodic cooling section."),
    (756, 898, "cooling_3", "Third periodic cooling section."),
    (898, 1032, "cooling_4", "Fourth periodic cooling section."),
    (1032, 1638, "less_periodic_cooling_2", "Less periodic cooling section."),
    (1638, 1786, "transition_2", "Transition period before demoulding."),
    (1786, 99999, "demoulding", "Demoulding zone / run end."),
]


def read_run(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    df["elapsed_s"] = df["time"] - df["time"].min()
    return df


def zone_for_elapsed(elapsed: float) -> str:
    for start, end, name, _ in ZONES:
        if start <= elapsed < end:
            return name
    return ZONES[-1][2]


def build_seconds(df: pd.DataFrame) -> pd.DataFrame:
    max_sec = int(np.ceil(df["elapsed_s"].max()))
    seconds = pd.DataFrame({"elapsed_sec": np.arange(0, max_sec + 1)})

    available_imu = [col for col in IMU_SENSOR_PREFIXES if col in df.columns]
    motion = df.dropna(subset=["acc z", "gyro y"], how="all").copy()
    motion["elapsed_sec"] = np.floor(motion["elapsed_s"]).astype(int)
    agg_spec = {"sample_count": ("time", "count")}
    for col in available_imu:
        prefix = IMU_SENSOR_PREFIXES[col]
        agg_spec[f"{prefix}_mean"] = (col, "mean")
        agg_spec[f"{prefix}_std"] = (col, "std")
    motion_sec = motion.groupby("elapsed_sec").agg(**agg_spec).reset_index()
    seconds = seconds.merge(motion_sec, on="elapsed_sec", how="left")

    temp = df.dropna(subset=["T8"], how="all")[["elapsed_s", "T8"]].sort_values("elapsed_s")
    seconds["T8"] = np.interp(seconds["elapsed_sec"], temp["elapsed_s"], temp["T8"])
    for col in [c for c in seconds.columns if c.endswith("_mean")]:
        seconds[col] = seconds[col].interpolate(limit_direction="both")
    for col in [c for c in seconds.columns if c.endswith("_std")]:
        seconds[col] = seconds[col].fillna(seconds[col].median())
    seconds["zone_reference"] = seconds["elapsed_sec"].map(zone_for_elapsed)
    seconds["T8_smooth"] = seconds["T8"].rolling(15, center=True, min_periods=1).median()
    seconds["gyroy_std_roll"] = seconds["gyroy_std"].rolling(12, center=True, min_periods=3).median()
    seconds["accz_std_roll"] = seconds["accz_std"].rolling(12, center=True, min_periods=3).median()
    return seconds


def _segments_from_mask(seconds: pd.Series, mask: pd.Series, min_len: int = 5) -> list[tuple[int, int, int]]:
    segs = []
    start = None
    prev = None
    for sec, hit in zip(seconds.astype(int), mask.fillna(False).astype(bool)):
        if hit and start is None:
            start = int(sec)
        if start is not None and (not hit or (prev is not None and int(sec) != prev + 1)):
            end = int(prev)
            if end - start + 1 >= min_len:
                segs.append((start, end, end - start + 1))
            start = int(sec) if hit else None
        prev = int(sec)
    if start is not None and prev is not None and int(prev) - start + 1 >= min_len:
        segs.append((start, int(prev), int(prev) - start + 1))
    return segs


def _t8_valleys(seconds: pd.DataFrame, start_after: int, min_distance: int = 80) -> list[tuple[int, float]]:
    s = seconds[seconds["elapsed_sec"] >= start_after].copy()
    vals = s["T8_smooth"].to_numpy()
    secs = s["elapsed_sec"].to_numpy()
    raw = []
    for i in range(8, len(vals) - 8):
        window = vals[i - 8 : i + 9]
        if vals[i] <= np.nanmin(window) + 1e-9 and vals[i] < vals[i - 8] and vals[i] < vals[i + 8]:
            raw.append((int(secs[i]), float(vals[i])))
    grouped = []
    for sec, val in raw:
        if not grouped or sec - grouped[-1][-1][0] > 3:
            grouped.append([(sec, val)])
        else:
            grouped[-1].append((sec, val))
    candidates = []
    for group in grouped:
        candidates.append(min(group, key=lambda item: item[1]))
    selected = []
    for sec, val in candidates:
        if not selected or sec - selected[-1][0] >= min_distance:
            selected.append((sec, val))
        elif val < selected[-1][1]:
            selected[-1] = (sec, val)
    return selected


def detect_pattern_zones(seconds: pd.DataFrame, run_name: str) -> pd.DataFrame:
    high = seconds["gyroy_std_roll"] > 1.0
    high_segments = _segments_from_mask(seconds["elapsed_sec"], high, min_len=5)
    vib_candidates = [seg for seg in high_segments if 250 <= seg[0] <= 650]
    vibration = max(vib_candidates, key=lambda item: item[2]) if vib_candidates else (320, 450, 130)
    dep_candidates = [seg for seg in high_segments if 150 <= seg[0] < vibration[0]]
    deposition_event = dep_candidates[-1] if dep_candidates else (max(0, vibration[0] - 75), vibration[0] - 50, 25)

    before_vib = seconds[(seconds["elapsed_sec"] >= 40) & (seconds["elapsed_sec"] < vibration[0])]
    high_t8_idx = before_vib["T8_smooth"].idxmax() if not before_vib.empty else seconds.index[0]
    high_t8_time = int(seconds.loc[high_t8_idx, "elapsed_sec"])
    deposition_transfer_start = min(deposition_event[0] - 30, high_t8_time + 20)
    deposition_transfer_start = max(60, int(deposition_transfer_start))

    early = seconds[(seconds["elapsed_sec"] >= 10) & (seconds["elapsed_sec"] < deposition_transfer_start)]
    baseline = float(seconds.loc[seconds["elapsed_sec"] <= 10, "T8_smooth"].median())
    warm_cross = early[early["T8_smooth"] >= baseline + 1.0]
    early_imu = early[early["gyroy_std_roll"] > 0.5]
    movement_end_options = []
    if not warm_cross.empty:
        movement_end_options.append(int(warm_cross["elapsed_sec"].iloc[0]))
    if not early_imu.empty:
        movement_end_options.append(int(early_imu["elapsed_sec"].iloc[0]))
    movement_end = min(movement_end_options) if movement_end_options else max(20, deposition_transfer_start - 160)

    valleys = _t8_valleys(seconds, start_after=vibration[1], min_distance=80)
    # First five valleys define transition-to-cooling plus four periodic cooling sections.
    while len(valleys) < 5:
        last = valleys[-1][0] if valleys else vibration[1] + 60
        valleys.append((last + 120, np.nan))
    cooling_start, c2, c3, c4, less_periodic_start = [v[0] for v in valleys[:5]]

    # Use the last genuinely cold T8 valley before the final rise. Small late wiggles
    # around 13-14 C are transition behavior, not the start of transition_2.
    low_valley_limit = float(seconds["T8_smooth"].quantile(0.20) + 1.0)
    late_valleys = [v for v in valleys if less_periodic_start < v[0] and v[1] <= low_valley_limit]
    transition2_start = late_valleys[-1][0] if late_valleys else less_periodic_start

    # Demoulding starts at the sharp acc-z regime change after the last T8-minimum
    # transition, not merely when T8 begins rising.
    accz_med = seconds["accz_mean"].rolling(30, center=True, min_periods=5).median()
    stable = seconds[
        (seconds["elapsed_sec"] >= transition2_start)
        & (seconds["elapsed_sec"] <= min(transition2_start + 70, int(seconds["elapsed_sec"].max())))
    ]
    baseline = float(stable["accz_mean"].median()) if not stable.empty else float(seconds["accz_mean"].median())
    search = seconds[seconds["elapsed_sec"] >= transition2_start + 20].copy()
    search["accz_med"] = accz_med.loc[search.index]
    search["accz_shift"] = (search["accz_med"] - baseline).abs()
    candidates = search[
        (search["accz_shift"] > 0.35)
        | ((search["accz_mean"] - baseline).abs() > 0.55)
        | (search["accz_std_roll"] > 0.20)
    ]
    if not candidates.empty:
        demoulding_start = int(candidates["elapsed_sec"].iloc[0])
    else:
        late_rise = seconds[(seconds["elapsed_sec"] > transition2_start) & (seconds["T8_smooth"] >= 14.0)]
        demoulding_start = int(late_rise["elapsed_sec"].iloc[0]) if not late_rise.empty else int(seconds["elapsed_sec"].max() - 90)
    run_end = int(seconds["elapsed_sec"].max())

    boundaries = [
        (0, movement_end, "movement_to_conditioning", "first warm-up/IMU-change landmark", "medium" if movement_end_options else "low"),
        (movement_end, deposition_transfer_start, "mould_conditioning", "T8 high/conditioning phase before deposition event", "medium"),
        (deposition_transfer_start, vibration[0], "deposition_transfer", "pre-vibration deposition/transfer IMU event", "medium"),
        (vibration[0], vibration[1], "vibration_warming", "longest sustained high gyro-y variability segment", "high"),
        (vibration[1], cooling_start, "transition_to_cooling", "post-vibration transition until first T8 valley", "high"),
        (cooling_start, c2, "cooling_1", "T8 valley-to-valley cooling cycle", "high"),
        (c2, c3, "cooling_2", "T8 valley-to-valley cooling cycle", "high"),
        (c3, c4, "cooling_3", "T8 valley-to-valley cooling cycle", "high"),
        (c4, less_periodic_start, "cooling_4", "T8 valley-to-valley cooling cycle", "high"),
        (less_periodic_start, transition2_start, "less_periodic_cooling_2", "after periodic valleys until the last cold T8 minimum", "medium"),
        (transition2_start, demoulding_start, "transition_2", "last cold T8 minimum until sharp acc-z regime change", "medium"),
        (demoulding_start, run_end, "demoulding", "after sharp acc-z regime change", "medium"),
    ]

    rows = []
    previous_end = 0
    for start, end, zone, basis, confidence in boundaries:
        start = max(int(start), previous_end)
        end = max(start, min(int(end), run_end))
        rows.append(
            {
                "run": run_name,
                "zone": zone,
                "detected_start_s": start,
                "detected_end_s": end,
                "duration_s": end - start,
                "detection_basis": basis,
                "confidence": confidence,
            }
        )
        previous_end = end
    return pd.DataFrame(rows)


def dtw_map(irregular: pd.DataFrame, reference: pd.DataFrame) -> tuple[pd.DataFrame, float]:
    x = irregular[FEATURES].to_numpy(dtype=float)
    y = reference[FEATURES].to_numpy(dtype=float)
    combined = np.vstack([x, y])
    mean = combined.mean(axis=0)
    std = combined.std(axis=0)
    std[std == 0] = 1
    x = (x - mean) / std
    y = (y - mean) / std

    n, m = len(x), len(y)
    prev = np.full(m + 1, np.inf)
    curr = np.full(m + 1, np.inf)
    prev[0] = 0.0
    direction = np.zeros((n + 1, m + 1), dtype=np.uint8)
    for i in range(1, n + 1):
        curr[0] = np.inf
        xi = x[i - 1]
        for j in range(1, m + 1):
            cost = float(np.linalg.norm(xi - y[j - 1]))
            choices = (prev[j - 1], prev[j], curr[j - 1])
            best = int(np.argmin(choices))
            curr[j] = cost + choices[best]
            direction[i, j] = best
        prev, curr = curr, prev

    i, j = n, m
    pairs = []
    while i > 0 and j > 0:
        pairs.append((i - 1, j - 1))
        step = direction[i, j]
        if step == 0:
            i -= 1
            j -= 1
        elif step == 1:
            i -= 1
        else:
            j -= 1
    pairs.reverse()

    pair_df = pd.DataFrame(pairs, columns=["irregular_idx", "reference_idx"])
    map_df = (
        pair_df.groupby("irregular_idx")["reference_idx"]
        .median()
        .round()
        .astype(int)
        .reset_index()
    )
    map_df["irregular_elapsed_s"] = map_df["irregular_idx"]
    map_df["reference_elapsed_s"] = map_df["reference_idx"]
    map_df["reference_zone"] = map_df["reference_elapsed_s"].map(zone_for_elapsed)
    return map_df, float(prev[m])


def temp_with_zones(df: pd.DataFrame, run_name: str, map_df: pd.DataFrame | None = None) -> pd.DataFrame:
    temp = df.dropna(subset=TEMP_SENSORS, how="all").copy()
    temp["run"] = run_name
    temp["elapsed_sec"] = np.floor(temp["elapsed_s"]).astype(int)
    if map_df is None:
        temp["reference_elapsed_s"] = temp["elapsed_s"]
        temp["zone"] = temp["elapsed_s"].map(zone_for_elapsed)
    else:
        mapper = map_df.set_index("irregular_elapsed_s")
        temp["reference_elapsed_s"] = temp["elapsed_sec"].map(mapper["reference_elapsed_s"])
        temp["zone"] = temp["elapsed_sec"].map(mapper["reference_zone"])
    temp["product_mean_C"] = temp[PRODUCT_SENSORS].mean(axis=1)
    temp["product_spread_C"] = temp[PRODUCT_SENSORS].max(axis=1) - temp[PRODUCT_SENSORS].min(axis=1)
    temp["hottest_product_sensor"] = temp[PRODUCT_SENSORS].idxmax(axis=1)
    return temp


def temp_with_detected_zones(df: pd.DataFrame, run_name: str, detected: pd.DataFrame) -> pd.DataFrame:
    temp = df.dropna(subset=TEMP_SENSORS, how="all").copy()
    temp["run"] = run_name
    temp["zone"] = None
    for row in detected[detected["run"] == run_name].itertuples(index=False):
        mask = (temp["elapsed_s"] >= row.detected_start_s) & (temp["elapsed_s"] < row.detected_end_s)
        temp.loc[mask, "zone"] = row.zone
    temp["zone"] = temp["zone"].fillna(detected[detected["run"] == run_name]["zone"].iloc[-1])
    temp["product_mean_C"] = temp[PRODUCT_SENSORS].mean(axis=1)
    temp["product_spread_C"] = temp[PRODUCT_SENSORS].max(axis=1) - temp[PRODUCT_SENSORS].min(axis=1)
    temp["hottest_product_sensor"] = temp[PRODUCT_SENSORS].idxmax(axis=1)
    return temp


def product_zone_summary(temp: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for (run, zone), group in temp.groupby(["run", "zone"], sort=False):
        for sensor in PRODUCT_SENSORS:
            s = group[sensor].dropna()
            rows.append(
                {
                    "run": run,
                    "zone": zone,
                    "sensor": sensor,
                    "mean_C": s.mean(),
                    "min_C": s.min(),
                    "max_C": s.max(),
                    "std_C": s.std(),
                    "n_temp_points": len(s),
                }
            )
        rows.append(
            {
                "run": run,
                "zone": zone,
                "sensor": "PRODUCT_SPREAD",
                "mean_C": group["product_spread_C"].mean(),
                "min_C": group["product_spread_C"].min(),
                "max_C": group["product_spread_C"].max(),
                "std_C": group["product_spread_C"].std(),
                "n_temp_points": len(group),
            }
        )
    return pd.DataFrame(rows)


def product_zone_delta(summary: pd.DataFrame) -> pd.DataFrame:
    pivot = summary.pivot_table(index=["zone", "sensor"], columns="run", values="mean_C").reset_index()
    pivot.columns.name = None
    if COMPARISON_RUN_NAME in pivot.columns and REFERENCE_RUN_NAME in pivot.columns:
        pivot["mean_delta_comparison_minus_reference_C"] = (
            pivot[COMPARISON_RUN_NAME] - pivot[REFERENCE_RUN_NAME]
        )
    return pivot


def duration_by_zone(reference_seconds: pd.DataFrame, map_df: pd.DataFrame) -> pd.DataFrame:
    ref_counts = reference_seconds.groupby("zone_reference").size().rename("reference_duration_s")
    comp_counts = map_df.groupby("reference_zone").size().rename("comparison_pattern_duration_s")
    out = pd.concat([ref_counts, comp_counts], axis=1).fillna(0).reset_index().rename(columns={"index": "zone"})
    out["extra_duration_s"] = out["comparison_pattern_duration_s"] - out["reference_duration_s"]
    order = {name: i for i, (_, _, name, _) in enumerate(ZONES)}
    out["order"] = out["zone"].map(order)
    return out.sort_values("order").drop(columns=["order"])


def mechanical_zone_summary(seconds: pd.DataFrame, run: str, zone_col: str) -> pd.DataFrame:
    rows = []
    for zone, group in seconds.groupby(zone_col, sort=False):
        row = {
            "run": run,
            "zone": zone,
            "T8_mean_C": group["T8"].mean(),
            "T8_min_C": group["T8"].min(),
            "T8_max_C": group["T8"].max(),
        }
        for prefix in IMU_SENSOR_PREFIXES.values():
            mean_col = f"{prefix}_mean"
            std_col = f"{prefix}_std"
            if mean_col in group.columns:
                row[f"{prefix}_mean"] = group[mean_col].mean()
                row[f"{prefix}_min"] = group[mean_col].min()
                row[f"{prefix}_max"] = group[mean_col].max()
            if std_col in group.columns:
                row[f"{prefix}_std_median"] = group[std_col].median()
                row[f"{prefix}_std_max"] = group[std_col].max()
        rows.append(row)
    return pd.DataFrame(rows)


def mechanical_delta(reference_seconds: pd.DataFrame, irregular_seconds: pd.DataFrame, map_df: pd.DataFrame) -> pd.DataFrame:
    comp = irregular_seconds.copy()
    mapper = map_df.set_index("irregular_elapsed_s")
    comp["reference_zone"] = comp["elapsed_sec"].map(mapper["reference_zone"])
    summary = pd.concat(
        [
            mechanical_zone_summary(reference_seconds, REFERENCE_RUN_NAME, "zone_reference"),
            mechanical_zone_summary(comp, COMPARISON_RUN_NAME, "reference_zone"),
        ],
        ignore_index=True,
    )
    metrics = [
        col
        for col in summary.columns
        if col not in {"run", "zone"} and pd.api.types.is_numeric_dtype(summary[col])
    ]
    pivot = summary.pivot_table(index="zone", columns="run", values=metrics)
    rows = []
    for zone in pivot.index:
        row = {"zone": zone}
        notable = []
        for metric in metrics:
            ref = pivot.loc[zone, (metric, REFERENCE_RUN_NAME)] if (metric, REFERENCE_RUN_NAME) in pivot else np.nan
            comp_val = pivot.loc[zone, (metric, COMPARISON_RUN_NAME)] if (metric, COMPARISON_RUN_NAME) in pivot else np.nan
            row[f"reference_{metric}"] = ref
            row[f"comparison_{metric}"] = comp_val
            row[f"delta_{metric}"] = comp_val - ref
            if pd.notna(ref) and pd.notna(comp_val):
                if metric.startswith("T8") and abs(comp_val - ref) >= 1.0:
                    notable.append(metric)
                if metric.startswith(("accz", "gyroy")) and abs(comp_val - ref) >= max(0.05, abs(ref) * 0.35):
                    notable.append(metric)
        row["notable_differences"] = ", ".join(notable)
        rows.append(row)
    order = {name: i for i, (_, _, name, _) in enumerate(ZONES)}
    return pd.DataFrame(rows).assign(order=lambda d: d["zone"].map(order)).sort_values("order").drop(columns=["order"])


def hotspot_by_zone(temp: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for (run, zone), group in temp.groupby(["run", "zone"], sort=False):
        product_median = group[PRODUCT_SENSORS].median(axis=1)
        for sensor in PRODUCT_SENSORS:
            delta = group[sensor] - product_median
            rows.append(
                {
                    "run": run,
                    "zone": zone,
                    "sensor": sensor,
                    "x_mm": COORDS[sensor][0],
                    "y_mm": COORDS[sensor][1],
                    "mean_vs_product_median_C": delta.mean(),
                    "pct_time_1C_above_product_median": (delta > 1.0).mean(),
                    "pct_time_1C_below_product_median": (delta < -1.0).mean(),
                }
            )
    return pd.DataFrame(rows)


def clearer_hotspots(temp: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    rows = []
    matrix_rows = []
    for (run, zone), group in temp.groupby(["run", "zone"], sort=False):
        means = group[PRODUCT_SENSORS].mean()
        zone_mean = means.mean()
        hottest = means.idxmax()
        coolest = means.idxmin()
        rows.append(
            {
                "run": run,
                "zone": zone,
                "zone_product_mean_C": zone_mean,
                "hottest_sensor": hottest,
                "hottest_mean_C": means[hottest],
                "coolest_sensor": coolest,
                "coolest_mean_C": means[coolest],
                "spread_hottest_minus_coolest_C": means[hottest] - means[coolest],
                "plain_language": f"{hottest} is warmest and {coolest} is coolest in this zone.",
            }
        )
        for sensor in PRODUCT_SENSORS:
            rows[-1][f"{sensor}_mean_C"] = means[sensor]
            rows[-1][f"{sensor}_vs_zone_mean_C"] = means[sensor] - zone_mean
            matrix_rows.append(
                {
                    "run": run,
                    "zone": zone,
                    "sensor": sensor,
                    "x_mm": COORDS[sensor][0],
                    "y_mm": COORDS[sensor][1],
                    "mean_C": means[sensor],
                    "delta_vs_zone_product_mean_C": means[sensor] - zone_mean,
                }
            )
    hotspot_summary = pd.DataFrame(rows)
    hotspot_long = pd.DataFrame(matrix_rows)
    delta_matrix = hotspot_long.pivot_table(
        index=["run", "zone"], columns="sensor", values="delta_vs_zone_product_mean_C"
    ).reset_index()
    delta_matrix.columns.name = None
    return hotspot_summary, hotspot_long, delta_matrix


def zone_findings(duration: pd.DataFrame, mech_delta: pd.DataFrame, product_delta: pd.DataFrame) -> pd.DataFrame:
    rows = []
    descriptions = {name: desc for _, _, name, desc in ZONES}
    duration_lookup = duration.set_index("zone")
    mech_lookup = mech_delta.set_index("zone")
    for _, _, zone, _ in ZONES:
        if zone not in duration_lookup.index:
            continue
        prod_zone = product_delta[(product_delta["zone"] == zone) & (product_delta["sensor"].isin(PRODUCT_SENSORS))]
        spread_zone = product_delta[(product_delta["zone"] == zone) & (product_delta["sensor"] == "PRODUCT_SPREAD")]
        mean_prod_delta = prod_zone["mean_delta_comparison_minus_reference_C"].mean()
        max_abs_sensor = prod_zone.loc[
            prod_zone["mean_delta_comparison_minus_reference_C"].abs().idxmax(), "sensor"
        ] if not prod_zone.empty else ""
        max_abs_delta = prod_zone["mean_delta_comparison_minus_reference_C"].abs().max() if not prod_zone.empty else np.nan
        spread_delta = (
            spread_zone["mean_delta_comparison_minus_reference_C"].iloc[0]
            if not spread_zone.empty
            else np.nan
        )
        notable = mech_lookup.loc[zone, "notable_differences"] if zone in mech_lookup.index else ""
        extra = duration_lookup.loc[zone, "extra_duration_s"]
        rows.append(
            {
                "zone": zone,
                "process_interpretation": descriptions[zone],
                "reference_duration_s": duration_lookup.loc[zone, "reference_duration_s"],
                "comparison_pattern_duration_s": duration_lookup.loc[zone, "comparison_pattern_duration_s"],
                "extra_duration_s": extra,
                "duration_reading": (
                    "main stretched section"
                    if extra >= 60
                    else "slightly longer"
                    if extra > 5
                    else "similar length"
                    if extra >= -5
                    else "shorter after alignment"
                ),
                "mechanical_reading": notable if notable else "no large zone-level IMU difference by this summary",
                "mean_product_delta_C": mean_prod_delta,
                "largest_sensor_delta": f"{max_abs_sensor} ({max_abs_delta:.2f} C abs mean delta)" if max_abs_sensor else "",
                "product_spread_delta_C": spread_delta,
                "temperature_reading": (
                    "comparison warmer than reference"
                    if mean_prod_delta >= 0.5
                    else "comparison cooler than reference"
                    if mean_prod_delta <= -0.5
                    else "similar mean product temperature"
                ),
            }
        )
    return pd.DataFrame(rows)


def write_df(ws, df: pd.DataFrame) -> None:
    for col_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(1, col_idx, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            if isinstance(val, (float, np.floating)) and not pd.isna(val):
                val = float(val)
            elif isinstance(val, (int, np.integer)) and not pd.isna(val):
                val = int(val)
            elif pd.isna(val):
                val = None
            ws.cell(row_idx, col_idx, val)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, col in enumerate(df.columns, 1):
        samples = [str(col), *[str(v) for v in df[col].head(200).fillna("").tolist()]]
        ws.column_dimensions[get_column_letter(idx)].width = min(max(len(s) for s in samples) + 2, 34)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                if isinstance(cell.value, float):
                    cell.number_format = "0.000"
    wb.save(path)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / "aasted3_pattern_detected_hotspot_report_v2.xlsx"

    comparison = read_run(COMPARISON_CSV)
    reference = read_run(REFERENCE_CSV)
    comparison_seconds = build_seconds(comparison)
    reference_seconds = build_seconds(reference)
    map_df, dtw_cost = dtw_map(comparison_seconds, reference_seconds)
    detected_reference = detect_pattern_zones(reference_seconds, REFERENCE_RUN_NAME)
    detected_comparison = detect_pattern_zones(comparison_seconds, COMPARISON_RUN_NAME)
    detected_zones = pd.concat([detected_reference, detected_comparison], ignore_index=True)

    comparison_temp = temp_with_zones(comparison, COMPARISON_RUN_NAME, map_df)
    reference_temp = temp_with_zones(reference, REFERENCE_RUN_NAME)
    temp_all = pd.concat([reference_temp, comparison_temp], ignore_index=True)
    detected_reference_temp = temp_with_detected_zones(reference, REFERENCE_RUN_NAME, detected_reference)
    detected_comparison_temp = temp_with_detected_zones(comparison, COMPARISON_RUN_NAME, detected_comparison)
    detected_temp_all = pd.concat([detected_reference_temp, detected_comparison_temp], ignore_index=True)

    duration = duration_by_zone(reference_seconds, map_df)
    mech_delta = mechanical_delta(reference_seconds, comparison_seconds, map_df)
    product_summary = product_zone_summary(temp_all)
    product_delta = product_zone_delta(product_summary)
    hotspots = hotspot_by_zone(temp_all)
    findings = zone_findings(duration, mech_delta, product_delta)
    detected_product_summary = product_zone_summary(detected_temp_all)
    detected_product_delta = product_zone_delta(detected_product_summary)
    hotspot_summary, hotspot_long, hotspot_delta_matrix = clearer_hotspots(detected_temp_all)

    detected_duration = detected_zones.pivot_table(index="zone", columns="run", values="duration_s").reset_index()
    detected_duration.columns.name = None
    if COMPARISON_RUN_NAME in detected_duration.columns and REFERENCE_RUN_NAME in detected_duration.columns:
        detected_duration["duration_delta_comparison_minus_reference_s"] = (
            detected_duration[COMPARISON_RUN_NAME] - detected_duration[REFERENCE_RUN_NAME]
        )
    order = {name: i for i, (_, _, name, _) in enumerate(ZONES)}
    detected_duration["order"] = detected_duration["zone"].map(order)
    detected_duration = detected_duration.sort_values("order").drop(columns=["order"])

    zone_defs = pd.DataFrame(
        [
            {
                "reference_start_s": start,
                "reference_end_s": end if end < 99999 else "run_end",
                "zone": name,
                "interpretation": description,
            }
            for start, end, name, description in ZONES
        ]
    )

    extra_wall = (comparison["elapsed_s"].max() - reference["elapsed_s"].max())
    summary = pd.DataFrame(
        [
            ["reference_run", REFERENCE_CSV.name],
            ["comparison_run", COMPARISON_CSV.name],
            ["comparison_label", COMPARISON_RUN_NAME],
            ["method", "Pattern-based DTW alignment using T8, acc z mean/std, gyro y mean/std per second"],
            ["new_in_this_version", "Adds experimental per-run pattern-derived zones and clearer product-hotspot figures."],
            ["extra_wall_time_s", extra_wall],
            ["meaning_of_extra_wall_time", "Comparison run elapsed duration minus reference elapsed duration; not itself a stop window."],
            ["dtw_cost", dtw_cost],
            ["largest_zone_extra_s", duration.loc[duration["extra_duration_s"].idxmax(), "zone"]],
            ["largest_zone_extra_duration_s", duration["extra_duration_s"].max()],
            ["product_sensors", ", ".join(PRODUCT_SENSORS)],
            ["ambient/location_marker", "T8"],
        ],
        columns=["metric", "value"],
    )

    readme = pd.DataFrame(
        [
            ["rolling_window", "A moving local summary, e.g. median over nearby 30 seconds. Useful for smoothing but poor for distinguishing expected quiet zones from true process stops."],
            ["extra_wall_time", "The comparison file duration minus the reference duration. This is only the duration difference between files."],
            ["low_motion_threshold", "The prior report took the 5th percentile of reference low-motion behavior as a cut-off. That can mislabel normal quiet zones as stops."],
            ["pattern_alignment", "This report aligns the comparison run to the reference process pattern using T8, acc z, and gyro y features, then asks which zones absorbed extra time."],
            ["zone_extra_duration", "For each reference process zone, this is comparison pattern duration minus reference duration after alignment."],
            ["temperature_comparison", "Product temperatures are compared inside the aligned process zones, not only against absolute wall-clock time."],
            ["product_hotspot_definition", "For the remade hotspot sheets, a product hotspot means a product sensor (T2, T3, T4, T5, T7) is warmer than the product-sensor average within the same detected zone."],
            ["pattern_detected_zones", "Experimental per-run segmentation from T8 valleys, sustained gyro-y variability, acc-z variability, and late T8 rise. Early zones are lower confidence than vibration/cooling cycles."],
        ],
        columns=["term", "plain_language_explanation"],
    )

    path_sample = map_df.iloc[:: max(1, len(map_df) // 500)].copy()
    path_sample["comparison_zone_assigned"] = path_sample["reference_zone"]

    wb = Workbook()
    wb.remove(wb.active)
    sheets = {
        "Read Me": readme,
        "Summary": summary,
        "Zone Definitions": zone_defs,
        "Zone Findings": findings,
        "Detected Zones": detected_zones,
        "Detected Duration": detected_duration,
        "Zone Duration": duration,
        "Mechanical By Zone": mech_delta,
        "Product Temp By Zone": product_summary,
        "Product Delta By Zone": product_delta,
        "Old Hotspots By Zone": hotspots,
        "Hotspot Summary": hotspot_summary,
        "Hotspot Sensor Data": hotspot_long,
        "Hotspot Delta Matrix": hotspot_delta_matrix,
        "Detected Product By Zone": detected_product_summary,
        "Detected Product Delta": detected_product_delta,
        "Alignment Path Sample": path_sample,
    }
    for name, df in sheets.items():
        ws = wb.create_sheet(name)
        write_df(ws, df)

    ws = wb["Zone Duration"]
    if ws.max_row > 1:
        ws.conditional_formatting.add(
            f"D2:D{ws.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="63BE7B",
                mid_type="num",
                mid_value=0,
                mid_color="FFFFFF",
                end_type="max",
                end_color="F8696B",
            ),
        )
        chart = BarChart()
        chart.title = "Extra Duration by Process Zone"
        chart.y_axis.title = "seconds"
        chart.x_axis.title = "zone"
        data = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18
        ws.add_chart(chart, "F2")

    ws = wb["Detected Duration"]
    if ws.max_row > 1 and ws.max_column >= 4:
        ws.conditional_formatting.add(
            f"D2:D{ws.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="63BE7B",
                mid_type="num",
                mid_value=0,
                mid_color="FFFFFF",
                end_type="max",
                end_color="F8696B",
            ),
        )
        chart = BarChart()
        chart.title = "Detected Zone Duration Delta"
        chart.y_axis.title = "seconds"
        chart.x_axis.title = "zone"
        data = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18
        ws.add_chart(chart, "F2")

    ws = wb["Product Delta By Zone"]
    if ws.max_row > 1 and ws.max_column >= 5:
        ws.conditional_formatting.add(
            f"E2:E{ws.max_row}",
            ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="num", mid_value=0, mid_color="FFFFFF", end_type="max", end_color="F8696B"),
        )

    ws = wb["Hotspot Delta Matrix"]
    if ws.max_row > 1 and ws.max_column >= 7:
        ws.conditional_formatting.add(
            f"C2:G{ws.max_row}",
            ColorScaleRule(start_type="min", start_color="4472C4", mid_type="num", mid_value=0, mid_color="FFFFFF", end_type="max", end_color="F8696B"),
        )

    ws = wb["Hotspot Summary"]
    if ws.max_row > 1:
        spread_col = None
        for cell in ws[1]:
            if cell.value == "spread_hottest_minus_coolest_C":
                spread_col = cell.column
                break
        if spread_col:
            col = get_column_letter(spread_col)
            ws.conditional_formatting.add(
                f"{col}2:{col}{ws.max_row}",
                ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="F8696B"),
            )
            chart = BarChart()
            chart.title = "Product Hotspot Spread by Detected Zone"
            chart.y_axis.title = "deg C"
            chart.x_axis.title = "run/zone row"
            data = Reference(ws, min_col=spread_col, min_row=1, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.height = 8
            chart.width = 18
            ws.add_chart(chart, "U2")

    # Add compact trajectory sheet for charting the alignment.
    chart_df = reference_seconds[["elapsed_sec", "T8", "accz_mean", "gyroy_mean", "zone_reference"]].iloc[::2].copy()
    ws = wb.create_sheet("Reference Pattern")
    write_df(ws, chart_df)
    chart = LineChart()
    chart.title = "Reference Pattern: T8, acc z, gyro y"
    chart.y_axis.title = "scaled raw values"
    chart.x_axis.title = "elapsed s"
    data = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 20
    ws.add_chart(chart, "G2")

    chart_df = comparison_seconds[["elapsed_sec", "T8", "accz_mean", "gyroy_mean"]].iloc[::2].copy()
    ws = wb.create_sheet("Comparison Pattern")
    write_df(ws, chart_df)
    chart = LineChart()
    chart.title = "Comparison Pattern: T8, acc z, gyro y"
    chart.y_axis.title = "scaled raw values"
    chart.x_axis.title = "elapsed s"
    data = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 20
    ws.add_chart(chart, "F2")

    wb.save(out_path)
    style_workbook(out_path)
    print(out_path)


if __name__ == "__main__":
    main()
