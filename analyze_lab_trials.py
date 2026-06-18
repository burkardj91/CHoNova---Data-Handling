from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


WORKSPACE = Path(r"C:\Users\BurkardJohannes\Documents\CHoNova _ Data Understanding")
INPUT_DIR = WORKSPACE / "inputs" / "lab_trials"
OUTPUT_DIR = WORKSPACE / "outputs" / "lab_trials"

RAW_XLSX = INPUT_DIR / "lab_trials_raw.xlsx"
SUMMARY_XLSX = INPUT_DIR / "lab_trials_summary.xlsx"
SETUP_XLSX = INPUT_DIR / "lab_trials_experimental setup.xlsx"
CSV_INPUT_DIR = Path(r"C:\Users\BurkardJohannes\Desktop\Temp Files\AAK")

PRODUCT_TEMPS = ["T1", "T3", "T4", "T5"]
MOULD_TEMPS = ["T2"]
AMBIENT_TEMPS = ["T6"]
PRIMARY_US = ["Rx1Tx1", "Rx2Tx2"]
ALL_US = ["Rx1Tx1", "Rx2Tx2", "Rx1Tx2", "Rx2Tx1"]
TIME_COL = "Absolut[s]"
SENSOR_CODE_TO_PRIMARY_US = {"11": "Rx1Tx1", "22": "Rx2Tx2"}

CSV_RAW_SOURCES = {
    "at2455_a_test": CSV_INPUT_DIR / "a_test.csv",
    "at2455_a_test3": CSV_INPUT_DIR / "a_test_3.csv",
    "at2455_a_test4": CSV_INPUT_DIR / "a_test_4.csv",
    "at2455_a5": CSV_INPUT_DIR / "a5.csv",
    "at2455_a6": CSV_INPUT_DIR / "a6.csv",
    "at2455_a7": CSV_INPUT_DIR / "a7.csv",
    "at2455_b": CSV_INPUT_DIR / "b1.csv",
    "at2455_b2": CSV_INPUT_DIR / "b2.csv",
    "at2455_b3": CSV_INPUT_DIR / "b3.csv",
    "at2455_b4": CSV_INPUT_DIR / "b4.csv",
    "at2561_a1": CSV_INPUT_DIR / "a1.csv",
    "at2561_a2": CSV_INPUT_DIR / "a2.csv",
    "at2561_a3": CSV_INPUT_DIR / "a3.csv",
    "at2561_a4": CSV_INPUT_DIR / "a4.csv",
}

PROCESS_START_TIMES = {
    "2455_a_test": {"deposition_s": 155.0, "cooling_start_s": 210.0, "zone_profile": "AAK_cooling_tunnel"},
    "2455_a_test3": {"deposition_s": 106.0, "cooling_start_s": 155.0, "zone_profile": "AAK_cooling_tunnel"},
    "2455_a_test4": {"deposition_s": 676.0, "cooling_start_s": 728.0, "zone_profile": "AAK_cooling_tunnel"},
    "2455_a5": {"deposition_s": 650.0, "cooling_start_s": 695.0, "zone_profile": "lab_device_simulator"},
    "2455_a6": {"deposition_s": 1245.0, "cooling_start_s": 1285.0, "zone_profile": "lab_device_simulator"},
    "2455_a7": {"deposition_s": 605.0, "cooling_start_s": 660.0, "zone_profile": "lab_device_simulator"},
    "2455_b": {"deposition_s": 378.0, "cooling_start_s": 406.0, "zone_profile": "lab_device_simulator"},
    "2455_b2": {"deposition_s": 156.0, "cooling_start_s": 217.0, "zone_profile": "lab_device_simulator"},
    "2455_b3": {"deposition_s": 124.0, "cooling_start_s": 162.0, "zone_profile": "lab_device_simulator"},
    "2455_b4": {"deposition_s": 178.0, "cooling_start_s": 228.0, "zone_profile": "lab_device_simulator"},
    "2561_a1": {"deposition_s": 315.0, "cooling_start_s": 382.0, "zone_profile": "lab_device_simulator"},
    "2561_a2": {"deposition_s": 516.0, "cooling_start_s": 562.0, "zone_profile": "lab_device_simulator"},
    "2561_a3": {"deposition_s": 622.0, "cooling_start_s": 660.0, "zone_profile": "lab_device_simulator"},
    "2561_a4": {"deposition_s": 598.0, "cooling_start_s": 645.0, "zone_profile": "lab_device_simulator"},
}

LAB_DEVICE_REFERENCE = {
    "reference_rep_id": "2455_a5",
    "zone_2_end_s": 2000.0,
    "demoulding_start_s": 2550.0,
}


def norm_id(value: object) -> str:
    s = str(value or "").strip().lower()
    if s.startswith("at"):
        s = s[2:]
    return s.replace(" ", "_")


def load_setup() -> pd.DataFrame:
    df = pd.read_excel(SETUP_XLSX, sheet_name=0)
    df = df.dropna(how="all").copy()
    df["rep_id"] = df["Experimental Name"].map(norm_id)
    return df


def parse_summary() -> tuple[pd.DataFrame, pd.DataFrame]:
    ws = load_workbook(SUMMARY_XLSX, read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    dataset_rows = [idx for idx, row in enumerate(rows) if row and row[0] == "Dataset"]
    dataset_rows.append(len(rows))

    blocks = []
    metrics = []
    for block_idx, start in enumerate(dataset_rows[:-1]):
        end = dataset_rows[block_idx + 1]
        block = rows[start:end]
        dataset = block[0][1]
        chocolate = next((row[1] for row in block if row and row[0] == "Chocolate Type"), None)
        tempering = next((row[1] for row in block if row and row[0] == "Tempering"), None)
        cooling = next((row[1] for row in block if row and row[0] == "Cooling"), None)
        blocks.append(
            {
                "dataset": dataset,
                "summary_chocolate_type": chocolate,
                "summary_tempering": tempering,
                "summary_cooling": cooling,
            }
        )

        rep_row_idx = next((i for i, row in enumerate(block) if row and row[0] == "Repetitions"), None)
        sensor_row_idx = next((i for i, row in enumerate(block) if row and row[0] == "Sensors"), None)
        if rep_row_idx is None or sensor_row_idx is None:
            continue
        rep_row = block[rep_row_idx]
        sensor_row = block[sensor_row_idx]
        col_to_rep_sensor = {}
        current_rep = None
        for col_idx in range(1, min(len(rep_row), 10)):
            rep = rep_row[col_idx]
            if rep is not None:
                current_rep = rep
            sensor = sensor_row[col_idx] if col_idx < len(sensor_row) else None
            if current_rep is not None and sensor is not None:
                col_to_rep_sensor[col_idx] = (norm_id(current_rep), str(int(sensor)) if isinstance(sensor, (int, float)) else str(sensor))

        for row in block[sensor_row_idx + 1 :]:
            label = row[0] if row else None
            if label not in {
                "Viscosity Damping",
                "Cryst. Onset (rel.)",
                "Cryst. Efficiency",
                "Detachment Onset (rel.)",
                "Detachment Completion (rel.)",
                "Detachment Duration",
                "Cryst. Length",
            }:
                continue
            for col_idx, (rep_id, sensor_code) in col_to_rep_sensor.items():
                value = row[col_idx] if col_idx < len(row) else None
                if value is None:
                    continue
                metrics.append(
                    {
                        "dataset": dataset,
                        "rep_id": rep_id,
                        "sensor_code": sensor_code,
                        "summary_metric": label,
                        "value": value,
                    }
                )
    return pd.DataFrame(blocks), pd.DataFrame(metrics)


def build_landmarks(summary_metrics: pd.DataFrame, setup: pd.DataFrame, blocks: pd.DataFrame) -> pd.DataFrame:
    pivot = summary_metrics.pivot_table(
        index=["dataset", "rep_id", "sensor_code"],
        columns="summary_metric",
        values="value",
        aggfunc="first",
    ).reset_index()
    pivot.columns.name = None
    pivot = pivot.merge(setup, on="rep_id", how="left")
    pivot = pivot.merge(blocks, on="dataset", how="left")
    pivot["raw_sheet"] = "at" + pivot["rep_id"]
    pivot["cooling_start_s_assumed"] = 0.0
    return pivot


def read_raw_sheet(sheet: str) -> pd.DataFrame | None:
    csv_path = CSV_RAW_SOURCES.get(sheet)
    if csv_path is not None and csv_path.exists():
        df = pd.read_csv(csv_path)
        if "time" in df.columns and TIME_COL not in df.columns:
            df = df.rename(columns={"time": TIME_COL})
    else:
        try:
            df = pd.read_excel(RAW_XLSX, sheet_name=sheet)
        except ValueError:
            return None
    if df.empty or TIME_COL not in df.columns:
        return None
    df = df.dropna(subset=[TIME_COL]).copy()
    return df


def available_raw_sheets() -> list[str]:
    sheets = [sheet for sheet, path in CSV_RAW_SOURCES.items() if path.exists()]
    if sheets:
        return sheets
    return load_workbook(RAW_XLSX, read_only=True, data_only=True).sheetnames


def rebuild_lab_raw_workbook_from_csv() -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet in available_raw_sheets():
        df = read_raw_sheet(sheet)
        if df is None:
            continue
        ws = wb.create_sheet(sheet)
        for c_idx, col in enumerate(df.columns, 1):
            ws.cell(1, c_idx, col)
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                if isinstance(value, (np.floating, float)) and pd.notna(value):
                    value = float(value)
                elif isinstance(value, (np.integer, int)) and pd.notna(value):
                    value = int(value)
                elif pd.isna(value):
                    value = None
                ws.cell(r_idx, c_idx, value)
    try:
        wb.save(RAW_XLSX)
        return RAW_XLSX
    except PermissionError:
        fallback = INPUT_DIR / "lab_trials_raw_from_csv.xlsx"
        wb.save(fallback)
        return fallback


def zone_bounds(row: pd.Series) -> list[dict[str, object]]:
    cryst = row.get("Cryst. Onset (rel.)")
    det_on = row.get("Detachment Onset (rel.)")
    det_comp = row.get("Detachment Completion (rel.)")
    zones = [
        {
            "zone": "zone_1_deposition_to_cooling_start",
            "start_s": np.nan,
            "end_s": 0.0,
            "status": "not_computed_missing_deposition_to_cooling_start_time",
        },
        {
            "zone": "zone_2_cooling_to_crystallization_start",
            "start_s": 0.0,
            "end_s": cryst,
            "status": "computed_assuming_cooling_start_at_0s",
        },
        {
            "zone": "zone_3_crystallization_to_detachment_onset",
            "start_s": cryst,
            "end_s": det_on,
            "status": "computed_from_summary_landmarks",
        },
        {
            "zone": "zone_4_detachment_onset_to_detachment_completion",
            "start_s": det_on,
            "end_s": det_comp,
            "status": "computed_from_summary_landmarks",
        },
    ]
    return zones


def safe_slope(x: pd.Series, y: pd.Series) -> float:
    valid = pd.DataFrame({"x": x, "y": y}).dropna()
    if len(valid) < 2 or valid["x"].nunique() < 2:
        return np.nan
    return float(np.polyfit(valid["x"], valid["y"], 1)[0])


def auc(x: pd.Series, y: pd.Series) -> float:
    valid = pd.DataFrame({"x": x, "y": y}).dropna()
    if len(valid) < 2:
        return np.nan
    return float(np.trapezoid(valid["y"], valid["x"]))


def resample_curve(seg: pd.DataFrame, channel: str, n: int = 101) -> np.ndarray | None:
    valid = seg[[TIME_COL, channel]].dropna().sort_values(TIME_COL)
    if len(valid) < 2 or valid[TIME_COL].nunique() < 2:
        return None
    x = valid[TIME_COL].to_numpy(dtype=float)
    y = valid[channel].to_numpy(dtype=float)
    xn = (x - x.min()) / (x.max() - x.min())
    grid = np.linspace(0, 1, n)
    return np.interp(grid, xn, y)


def analyze() -> tuple[pd.DataFrame, dict[str, pd.DataFrame]]:
    setup = load_setup()
    blocks, summary_metrics = parse_summary()
    landmarks = build_landmarks(summary_metrics, setup, blocks)

    raw_sheets = available_raw_sheets()
    raw_availability = pd.DataFrame(
        [{"raw_sheet": sheet, "rep_id": norm_id(sheet), "available": True} for sheet in raw_sheets]
    )

    us_metrics = []
    temp_metrics = []
    zone_rows = []
    curve_rows = []
    raw_cache: dict[str, pd.DataFrame | None] = {}

    for _, lm in landmarks.iterrows():
        raw_sheet = lm["raw_sheet"]
        if raw_sheet not in raw_cache:
            raw_cache[raw_sheet] = read_raw_sheet(raw_sheet)
        df = raw_cache[raw_sheet]
        for zone in zone_bounds(lm):
            zrow = {
                "dataset": lm["dataset"],
                "rep_id": lm["rep_id"],
                "sensor_code": lm["sensor_code"],
                "raw_sheet": raw_sheet,
                **zone,
            }
            if df is None:
                zrow["raw_status"] = "missing_or_empty_raw_sheet"
                zone_rows.append(zrow)
                continue
            if pd.isna(zone["start_s"]) or pd.isna(zone["end_s"]) or float(zone["end_s"]) <= float(zone["start_s"]):
                zrow["raw_status"] = "zone_not_computed"
                zone_rows.append(zrow)
                continue
            seg = df[(df[TIME_COL] >= float(zone["start_s"])) & (df[TIME_COL] <= float(zone["end_s"]))].copy()
            zrow["raw_status"] = "ok" if not seg.empty else "no_rows_in_zone"
            zrow["n_rows"] = len(seg)
            zone_rows.append(zrow)
            if seg.empty:
                continue

            for ch in ALL_US:
                if ch not in seg.columns:
                    continue
                vals = seg[ch].dropna()
                us_metrics.append(
                    {
                        **{k: lm.get(k) for k in ["dataset", "rep_id", "sensor_code", "Chocolate Recipe", "Cooling Configuration", "Cooling Temperature (°C)", "Cooling Speed (%)", "Tempering", "Tempering Degree", "Trials"]},
                        "zone": zone["zone"],
                        "channel": ch,
                        "role": "primary" if ch in PRIMARY_US else "secondary",
                        "duration_s": float(zone["end_s"]) - float(zone["start_s"]),
                        "mean": vals.mean(),
                        "std": vals.std(),
                        "min": vals.min(),
                        "max": vals.max(),
                        "range": vals.max() - vals.min() if len(vals) else np.nan,
                        "auc": auc(seg[TIME_COL], seg[ch]),
                        "slope": safe_slope(seg[TIME_COL], seg[ch]),
                        "n_points": len(vals),
                    }
                )
                curve = resample_curve(seg, ch)
                if curve is not None:
                    curve_rows.append(
                        {
                            "dataset": lm["dataset"],
                            "rep_id": lm["rep_id"],
                            "sensor_code": lm["sensor_code"],
                            "zone": zone["zone"],
                            "channel": ch,
                            "curve": curve,
                        }
                    )

            available_product = [c for c in PRODUCT_TEMPS if c in seg.columns]
            product_frame = seg[available_product].copy()
            product_spread = product_frame.max(axis=1) - product_frame.min(axis=1) if available_product else pd.Series(dtype=float)
            for ch in PRODUCT_TEMPS + MOULD_TEMPS + AMBIENT_TEMPS:
                if ch not in seg.columns:
                    continue
                vals = seg[ch].dropna()
                temp_metrics.append(
                    {
                        **{k: lm.get(k) for k in ["dataset", "rep_id", "sensor_code", "Chocolate Recipe", "Cooling Configuration", "Cooling Temperature (°C)", "Cooling Speed (%)", "Tempering", "Tempering Degree", "Trials"]},
                        "zone": zone["zone"],
                        "sensor": ch,
                        "sensor_role": "product" if ch in PRODUCT_TEMPS else "mould" if ch in MOULD_TEMPS else "ambient",
                        "duration_s": float(zone["end_s"]) - float(zone["start_s"]),
                        "mean_C": vals.mean(),
                        "std_C": vals.std(),
                        "min_C": vals.min(),
                        "max_C": vals.max(),
                        "range_C": vals.max() - vals.min() if len(vals) else np.nan,
                        "slope_C_per_s": safe_slope(seg[TIME_COL], seg[ch]),
                        "mean_product_spread_C": product_spread.mean() if ch in PRODUCT_TEMPS and len(product_spread) else np.nan,
                        "n_points": len(vals),
                    }
                )

    us_df = pd.DataFrame(us_metrics)
    temp_df = pd.DataFrame(temp_metrics)
    curves = pd.DataFrame(curve_rows)

    repeatability_rows = []
    if not curves.empty:
        for keys, group in curves.groupby(["dataset", "sensor_code", "zone", "channel"]):
            arr = np.vstack(group["curve"].to_numpy())
            mean_curve = np.nanmean(arr, axis=0)
            for _, row in group.iterrows():
                curve = row["curve"]
                rmse = float(np.sqrt(np.nanmean((curve - mean_curve) ** 2)))
                if np.nanstd(curve) == 0 or np.nanstd(mean_curve) == 0:
                    corr = np.nan
                else:
                    corr = float(np.corrcoef(curve, mean_curve)[0, 1])
                repeatability_rows.append(
                    {
                        "dataset": row["dataset"],
                        "sensor_code": row["sensor_code"],
                        "zone": row["zone"],
                        "channel": row["channel"],
                        "rep_id": row["rep_id"],
                        "normalized_curve_rmse_to_group_mean": rmse,
                        "correlation_to_group_mean": corr,
                    }
                )
    repeat_df = pd.DataFrame(repeatability_rows)
    if not us_df.empty and not repeat_df.empty:
        us_df = us_df.merge(repeat_df, on=["dataset", "rep_id", "sensor_code", "zone", "channel"], how="left")

    batch_us = summarize_groups(us_df, ["dataset", "sensor_code", "zone", "channel"], "us")
    setup_us = summarize_groups(us_df, ["Chocolate Recipe", "Cooling Configuration", "Tempering", "Tempering Degree", "sensor_code", "zone", "channel"], "us")
    batch_temp = summarize_groups(temp_df, ["dataset", "sensor_code", "zone", "sensor"], "temp")
    setup_temp = summarize_groups(temp_df, ["Chocolate Recipe", "Cooling Configuration", "Tempering", "Tempering Degree", "sensor_code", "zone", "sensor"], "temp")

    tables = {
        "Setup Metadata": setup,
        "Summary Blocks": blocks,
        "Parsed Landmarks": landmarks,
        "Zone Boundaries": pd.DataFrame(zone_rows),
        "US Zone Metrics": us_df,
        "US Batch Repeatability": batch_us,
        "US Setup Comparison": setup_us,
        "Temperature Zone Metrics": temp_df,
        "Temperature Batch Repeatability": batch_temp,
        "Temperature Setup Comparison": setup_temp,
        "Raw Availability": raw_availability,
    }
    return landmarks, tables


def summarize_groups(df: pd.DataFrame, keys: list[str], kind: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    keys = [k for k in keys if k in df.columns]
    if kind == "us":
        agg = {
            "mean": ["mean", "std"],
            "range": ["mean", "std"],
            "auc": ["mean", "std"],
            "slope": ["mean", "std"],
            "normalized_curve_rmse_to_group_mean": ["mean", "max"],
            "correlation_to_group_mean": ["mean", "min"],
            "rep_id": "nunique",
        }
    else:
        agg = {
            "mean_C": ["mean", "std"],
            "range_C": ["mean", "std"],
            "slope_C_per_s": ["mean", "std"],
            "mean_product_spread_C": ["mean", "std"],
            "rep_id": "nunique",
        }
    out = df.groupby(keys, dropna=False).agg(agg)
    out.columns = ["_".join(c for c in col if c) for col in out.columns.to_flat_index()]
    return out.reset_index()


def display_factor_value(value: object) -> str:
    if pd.isna(value):
        return ""
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return f"{int(value.day):02d}-{int(value.month):02d}-{int(value.year) % 100:02d}"
    return str(value)


def effect_size_eta2(df: pd.DataFrame, response: str, factor: str) -> tuple[float, int, str]:
    if df.empty or response not in df.columns or factor not in df.columns:
        return np.nan, 0, "not_available"
    work = df[[response, factor]].dropna()
    work[factor] = work[factor].map(display_factor_value)
    if work.empty or work[factor].nunique(dropna=True) < 2 or work[response].nunique(dropna=True) < 2:
        return np.nan, int(work[factor].nunique(dropna=True)), "insufficient_variation"
    overall = work[response].mean()
    ss_total = float(((work[response] - overall) ** 2).sum())
    if ss_total <= 0:
        return np.nan, int(work[factor].nunique(dropna=True)), "insufficient_variation"
    ss_between = 0.0
    for _, group in work.groupby(factor):
        ss_between += len(group) * float((group[response].mean() - overall) ** 2)
    eta2 = ss_between / ss_total
    if eta2 >= 0.35:
        strength = "large"
    elif eta2 >= 0.14:
        strength = "moderate"
    elif eta2 >= 0.06:
        strength = "small"
    else:
        strength = "weak"
    return eta2, int(work[factor].nunique(dropna=True)), strength


def best_group_difference(df: pd.DataFrame, response: str, factor: str) -> str:
    if df.empty or response not in df.columns or factor not in df.columns:
        return ""
    work = df[[response, factor]].dropna()
    work[factor] = work[factor].map(display_factor_value)
    if work[factor].nunique() < 2:
        return ""
    means = work.groupby(factor)[response].mean().sort_values()
    low_label, high_label = means.index[0], means.index[-1]
    return f"{high_label} ({means.iloc[-1]:.4g}) vs {low_label} ({means.iloc[0]:.4g})"


def screen_effects(df: pd.DataFrame, response_specs: list[tuple[str, str]], factors: list[str], source: str) -> pd.DataFrame:
    rows = []
    for response, response_label in response_specs:
        if response not in df.columns:
            continue
        for factor in factors:
            eta2, n_levels, strength = effect_size_eta2(df, response, factor)
            rows.append(
                {
                    "source": source,
                    "response": response,
                    "response_label": response_label,
                    "factor": factor,
                    "n_levels": n_levels,
                    "eta2_effect_size": eta2,
                    "effect_strength": strength,
                    "largest_mean_difference": best_group_difference(df, response, factor),
                    "n_observations": int(df[[response, factor]].dropna().shape[0]) if factor in df.columns else 0,
                    "model_note": "Effect-size screening fallback because statsmodels MixedLM is unavailable in this runtime.",
                }
            )
    return pd.DataFrame(rows).sort_values(["source", "response", "eta2_effect_size"], ascending=[True, True, False])


def make_question_summaries(tables: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    us_df = tables["US Zone Metrics"].copy()
    temp_df = tables["Temperature Zone Metrics"].copy()
    landmarks = tables["Parsed Landmarks"].copy()

    question_a_rows = [
        {
            "question": "A",
            "analysis": "Within experimental set / batch repeatability by zone",
            "included_signals": "Primary and secondary ultrasound; product temperature T1/T3/T4/T5; mould T2; ambient T6",
            "main_tables": "US Batch Repeatability; Temperature Batch Repeatability",
            "interpretation": "Low std/RMSE and high correlation indicate repeatable zone behavior within a set.",
        },
        {
            "question": "B",
            "analysis": "Differences between experimental setups",
            "included_factors": "Chocolate Recipe, Cooling Configuration, Cooling Temperature, Cooling Speed, Tempering, Tempering Degree",
            "main_tables": "US Setup Comparison; Temperature Setup Comparison",
            "interpretation": "Compares whether batches with different setups have shifted ultrasound or thermal summaries.",
        },
        {
            "question": "C",
            "analysis": "Mixed-effect style screening of extracted parameters",
            "included_responses": "Summary parameters plus per-zone ultrasound and temperature features",
            "main_tables": "Effect Screening; Summary Parameter Screening",
            "interpretation": "Uses eta-squared effect sizes as a fallback screen for setup/zone/sensor effects.",
        },
    ]

    temp_repeatability = (
        temp_df.groupby(["dataset", "zone", "sensor_role"], dropna=False)
        .agg(
            repetitions=("rep_id", "nunique"),
            sensors=("sensor", "nunique"),
            mean_temperature_std_across_reps_C=("mean_C", "std"),
            mean_range_C=("range_C", "mean"),
            mean_abs_slope_C_per_s=("slope_C_per_s", lambda s: s.abs().mean()),
            mean_product_spread_C=("mean_product_spread_C", "mean"),
        )
        .reset_index()
        if not temp_df.empty
        else pd.DataFrame()
    )
    us_repeatability = (
        us_df.groupby(["dataset", "zone", "role", "channel"], dropna=False)
        .agg(
            repetitions=("rep_id", "nunique"),
            mean_signal_std_across_reps=("mean", "std"),
            mean_range=("range", "mean"),
            mean_curve_rmse=("normalized_curve_rmse_to_group_mean", "mean"),
            min_curve_correlation=("correlation_to_group_mean", "min"),
        )
        .reset_index()
        if not us_df.empty
        else pd.DataFrame()
    )

    setup_factors = [
        "Chocolate Recipe",
        "Cooling Configuration",
        "Cooling Temperature (°C)",
        "Cooling Speed (%)",
        "Tempering",
        "Tempering Degree",
        "dataset",
        "zone",
        "sensor_code",
        "channel",
        "sensor",
        "sensor_role",
    ]
    effect_tables = []
    effect_tables.append(
        screen_effects(
            us_df,
            [
                ("mean", "ultrasound mean amplitude"),
                ("range", "ultrasound range"),
                ("auc", "ultrasound area under curve"),
                ("normalized_curve_rmse_to_group_mean", "ultrasound curve RMSE to group mean"),
                ("correlation_to_group_mean", "ultrasound curve correlation to group mean"),
            ],
            setup_factors,
            "ultrasound_zone_metrics",
        )
    )
    effect_tables.append(
        screen_effects(
            temp_df,
            [
                ("mean_C", "temperature mean"),
                ("range_C", "temperature range"),
                ("slope_C_per_s", "temperature slope"),
                ("mean_product_spread_C", "product temperature spread"),
            ],
            setup_factors,
            "temperature_zone_metrics",
        )
    )
    effect_screening = pd.concat(effect_tables, ignore_index=True)

    summary_effects = pd.DataFrame()
    if not landmarks.empty:
        summary_long = landmarks.melt(
            id_vars=[
                "dataset",
                "rep_id",
                "sensor_code",
                "Chocolate Recipe",
                "Cooling Configuration",
                "Cooling Temperature (°C)",
                "Cooling Speed (%)",
                "Tempering",
                "Tempering Degree",
                "Trials",
            ],
            value_vars=[
                c
                for c in [
                    "Viscosity Damping",
                    "Cryst. Onset (rel.)",
                    "Cryst. Efficiency",
                    "Detachment Onset (rel.)",
                    "Detachment Completion (rel.)",
                    "Detachment Duration",
                    "Cryst. Length",
                ]
                if c in landmarks.columns
            ],
            var_name="summary_parameter",
            value_name="summary_value",
        ).dropna(subset=["summary_value"])
        rows = []
        for parameter, group in summary_long.groupby("summary_parameter"):
            screened = screen_effects(
                group,
                [("summary_value", parameter)],
                [
                    "Chocolate Recipe",
                    "Cooling Configuration",
                    "Cooling Temperature (°C)",
                    "Cooling Speed (%)",
                    "Tempering",
                    "Tempering Degree",
                    "dataset",
                    "sensor_code",
                ],
                "lab_trials_summary",
            )
            screened["summary_parameter"] = parameter
            rows.append(screened)
        summary_effects = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()

    top_findings = pd.concat(
        [
            effect_screening[effect_screening["effect_strength"].isin(["large", "moderate"])].head(40),
            summary_effects[summary_effects["effect_strength"].isin(["large", "moderate"])].head(40),
        ],
        ignore_index=True,
    )

    return {
        "Research Questions": pd.DataFrame(question_a_rows),
        "A Temp Repeatability": temp_repeatability,
        "A US Repeatability": us_repeatability,
        "C Effect Screening": effect_screening,
        "C Summary Param Screening": summary_effects,
        "Top Effect Findings": top_findings,
    }


def zone_short(zone: object) -> str:
    labels = {
        "zone_1_deposition_to_cooling_start": "Z1 deposition -> cooling start",
        "zone_2_cooling_to_crystallization_start": "Z2 cooling -> crystallization start",
        "zone_3_crystallization_to_detachment_onset": "Z3 crystallization -> detachment onset",
        "zone_4_detachment_onset_to_detachment_completion": "Z4 detachment onset -> completion",
    }
    return labels.get(str(zone), str(zone))


def zone_order(zone: object) -> int:
    s = str(zone)
    if "zone_1" in s:
        return 1
    if "zone_2" in s:
        return 2
    if "zone_3" in s:
        return 3
    if "zone_4" in s:
        return 4
    return 99


def fmt_num(value: object, digits: int = 3) -> str:
    if pd.isna(value):
        return "n/a"
    return f"{float(value):.{digits}f}"


def setup_text(row: pd.Series) -> str:
    parts = []
    for col in ["Chocolate Recipe", "Cooling Configuration", "Cooling Temperature (°C)", "Cooling Speed (%)", "Tempering", "Tempering Degree"]:
        if col in row and pd.notna(row[col]):
            parts.append(f"{col}: {display_factor_value(row[col])}")
    return "; ".join(parts)


def make_how_to_read_sheet() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "step": 1,
                "sheet": "How To Read",
                "purpose": "Explains the workflow and the meaning of the front summary sheets.",
                "how_to_read": "Start here, then read the zone sheets from left to right.",
            },
            {
                "step": 2,
                "sheet": "Zone Within Setup",
                "purpose": "Question A: for each zone and each experimental setup, checks repeatability between repetitions.",
                "how_to_read": "Lower temperature standard deviation, lower ultrasound RMSE, and higher ultrasound correlation mean better repeatability.",
            },
            {
                "step": 3,
                "sheet": "Zone Between Setups",
                "purpose": "Question B: for each zone, compares whether the experimental sets sit at different temperature or ultrasound levels.",
                "how_to_read": "The between-setup range shows how far apart the setup means are; the high/low labels show which setup drives the spread.",
            },
            {
                "step": 4,
                "sheet": "Zone Summary Links",
                "purpose": "Links zone behavior to summary outcomes such as crystallization start and detachment timing.",
                "how_to_read": "Use this when asking whether, for example, zone 2 temperature or ultrasound behavior may explain differences in crystallization start.",
            },
            {
                "step": 5,
                "sheet": "Condition Effects",
                "purpose": "Question C: relates extracted parameters to chocolate type, cooling, tempering, zone, and sensor/channel factors.",
                "how_to_read": "Eta-squared is a screening effect size: weak < 0.06, small 0.06-0.14, moderate 0.14-0.35, large >= 0.35.",
            },
            {
                "step": 6,
                "sheet": "Detailed Data",
                "purpose": "The later sheets are the audit trail with all parsed and calculated values.",
                "how_to_read": "Use these only when you want to inspect the numbers behind a front-sheet conclusion.",
            },
            {
                "step": 7,
                "sheet": "Zone note",
                "purpose": "Zone 1 is not interpreted yet.",
                "how_to_read": "The uploaded summary gives crystallization and detachment landmarks relative to cooling start, but not a reliable deposition-to-cooling-start boundary for every repetition.",
            },
        ]
    )


def make_zone_within_setup_summary(tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    temp_df = tables["Temperature Zone Metrics"].copy()
    us_df = tables["US Zone Metrics"].copy()
    if temp_df.empty and us_df.empty:
        return pd.DataFrame()

    setup_cols = ["Chocolate Recipe", "Cooling Configuration", "Cooling Temperature (°C)", "Cooling Speed (%)", "Tempering", "Tempering Degree"]
    setup_meta = tables["Parsed Landmarks"].drop_duplicates("dataset").set_index("dataset")
    datasets = sorted(set(temp_df.get("dataset", pd.Series(dtype=object))).union(set(us_df.get("dataset", pd.Series(dtype=object)))))
    zones = sorted(set(temp_df.get("zone", pd.Series(dtype=object))).union(set(us_df.get("zone", pd.Series(dtype=object)))), key=zone_order)

    rows = []
    for dataset in datasets:
        for zone in zones:
            t = temp_df[(temp_df["dataset"] == dataset) & (temp_df["zone"] == zone)] if not temp_df.empty else pd.DataFrame()
            u = us_df[(us_df["dataset"] == dataset) & (us_df["zone"] == zone) & (us_df["role"] == "primary")] if not us_df.empty else pd.DataFrame()
            if t.empty and u.empty:
                continue

            sensor_repeat = pd.DataFrame()
            if not t.empty:
                sensor_repeat = (
                    t.groupby(["sensor_role", "sensor"], dropna=False)
                    .agg(
                        repetitions=("rep_id", "nunique"),
                        mean_temperature_C=("mean_C", "mean"),
                        rep_to_rep_std_C=("mean_C", "std"),
                        mean_range_C=("range_C", "mean"),
                        mean_abs_slope_C_per_s=("slope_C_per_s", lambda s: s.abs().mean()),
                        mean_product_spread_C=("mean_product_spread_C", "mean"),
                    )
                    .reset_index()
                )

            def role_value(role: str, col: str, func: str = "mean") -> float:
                if sensor_repeat.empty:
                    return np.nan
                vals = sensor_repeat.loc[sensor_repeat["sensor_role"] == role, col].dropna()
                if vals.empty:
                    return np.nan
                return float(vals.max() if func == "max" else vals.mean())

            us_channel = pd.DataFrame()
            if not u.empty:
                us_channel = (
                    u.groupby("channel", dropna=False)
                    .agg(
                        repetitions=("rep_id", "nunique"),
                        mean_signal=("mean", "mean"),
                        rep_to_rep_signal_std=("mean", "std"),
                        mean_curve_rmse=("normalized_curve_rmse_to_group_mean", "mean"),
                        max_curve_rmse=("normalized_curve_rmse_to_group_mean", "max"),
                        min_curve_correlation=("correlation_to_group_mean", "min"),
                    )
                    .reset_index()
                )

            us_rmse = float(us_channel["mean_curve_rmse"].mean()) if not us_channel.empty else np.nan
            us_corr = float(us_channel["min_curve_correlation"].min()) if not us_channel.empty else np.nan
            prod_std = role_value("product", "rep_to_rep_std_C")
            mould_std = role_value("mould", "rep_to_rep_std_C")
            ambient_std = role_value("ambient", "rep_to_rep_std_C")
            prod_spread = role_value("product", "mean_product_spread_C")
            n_reps = int(max(t["rep_id"].nunique() if not t.empty else 0, u["rep_id"].nunique() if not u.empty else 0))
            meta = setup_meta.loc[dataset] if dataset in setup_meta.index else pd.Series(dtype=object)
            rows.append(
                {
                    "zone": zone_short(zone),
                    "dataset": dataset,
                    "n_repetitions": n_reps,
                    "experimental_setup": setup_text(meta),
                    "product_repeatability_mean_std_C": prod_std,
                    "mould_T2_repeatability_std_C": mould_std,
                    "ambient_T6_repeatability_std_C": ambient_std,
                    "mean_within_mould_product_spread_C": prod_spread,
                    "primary_us_mean_curve_rmse": us_rmse,
                    "primary_us_min_curve_correlation": us_corr,
                    "plain_language_read": (
                        f"Within this setup: product repeatability std {fmt_num(prod_std)} °C; "
                        f"mould T2 std {fmt_num(mould_std)} °C; ambient T6 std {fmt_num(ambient_std)} °C; "
                        f"mean product spread inside mould {fmt_num(prod_spread)} °C; "
                        f"primary ultrasound RMSE {fmt_num(us_rmse, 4)} and min correlation {fmt_num(us_corr, 3)}."
                    ),
                }
            )
    return pd.DataFrame(rows).sort_values(["zone", "dataset"])


def make_zone_between_setup_summary(tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    temp_df = tables["Temperature Zone Metrics"].copy()
    us_df = tables["US Zone Metrics"].copy()
    rows = []

    if not temp_df.empty:
        temp_means = (
            temp_df.groupby(["zone", "dataset", "sensor_role"], dropna=False)
            .agg(mean_value=("mean_C", "mean"))
            .reset_index()
        )
        for (zone, role), group in temp_means.groupby(["zone", "sensor_role"], dropna=False):
            if group["dataset"].nunique() < 2:
                continue
            ordered = group.sort_values("mean_value")
            low = ordered.iloc[0]
            high = ordered.iloc[-1]
            rows.append(
                {
                    "zone": zone_short(zone),
                    "signal": f"{role} temperature",
                    "comparison_metric": "mean temperature across repetitions and sensors",
                    "between_setup_range": float(high["mean_value"] - low["mean_value"]),
                    "unit": "°C",
                    "lowest_setup": low["dataset"],
                    "lowest_setup_mean": float(low["mean_value"]),
                    "highest_setup": high["dataset"],
                    "highest_setup_mean": float(high["mean_value"]),
                    "plain_language_read": f"{role} temperature differs by {fmt_num(high['mean_value'] - low['mean_value'])} °C between the lowest and highest setup means.",
                }
            )

    if not us_df.empty:
        us_means = (
            us_df[us_df["role"] == "primary"]
            .groupby(["zone", "dataset", "channel"], dropna=False)
            .agg(mean_value=("mean", "mean"), curve_rmse=("normalized_curve_rmse_to_group_mean", "mean"))
            .reset_index()
        )
        for (zone, channel), group in us_means.groupby(["zone", "channel"], dropna=False):
            if group["dataset"].nunique() < 2:
                continue
            ordered = group.sort_values("mean_value")
            low = ordered.iloc[0]
            high = ordered.iloc[-1]
            rows.append(
                {
                    "zone": zone_short(zone),
                    "signal": f"{channel} ultrasound",
                    "comparison_metric": "mean primary ultrasound amplitude",
                    "between_setup_range": float(high["mean_value"] - low["mean_value"]),
                    "unit": "a.u.",
                    "lowest_setup": low["dataset"],
                    "lowest_setup_mean": float(low["mean_value"]),
                    "highest_setup": high["dataset"],
                    "highest_setup_mean": float(high["mean_value"]),
                    "plain_language_read": f"{channel} ultrasound amplitude differs by {fmt_num(high['mean_value'] - low['mean_value'], 4)} a.u. between setup means.",
                }
            )

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return out.sort_values(["zone", "signal"])


def make_condition_effect_summary(question_tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    factor_priority = [
        "Chocolate Recipe",
        "Cooling Configuration",
        "Cooling Temperature (°C)",
        "Cooling Temperature (Â°C)",
        "Cooling Speed (%)",
        "Tempering",
        "Tempering Degree",
        "zone",
    ]
    rows = []
    for sheet_name, source_label in [
        ("C Summary Param Screening", "summary landmarks"),
        ("C Effect Screening", "per-zone curves"),
    ]:
        df = question_tables.get(sheet_name, pd.DataFrame()).copy()
        if df.empty:
            continue
        df = df[df["factor"].isin(factor_priority)].copy()
        df = df[df["effect_strength"].isin(["large", "moderate", "small"])].copy()
        if df.empty:
            continue
        df = df.sort_values("eta2_effect_size", ascending=False)
        for _, row in df.head(30).iterrows():
            response = row.get("response_label") or row.get("summary_parameter") or row.get("response")
            rows.append(
                {
                    "source": source_label,
                    "factor": row["factor"].replace("Â", ""),
                    "response": response,
                    "effect_strength": row["effect_strength"],
                    "eta2_effect_size": row["eta2_effect_size"],
                    "largest_mean_difference": row["largest_mean_difference"],
                    "plain_language_read": (
                        f"{row['factor'].replace('Â', '')} shows a {row['effect_strength']} screened effect "
                        f"on {response}; largest mean contrast: {row['largest_mean_difference']}."
                    ),
                    "method_note": "Screening effect size, not a formal mixed-effects p-value.",
                }
            )
    return pd.DataFrame(rows)


def pearson_r(x: pd.Series, y: pd.Series) -> tuple[float, int]:
    work = pd.DataFrame({"x": x, "y": y}).dropna()
    if len(work) < 3 or work["x"].nunique() < 2 or work["y"].nunique() < 2:
        return np.nan, int(len(work))
    return float(work["x"].corr(work["y"])), int(len(work))


def corr_strength(r: object) -> str:
    if pd.isna(r):
        return "not enough variation"
    ar = abs(float(r))
    if ar >= 0.8:
        return "very strong"
    if ar >= 0.6:
        return "strong"
    if ar >= 0.4:
        return "moderate"
    if ar >= 0.2:
        return "weak"
    return "very weak"


def make_zone_summary_link_sheet(tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Relate zone-level curve features to the manually summarized process landmarks."""
    landmarks = tables["Parsed Landmarks"].copy()
    temp_df = tables["Temperature Zone Metrics"].copy()
    us_df = tables["US Zone Metrics"].copy()
    if landmarks.empty:
        return pd.DataFrame()

    outcomes = [
        "Viscosity Damping",
        "Cryst. Onset (rel.)",
        "Cryst. Efficiency",
        "Detachment Onset (rel.)",
        "Detachment Completion (rel.)",
        "Detachment Duration",
        "Cryst. Length",
    ]
    outcomes = [c for c in outcomes if c in landmarks.columns]
    id_cols = ["dataset", "rep_id", "sensor_code"]

    predictors = []
    if not temp_df.empty:
        temp_agg = (
            temp_df.groupby(id_cols + ["zone", "sensor_role"], dropna=False)
            .agg(
                mean_temp_C=("mean_C", "mean"),
                range_temp_C=("range_C", "mean"),
                abs_slope_temp_C_per_s=("slope_C_per_s", lambda s: s.abs().mean()),
                product_spread_C=("mean_product_spread_C", "mean"),
            )
            .reset_index()
        )
        for _, row in temp_agg.iterrows():
            for metric in ["mean_temp_C", "range_temp_C", "abs_slope_temp_C_per_s", "product_spread_C"]:
                if pd.notna(row[metric]):
                    predictors.append(
                        {
                            **{c: row[c] for c in id_cols},
                            "zone": row["zone"],
                            "predictor_group": f"{row['sensor_role']} temperature",
                            "predictor_metric": metric,
                            "predictor_value": row[metric],
                        }
                    )

    if not us_df.empty:
        us_agg = (
            us_df[us_df["role"] == "primary"]
            .groupby(id_cols + ["zone", "channel"], dropna=False)
            .agg(
                mean_us=("mean", "mean"),
                range_us=("range", "mean"),
                curve_rmse=("normalized_curve_rmse_to_group_mean", "mean"),
                curve_correlation=("correlation_to_group_mean", "mean"),
            )
            .reset_index()
        )
        for _, row in us_agg.iterrows():
            for metric in ["mean_us", "range_us", "curve_rmse", "curve_correlation"]:
                if pd.notna(row[metric]):
                    predictors.append(
                        {
                            **{c: row[c] for c in id_cols},
                            "zone": row["zone"],
                            "predictor_group": f"{row['channel']} ultrasound",
                            "predictor_metric": metric,
                            "predictor_value": row[metric],
                        }
                    )

    pred = pd.DataFrame(predictors)
    if pred.empty:
        return pred
    merged = pred.merge(landmarks[id_cols + outcomes], on=id_cols, how="left")

    rows = []
    group_cols = ["dataset", "zone", "predictor_group", "predictor_metric"]
    for keys, group in merged.groupby(group_cols, dropna=False):
        dataset, zone, predictor_group, predictor_metric = keys
        for outcome in outcomes:
            r, n = pearson_r(group["predictor_value"], group[outcome])
            rows.append(
                {
                    "scope": "within experimental setup",
                    "dataset": dataset,
                    "zone": zone_short(zone),
                    "predictor_group": predictor_group,
                    "predictor_metric": predictor_metric,
                    "summary_outcome": outcome,
                    "n_pairs": n,
                    "pearson_r": r,
                    "strength": corr_strength(r),
                    "plain_language_read": (
                        f"Within {dataset}, {zone_short(zone)}: {predictor_group} {predictor_metric} "
                        f"vs {outcome} gives r={fmt_num(r)} from {n} paired values."
                    ),
                    "caution": "Small n within a setup; use as pattern evidence, not proof.",
                }
            )

    # Across-setup link: first average repetitions inside each dataset, then compare setup means.
    across = (
        merged.groupby(["dataset", "zone", "predictor_group", "predictor_metric"], dropna=False)
        .agg(predictor_value=("predictor_value", "mean"), **{out: (out, "mean") for out in outcomes})
        .reset_index()
    )
    for keys, group in across.groupby(["zone", "predictor_group", "predictor_metric"], dropna=False):
        zone, predictor_group, predictor_metric = keys
        for outcome in outcomes:
            r, n = pearson_r(group["predictor_value"], group[outcome])
            rows.append(
                {
                    "scope": "between experimental setups",
                    "dataset": "all setup means",
                    "zone": zone_short(zone),
                    "predictor_group": predictor_group,
                    "predictor_metric": predictor_metric,
                    "summary_outcome": outcome,
                    "n_pairs": n,
                    "pearson_r": r,
                    "strength": corr_strength(r),
                    "plain_language_read": (
                        f"Between setup means, {zone_short(zone)}: {predictor_group} {predictor_metric} "
                        f"vs {outcome} gives r={fmt_num(r)} from {n} setup means."
                    ),
                    "caution": "Only a few setup means are available; this is exploratory.",
                }
            )

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out["abs_r"] = out["pearson_r"].abs()
    return out.sort_values(["scope", "zone", "abs_r"], ascending=[True, True, False]).drop(columns=["abs_r"])


def make_zone_link_highlights(zone_links: pd.DataFrame) -> pd.DataFrame:
    if zone_links.empty:
        return zone_links
    preferred_outcomes = [
        "Cryst. Onset (rel.)",
        "Cryst. Efficiency",
        "Detachment Onset (rel.)",
        "Detachment Completion (rel.)",
        "Detachment Duration",
        "Cryst. Length",
    ]
    work = zone_links[zone_links["summary_outcome"].isin(preferred_outcomes)].copy()
    work = work[work["n_pairs"] >= 4].copy()
    if work.empty:
        return pd.DataFrame()
    work["abs_r"] = work["pearson_r"].abs()
    pieces = []
    for (scope, zone), group in work.groupby(["scope", "zone"], dropna=False):
        pieces.append(group.sort_values("abs_r", ascending=False).head(5))
    out = pd.concat(pieces, ignore_index=True).sort_values(["scope", "zone", "abs_r"], ascending=[True, True, False])
    cols = [
        "scope",
        "zone",
        "predictor_group",
        "predictor_metric",
        "summary_outcome",
        "n_pairs",
        "pearson_r",
        "strength",
        "plain_language_read",
        "caution",
    ]
    return out[cols]


def make_executive_summary(clear_tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    within = clear_tables.get("Zone Within Setup", pd.DataFrame())
    between = clear_tables.get("Zone Between Setups", pd.DataFrame())
    effects = clear_tables.get("Condition Effects", pd.DataFrame())

    rows = [
        {
            "topic": "What was done",
            "summary": "The lab-trial raw curves were divided into zones from the uploaded summary landmarks, then ultrasound and temperature features were extracted per repetition, zone, and sensor/channel.",
            "where_to_check": "Detailed calculations: US Zone Metrics and Temperature Zone Metrics",
        },
        {
            "topic": "Question A - within setup",
            "summary": "For each experimental setup and zone, repetitions are compared using product, mould, and ambient temperature repeatability plus primary ultrasound shape similarity.",
            "where_to_check": "Zone Within Setup",
        },
        {
            "topic": "Question B - between setups",
            "summary": "For each zone, setup means are compared to show which setup has the lowest and highest temperature or ultrasound level.",
            "where_to_check": "Zone Between Setups",
        },
        {
            "topic": "Question C - relation to experimental conditions",
            "summary": "Experimental factors such as chocolate recipe, cooling configuration, cooling temperature/speed, tempering, and zone are screened against the extracted features.",
            "where_to_check": "Condition Effects",
        },
        {
            "topic": "Zone behavior vs summary outcomes",
            "summary": "Zone features are now also correlated with summary outcomes, for example whether zone 2 cooling temperature or ultrasound behavior aligns with crystallization start.",
            "where_to_check": "Zone Summary Links",
        },
        {
            "topic": "Caution",
            "summary": "The condition-effect sheet is effect-size screening, not a formal mixed-effects p-value model. It is meant to point to likely drivers before formal modelling.",
            "where_to_check": "Condition Effects",
        },
        {
            "topic": "Zone 1",
            "summary": "Zone 1 is not calculated yet because the uploaded summary does not provide a reliable deposition-to-cooling-start boundary for every repetition.",
            "where_to_check": "Read Me / Zone Boundaries",
        },
    ]

    if not within.empty and "product_repeatability_mean_std_C" in within.columns:
        valid = within.dropna(subset=["product_repeatability_mean_std_C"])
        if not valid.empty:
            best = valid.sort_values("product_repeatability_mean_std_C").iloc[0]
            worst = valid.sort_values("product_repeatability_mean_std_C").iloc[-1]
            rows.append(
                {
                    "topic": "Within-setup temperature highlight",
                    "summary": (
                        f"Best product-temperature repeatability in the current extraction: {best['dataset']} / {best['zone']} "
                        f"({fmt_num(best['product_repeatability_mean_std_C'])} °C std). Highest product-temperature variation: "
                        f"{worst['dataset']} / {worst['zone']} ({fmt_num(worst['product_repeatability_mean_std_C'])} °C std)."
                    ),
                    "where_to_check": "Zone Within Setup",
                }
            )

    if not between.empty and "between_setup_range" in between.columns:
        valid = between.dropna(subset=["between_setup_range"]).sort_values("between_setup_range", ascending=False)
        if not valid.empty:
            top = valid.iloc[0]
            rows.append(
                {
                    "topic": "Between-setup highlight",
                    "summary": (
                        f"Largest between-setup spread in the front summary: {top['signal']} in {top['zone']} "
                        f"({fmt_num(top['between_setup_range'])} {top['unit']}; {top['highest_setup']} vs {top['lowest_setup']})."
                    ),
                    "where_to_check": "Zone Between Setups",
                }
            )

    if not effects.empty and "eta2_effect_size" in effects.columns:
        valid = effects.dropna(subset=["eta2_effect_size"]).sort_values("eta2_effect_size", ascending=False)
        if not valid.empty:
            top = valid.iloc[0]
            rows.append(
                {
                    "topic": "Condition-effect highlight",
                    "summary": (
                        f"Strongest screened setup/zone relation: {top['factor']} vs {top['response']} "
                        f"({top['effect_strength']}, eta² {fmt_num(top['eta2_effect_size'])})."
                    ),
                    "where_to_check": "Condition Effects",
                }
            )

    return pd.DataFrame(rows)


MAIN_FACTORS = [
    "Chocolate Recipe",
    "Cooling Configuration",
    "Cooling Temperature (°C)",
    "Cooling Speed (%)",
    "Tempering",
    "Tempering Degree",
    "Trials",
]


SUMMARY_PARAMETERS = [
    "Viscosity Damping",
    "Cryst. Onset (rel.)",
    "Cryst. Efficiency",
    "Detachment Onset (rel.)",
    "Detachment Completion (rel.)",
    "Detachment Duration",
    "Cryst. Length",
]


def summary_long_table(landmarks: pd.DataFrame) -> pd.DataFrame:
    value_vars = [c for c in SUMMARY_PARAMETERS if c in landmarks.columns]
    id_vars = ["dataset", "rep_id", "sensor_code", *[c for c in MAIN_FACTORS if c in landmarks.columns]]
    return landmarks.melt(
        id_vars=id_vars,
        value_vars=value_vars,
        var_name="summary_parameter",
        value_name="summary_value",
    ).dropna(subset=["summary_value"])


def factor_effects_for_summary(tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    landmarks = tables["Parsed Landmarks"].copy()
    long = summary_long_table(landmarks)
    rows = []
    for parameter, group in long.groupby("summary_parameter", dropna=False):
        for factor in [f for f in MAIN_FACTORS if f in group.columns]:
            eta2, n_levels, strength = effect_size_eta2(group, "summary_value", factor)
            if n_levels < 2:
                continue
            rows.append(
                {
                    "summary_parameter": parameter,
                    "independent_variable": factor,
                    "n_factor_levels": n_levels,
                    "n_measurements": int(group[["summary_value", factor]].dropna().shape[0]),
                    "eta2_effect_size": eta2,
                    "effect_strength": strength,
                    "largest_mean_difference": best_group_difference(group, "summary_value", factor),
                    "interpretation": (
                        f"{factor} has a {strength} screened association with {parameter}. "
                        f"Largest mean contrast: {best_group_difference(group, 'summary_value', factor)}."
                    ),
                }
            )
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return out.sort_values(["summary_parameter", "eta2_effect_size"], ascending=[True, False])


def repeatability_cv_summary(tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    landmarks = tables["Parsed Landmarks"].copy()
    long = summary_long_table(landmarks)
    rows = []
    for (dataset, sensor_code, parameter), group in long.groupby(["dataset", "sensor_code", "summary_parameter"], dropna=False):
        vals = pd.to_numeric(group["summary_value"], errors="coerce").dropna()
        if vals.empty:
            continue
        mean = float(vals.mean())
        std = float(vals.std()) if len(vals) > 1 else np.nan
        cv = float(std / abs(mean) * 100) if len(vals) > 1 and mean != 0 and pd.notna(std) else np.nan
        if pd.isna(cv):
            flag = "not enough repetitions"
        elif cv < 5:
            flag = "high repeatability"
        elif cv < 15:
            flag = "moderate repeatability"
        else:
            flag = "low repeatability / inspect"
        setup = group.iloc[0]
        rows.append(
            {
                "dataset": dataset,
                "sensor_code": sensor_code,
                "summary_parameter": parameter,
                "n_repetitions": int(vals.shape[0]),
                "mean": mean,
                "std": std,
                "cv_percent": cv,
                "min": float(vals.min()),
                "max": float(vals.max()),
                "repeatability_flag": flag,
                "experimental_setup": setup_text(setup),
                "interpretation": (
                    f"{parameter} in {dataset}, sensor {sensor_code}: CV {fmt_num(cv, 1)}%; {flag}."
                ),
            }
        )
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return out.sort_values(["dataset", "summary_parameter", "sensor_code"])


def temp_predictor_table(tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    temp = tables["Temperature Zone Metrics"].copy()
    if temp.empty:
        return pd.DataFrame()
    return (
        temp.groupby(["dataset", "rep_id", "sensor_code", "zone", "sensor_role"], dropna=False)
        .agg(
            mean_temp_C=("mean_C", "mean"),
            range_temp_C=("range_C", "mean"),
            abs_slope_temp_C_per_s=("slope_C_per_s", lambda s: s.abs().mean()),
            product_spread_C=("mean_product_spread_C", "mean"),
        )
        .reset_index()
    )


def targeted_temp_outcome_links(tables: dict[str, pd.DataFrame], repeatability: pd.DataFrame | None = None) -> pd.DataFrame:
    landmarks = tables["Parsed Landmarks"].copy()
    temp_pred = temp_predictor_table(tables)
    if landmarks.empty or temp_pred.empty:
        return pd.DataFrame()

    if repeatability is None or repeatability.empty:
        repeatability = repeatability_cv_summary(tables)
    poor_repeatability = repeatability[
        pd.to_numeric(repeatability.get("cv_percent"), errors="coerce").gt(20)
    ].copy()
    if poor_repeatability.empty:
        return pd.DataFrame(
            [
                {
                    "question": "Can temperature behavior explain summary variation?",
                    "scope": "only poor-repeatability cases",
                    "dataset": "",
                    "zone": "",
                    "temperature_group": "",
                    "temperature_metric": "",
                    "summary_outcome": "",
                    "n_pairs": 0,
                    "pearson_r": np.nan,
                    "strength": "",
                    "interpretation": "No summary parameter had CV > 20%, so no targeted temperature-explanation analysis was needed.",
                    "caution": "Threshold used: CV > 20%.",
                }
            ]
        )

    outcomes = [c for c in SUMMARY_PARAMETERS if c in landmarks.columns]
    merged = temp_pred.merge(landmarks[["dataset", "rep_id", "sensor_code", *outcomes]], on=["dataset", "rep_id", "sensor_code"], how="left")
    target_map = {
        "Cryst. Onset (rel.)": ["zone_2_cooling_to_crystallization_start"],
        "Cryst. Efficiency": ["zone_2_cooling_to_crystallization_start"],
        "Cryst. Length": ["zone_2_cooling_to_crystallization_start", "zone_3_crystallization_to_detachment_onset"],
        "Detachment Onset (rel.)": ["zone_3_crystallization_to_detachment_onset", "zone_4_detachment_onset_to_detachment_completion"],
        "Detachment Completion (rel.)": ["zone_4_detachment_onset_to_detachment_completion"],
        "Detachment Duration": ["zone_4_detachment_onset_to_detachment_completion"],
    }
    metrics = ["mean_temp_C", "range_temp_C", "abs_slope_temp_C_per_s", "product_spread_C"]
    rows = []
    poor_cases = poor_repeatability[["dataset", "sensor_code", "summary_parameter", "cv_percent", "repeatability_flag"]].drop_duplicates()
    for _, poor in poor_cases.iterrows():
        outcome = poor["summary_parameter"]
        zones = target_map.get(outcome)
        if not zones:
            continue
        if outcome not in merged.columns:
            continue
        work = merged[
            (merged["dataset"] == poor["dataset"])
            & (merged["sensor_code"].astype(str) == str(poor["sensor_code"]))
            & (merged["zone"].isin(zones))
        ].copy()
        if work.empty:
            continue
        for scope_name, group_keys in [
            ("within experimental setup", ["dataset", "zone", "sensor_role"]),
        ]:
            source = work
            for keys, group in source.groupby(group_keys, dropna=False):
                dataset, zone, role = keys
                for metric in metrics:
                    if metric == "product_spread_C" and role != "product":
                        continue
                    r, n = pearson_r(group[metric], group[outcome])
                    rows.append(
                        {
                            "question": "Can temperature explain poor repeatability?",
                            "scope": scope_name,
                            "dataset": dataset,
                            "zone": zone_short(zone),
                            "temperature_group": f"{role} temperature",
                            "temperature_metric": metric,
                            "summary_outcome": outcome,
                            "poor_repeatability_cv_percent": poor["cv_percent"],
                            "repeatability_flag": poor["repeatability_flag"],
                            "n_pairs": n,
                            "pearson_r": r,
                            "strength": corr_strength(r),
                            "interpretation": (
                                f"{dataset}, sensor {poor['sensor_code']}: {outcome} had CV {fmt_num(poor['cv_percent'], 1)}%. "
                                f"{zone_short(zone)} {role} {metric} vs {outcome}: r={fmt_num(r)} ({corr_strength(r)}), n={n}."
                            ),
                            "caution": "Only cases with CV > 20% are shown; correlations are exploratory and often based on few repetitions.",
                        }
                    )
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out["abs_r"] = out["pearson_r"].abs()
    out = out[out["n_pairs"] >= 3].copy()
    pieces = []
    for (scope, outcome), group in out.groupby(["scope", "summary_outcome"], dropna=False):
        pieces.append(group.sort_values("abs_r", ascending=False).head(8))
    if not pieces:
        return pd.DataFrame(columns=[c for c in out.columns if c != "abs_r"])
    return pd.concat(pieces, ignore_index=True).sort_values(["scope", "summary_outcome", "abs_r"], ascending=[True, True, False]).drop(columns="abs_r")


def us_pattern_analysis(tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    us = tables["US Zone Metrics"].copy()
    temp = temp_predictor_table(tables)
    if us.empty:
        return pd.DataFrame()
    primary = us[us["role"] == "primary"].copy()
    pattern_rows = []
    for (dataset, sensor_code, zone, channel), group in primary.groupby(["dataset", "sensor_code", "zone", "channel"], dropna=False):
        n = group["rep_id"].nunique()
        mean_rmse = group["normalized_curve_rmse_to_group_mean"].mean()
        max_rmse = group["normalized_curve_rmse_to_group_mean"].max()
        min_corr = group["correlation_to_group_mean"].min()
        mean_amp = group["mean"].mean()
        std_amp = group["mean"].std()
        cv_amp = float(std_amp / abs(mean_amp) * 100) if pd.notna(std_amp) and mean_amp != 0 else np.nan
        if pd.isna(min_corr):
            flag = "not enough data"
        elif min_corr >= 0.85 and (pd.isna(cv_amp) or cv_amp < 10):
            flag = "repeatable US pattern"
        elif min_corr >= 0.6:
            flag = "moderately repeatable US pattern"
        else:
            flag = "different US pattern / inspect"
        pattern_rows.append(
            {
                "analysis_type": "US pattern repeatability",
                "dataset": dataset,
                "sensor_code": sensor_code,
                "zone": zone_short(zone),
                "channel": channel,
                "n_repetitions": int(n),
                "mean_curve_rmse": mean_rmse,
                "max_curve_rmse": max_rmse,
                "min_curve_correlation": min_corr,
                "amplitude_cv_percent": cv_amp,
                "pattern_flag": flag,
                "temperature_link": "",
                "interpretation": (
                    f"{dataset}, sensor {sensor_code}, {zone_short(zone)} {channel}: min curve correlation {fmt_num(min_corr)}, "
                    f"amplitude CV {fmt_num(cv_amp, 1)}%; {flag}."
                ),
            }
        )

    code_means = (
        primary.groupby(["dataset", "sensor_code", "zone", "channel"], dropna=False)
        .agg(
            mean_curve_rmse=("normalized_curve_rmse_to_group_mean", "mean"),
            mean_curve_correlation=("correlation_to_group_mean", "mean"),
            mean_amplitude=("mean", "mean"),
        )
        .reset_index()
    )
    for (dataset, zone, channel), group in code_means.groupby(["dataset", "zone", "channel"], dropna=False):
        codes = sorted([str(c) for c in group["sensor_code"].dropna().unique()])
        if not {"11", "22"}.issubset(set(codes)):
            continue
        g11 = group[group["sensor_code"].astype(str) == "11"].iloc[0]
        g22 = group[group["sensor_code"].astype(str) == "22"].iloc[0]
        pattern_rows.append(
            {
                "analysis_type": "US sensor-code comparison",
                "dataset": dataset,
                "sensor_code": "11 vs 22",
                "zone": zone_short(zone),
                "channel": channel,
                "n_repetitions": int(primary[(primary["dataset"] == dataset) & (primary["zone"] == zone) & (primary["channel"] == channel)]["rep_id"].nunique()),
                "mean_curve_rmse": np.nan,
                "max_curve_rmse": np.nan,
                "min_curve_correlation": np.nan,
                "amplitude_cv_percent": np.nan,
                "pattern_flag": "sensor comparison",
                "temperature_link": "",
                "interpretation": (
                    f"{dataset} {zone_short(zone)} {channel}: sensor 11 vs 22 mean amplitude difference "
                    f"{fmt_num(g11['mean_amplitude'] - g22['mean_amplitude'], 4)} a.u.; "
                    f"curve correlation difference {fmt_num(g11['mean_curve_correlation'] - g22['mean_curve_correlation'])}; "
                    f"RMSE difference {fmt_num(g11['mean_curve_rmse'] - g22['mean_curve_rmse'], 4)}."
                ),
            }
        )

    # Specifically connect detachment US pattern variability to product temperature distribution.
    det_us = primary[primary["zone"] == "zone_4_detachment_onset_to_detachment_completion"].copy()
    det_temp = temp[(temp["zone"] == "zone_4_detachment_onset_to_detachment_completion") & (temp["sensor_role"] == "product")].copy()
    if not det_us.empty and not det_temp.empty:
        merged = det_us.merge(
            det_temp[["dataset", "rep_id", "sensor_code", "product_spread_C", "mean_temp_C", "range_temp_C"]],
            on=["dataset", "rep_id", "sensor_code"],
            how="left",
        )
        for (dataset, sensor_code, channel), group in merged.groupby(["dataset", "sensor_code", "channel"], dropna=False):
            for predictor in ["product_spread_C", "mean_temp_C", "range_temp_C"]:
                for us_metric in ["normalized_curve_rmse_to_group_mean", "correlation_to_group_mean", "mean"]:
                    r, n = pearson_r(group[predictor], group[us_metric])
                    pattern_rows.append(
                        {
                            "analysis_type": "Detachment US pattern explained by temperature",
                            "dataset": dataset,
                            "sensor_code": sensor_code,
                            "zone": "Z4 detachment onset -> completion",
                            "channel": channel,
                            "n_repetitions": int(n),
                            "mean_curve_rmse": np.nan,
                            "max_curve_rmse": np.nan,
                            "min_curve_correlation": np.nan,
                            "amplitude_cv_percent": np.nan,
                            "pattern_flag": corr_strength(r),
                            "temperature_link": f"{predictor} vs {us_metric}",
                            "interpretation": (
                                f"During detachment for sensor {sensor_code}, {predictor} vs {channel} {us_metric}: "
                                f"r={fmt_num(r)} ({corr_strength(r)}), n={n}."
                            ),
                        }
                    )

    out = pd.DataFrame(pattern_rows)
    if out.empty:
        return out
    link_mask = out["analysis_type"].eq("Detachment US pattern explained by temperature")
    repeat = out[~link_mask].copy()
    links = out[link_mask].copy()
    if not links.empty:
        links["abs_r"] = links["interpretation"].str.extract(r"r=([-0-9.]+)").astype(float).abs()
        kept = []
        for (dataset, sensor_code, channel), group in links.groupby(["dataset", "sensor_code", "channel"], dropna=False):
            kept.append(group.sort_values("abs_r", ascending=False).head(5))
        links = pd.concat(kept, ignore_index=True).drop(columns="abs_r") if kept else links.drop(columns="abs_r")
    return pd.concat([repeat, links], ignore_index=True).sort_values(["analysis_type", "dataset", "sensor_code", "zone", "channel"])


def smooth_series(s: pd.Series, window: int = 17) -> pd.Series:
    return s.rolling(window, center=True, min_periods=max(3, window // 4)).median()


def value_at(df: pd.DataFrame, col: str, t: float) -> float:
    valid = df[[TIME_COL, col]].dropna()
    if valid.empty or pd.isna(t):
        return np.nan
    idx = (valid[TIME_COL] - t).abs().idxmin()
    return float(valid.loc[idx, col])


def t6_rise_boundary(df: pd.DataFrame, after: float, rise_C: float = 1.0, search_end: float | None = None) -> float | None:
    valid = df[[TIME_COL, "T6"]].dropna().sort_values(TIME_COL).copy()
    if valid.empty:
        return None
    if search_end is None:
        search_end = float(valid[TIME_COL].max())
    valid = valid[(valid[TIME_COL] >= after) & (valid[TIME_COL] <= search_end)].copy()
    if len(valid) < 8:
        return None
    valid["smooth_t6"] = smooth_series(valid["T6"], 21)
    low_idx = valid["smooth_t6"].idxmin()
    low_time = float(valid.loc[low_idx, TIME_COL])
    low_value = float(valid.loc[low_idx, "smooth_t6"])
    hit = valid[(valid[TIME_COL] > low_time) & (valid["smooth_t6"] >= low_value + rise_C)]
    return None if hit.empty else float(hit.iloc[0][TIME_COL])


def first_crossing_after(df: pd.DataFrame, col: str, threshold: float, after: float, direction: str) -> float | None:
    valid = df[[TIME_COL, col]].dropna().sort_values(TIME_COL)
    valid = valid[valid[TIME_COL] >= after]
    hit = valid[valid[col] <= threshold] if direction == "below" else valid[valid[col] >= threshold]
    return None if hit.empty else float(hit.iloc[0][TIME_COL])


def detect_lab_zones(tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    raw_cache: dict[str, pd.DataFrame | None] = {}
    ref_df = read_raw_sheet("at2455_a_test")
    ref_t6_b1 = value_at(ref_df, "T6", 990.0) if ref_df is not None else np.nan
    ref_t6_b3 = value_at(ref_df, "T6", 2195.0) if ref_df is not None else np.nan
    for sheet in available_raw_sheets():
        rep_id = norm_id(sheet)
        meta = PROCESS_START_TIMES.get(rep_id)
        if meta is None:
            continue
        df = raw_cache.get(sheet)
        if df is None:
            df = read_raw_sheet(sheet)
            raw_cache[sheet] = df
        if df is None or df.empty:
            continue
        dep = meta["deposition_s"]
        cool = meta["cooling_start_s"]
        profile = meta["zone_profile"]
        tmax = float(df[TIME_COL].max())
        if profile == "AAK_cooling_tunnel":
            if sheet == "at2455_a_test":
                b1 = 990.0
                b2 = t6_rise_boundary(df, b1, 1.0, 2195.0) or 1540.0
                b3 = 2195.0
                basis = "AAK cooling tunnel reference; zone 2 to 3 verified by T6 rise >1 C"
            else:
                b1 = first_crossing_after(df, "T6", ref_t6_b1, cool + 100.0, "below") or min(cool + 780.0, tmax)
                b2 = t6_rise_boundary(df, b1 + 60.0, 1.0) or min(max(b1 + 400.0, cool + 1290.0), tmax)
                tmp = df[[TIME_COL, "T6"]].copy()
                tmp["product_mean"] = df[PRODUCT_TEMPS].mean(axis=1)
                tmp["smooth_product"] = smooth_series(tmp["product_mean"], 21)
                tmp["slope_product"] = tmp["smooth_product"].diff() / tmp[TIME_COL].diff()
                after_b2 = tmp[tmp[TIME_COL] > b2 + 250.0]
                warm = after_b2[(after_b2["T6"] >= ref_t6_b3) & (after_b2["slope_product"] > 0.001)]
                if warm.empty:
                    warm = after_b2.sort_values("slope_product", ascending=False).head(1)
                b3 = float(warm.iloc[0][TIME_COL]) if not warm.empty else min(b2 + 650.0, tmax)
                basis = "AAK cooling tunnel pattern; T6 +1 C rule and warming transition"
            zones = [
                ("zone_1_deposition_to_cooling_start", dep, cool, "user-provided deposition/cooling start"),
                ("zone_2_cooling_t6_drop", cool, b1, basis),
                ("zone_3_t6_rise_after_low", b1, b2, basis),
                ("zone_4_until_demoulding", b2, b3, basis),
                ("demoulding_after_warming", b3, tmax, "after detected warming/demoulding transition"),
            ]
        else:
            ref_meta = PROCESS_START_TIMES[LAB_DEVICE_REFERENCE["reference_rep_id"]]
            zone_2_duration = LAB_DEVICE_REFERENCE["zone_2_end_s"] - ref_meta["cooling_start_s"]
            zone_3_duration = LAB_DEVICE_REFERENCE["demoulding_start_s"] - LAB_DEVICE_REFERENCE["zone_2_end_s"]
            zone_2_end = min(cool + zone_2_duration, tmax)
            demould = min(zone_2_end + zone_3_duration, tmax)
            zones = [
                ("zone_1_deposition_to_cooling_start", dep, cool, "user-provided deposition/cooling start"),
                ("zone_2_cooling_simulator", cool, zone_2_end, "lab-device cooling simulator: main cooling until first temperature increase pattern"),
                ("zone_3_cooling_simulator", zone_2_end, demould, "lab-device cooling simulator: intermediate cooling after first increase until strong warming"),
                ("demoulding_after_warming", demould, tmax, "demoulding only after later strong temperature increase"),
            ]
        for zone, start, end, basis in zones:
            rows.append(
                {
                    "rep_id": rep_id,
                    "raw_sheet": sheet,
                    "zone_profile": profile,
                    "zone": zone,
                    "start_s": start,
                    "end_s": end,
                    "duration_s": end - start if pd.notna(start) and pd.notna(end) else np.nan,
                    "basis": basis,
                    "T6_start_C": value_at(df, "T6", start),
                    "T6_end_C": value_at(df, "T6", end),
                    "product_mean_start_C": float(df.loc[(df[TIME_COL] - start).abs().idxmin(), PRODUCT_TEMPS].mean()) if pd.notna(start) else np.nan,
                    "product_mean_end_C": float(df.loc[(df[TIME_COL] - end).abs().idxmin(), PRODUCT_TEMPS].mean()) if pd.notna(end) else np.nan,
                }
            )
    return pd.DataFrame(rows)


def change_scores(t: np.ndarray, y: np.ndarray, window_points: int) -> np.ndarray:
    scores = np.full(len(y), np.nan)
    for i in range(window_points, len(y) - window_points):
        left = y[i - window_points : i]
        right = y[i : i + window_points]
        denom = np.nanstd(np.r_[left, right])
        if not np.isfinite(denom) or denom <= 1e-9:
            denom = 1e-9
        scores[i] = (np.nanmean(right) - np.nanmean(left)) / denom
    return scores


def detect_change_points(seg: pd.DataFrame, channel: str) -> list[dict[str, float]]:
    valid = seg[[TIME_COL, channel]].dropna().sort_values(TIME_COL)
    if len(valid) < 12:
        return []
    t = valid[TIME_COL].to_numpy(float)
    y = valid[channel].rolling(5, center=True, min_periods=2).median().to_numpy(float)
    dt = np.nanmedian(np.diff(t))
    window_points = max(4, int(round(45.0 / max(dt, 1.0))))
    scores = change_scores(t, y, window_points)
    positive = scores[np.isfinite(scores) & (scores > 0)]
    threshold = np.nanpercentile(positive, 80) if len(positive) else np.inf
    candidates = []
    for idx in np.argsort(np.nan_to_num(scores, nan=-np.inf))[::-1]:
        if not np.isfinite(scores[idx]) or scores[idx] <= 0 or scores[idx] < max(0.8, threshold):
            continue
        tt = float(t[idx])
        if any(abs(tt - c["change_point_s"]) < 60.0 for c in candidates):
            continue
        before = float(np.nanmedian(y[max(0, idx - window_points) : idx]))
        after = float(np.nanmedian(y[idx : min(len(y), idx + window_points)]))
        delta = after - before
        if delta < 0.025:
            continue
        candidates.append(
            {
                "change_point_s": tt,
                "change_score": float(scores[idx]),
                "local_before_median": before,
                "local_after_median": after,
                "local_delta_after_minus_before": delta,
            }
        )
        if len(candidates) >= 30:
            break
    return sorted(candidates, key=lambda c: c["change_point_s"])


def median_window(df: pd.DataFrame, channel: str, start: float, end: float) -> float:
    w = df[(df[TIME_COL] >= start) & (df[TIME_COL] <= end)][channel].dropna()
    if w.empty:
        idx = (df[TIME_COL] - end).abs().idxmin()
        return float(df.loc[idx, channel])
    return float(w.median())


def us_level_near(seg: pd.DataFrame, channel: str, t0: float, width: float = 5.0) -> float:
    w = seg[(seg[TIME_COL] >= t0) & (seg[TIME_COL] <= t0 + width)][channel].dropna()
    return np.nan if w.empty else float(w.median())


def lab_detachment_analysis(tables: dict[str, pd.DataFrame]) -> tuple[pd.DataFrame, pd.DataFrame]:
    landmarks = tables["Parsed Landmarks"].copy()
    if landmarks.empty:
        return pd.DataFrame(), pd.DataFrame()
    zones = detect_lab_zones(tables)
    zone_end = zones[zones["zone"].isin(["zone_4_until_demoulding", "zone_3_cooling_simulator"])].copy()
    zone_end = zone_end.sort_values("end_s").drop_duplicates("rep_id", keep="last").set_index("rep_id")
    cp_rows = []
    summary_rows = []
    raw_cache: dict[str, pd.DataFrame | None] = {}
    for _, lm in landmarks.iterrows():
        rep_id = lm["rep_id"]
        meta = PROCESS_START_TIMES.get(rep_id)
        if meta is None or "Detachment Onset (rel.)" not in lm or pd.isna(lm["Detachment Onset (rel.)"]):
            continue
        sheet = "at" + rep_id
        df = raw_cache.get(sheet)
        if df is None:
            df = read_raw_sheet(sheet)
            raw_cache[sheet] = df
        if df is None or df.empty:
            continue
        onset = float(meta["cooling_start_s"] + lm["Detachment Onset (rel.)"])
        if rep_id in zone_end.index:
            window_end = float(zone_end.loc[rep_id, "end_s"])
        else:
            window_end = float(df[TIME_COL].max())
        seg = df[(df[TIME_COL] >= onset) & (df[TIME_COL] <= window_end)].copy()
        if seg.empty:
            continue
        channel = SENSOR_CODE_TO_PRIMARY_US.get(str(lm["sensor_code"]))
        if channel is None:
            continue
        for channel in [channel]:
            reference = median_window(df, channel, meta["deposition_s"] - 5.0, meta["deposition_s"])
            threshold = reference - 0.10 * abs(reference)
            cps = detect_change_points(seg, channel)
            selected = None
            count_until = 0
            for idx, cp in enumerate(cps, 1):
                level = us_level_near(seg, channel, cp["change_point_s"], 5.0)
                met = bool(pd.notna(level) and level >= threshold)
                if selected is None:
                    count_until = idx
                cp_rows.append(
                    {
                        "rep_id": rep_id,
                        "sensor_code": lm["sensor_code"],
                        "channel": channel,
                        "change_point_index": idx,
                        **cp,
                        "reference_window_s": f"{meta['deposition_s'] - 5.0:.1f}-{meta['deposition_s']:.1f}",
                        "reference_us_level": reference,
                        "threshold_10pct_lower": threshold,
                        "us_level_at_change_point": level,
                        "threshold_met": met,
                        "note": "first passing change point used as detachment offset" if met and selected is None else ("later passing change point" if met else ""),
                    }
                )
                if met and selected is None:
                    selected = {**cp, "level": level, "idx": idx}
            if selected is not None:
                offset = selected["change_point_s"]
                status = "full_detachment_for_channel"
                selected_level = selected["level"]
                count_until = selected["idx"]
            elif cps:
                offset = np.nan
                status = "partial_detachment_no_change_point_met_threshold"
                selected_level = us_level_near(seg, channel, cps[-1]["change_point_s"], 5.0)
                count_until = len(cps)
            else:
                offset = np.nan
                status = "no_positive_change_point_detected"
                selected_level = np.nan
                count_until = 0
            pattern = (
                f"{len(cps)} positive upward change points from detachment onset {onset:.1f}s to cooling-window end {window_end:.1f}s. "
                f"{'First passing point at ' + f'{offset:.1f}s' if pd.notna(offset) else 'No change point reached the pre-deposition -10% threshold'}."
            )
            summary_rows.append(
                {
                    "dataset": lm.get("dataset", ""),
                    "rep_id": rep_id,
                    "sensor_code": lm["sensor_code"],
                    "channel": channel,
                    "zone_profile": meta["zone_profile"],
                    "deposition_s": meta["deposition_s"],
                    "cooling_start_s": meta["cooling_start_s"],
                    "detachment_onset_absolute_s": onset,
                    "analysis_window_end_s": window_end,
                    "detachment_offset_s": offset,
                    "detachment_offset_status": status,
                    "reference_us_level": reference,
                    "threshold_10pct_lower": threshold,
                    "us_level_at_decision": selected_level,
                    "change_points_detected_total": len(cps),
                    "change_points_until_detachment_or_last": count_until,
                    "pattern_description": pattern,
                    "method": "positive/upward local mean-shift change points; full detachment per channel if US at change point >= 10%-below-pre-deposition-reference threshold",
                }
            )
    return pd.DataFrame(summary_rows), pd.DataFrame(cp_rows)


def detachment_change_point_summary(detachment_offset: pd.DataFrame) -> pd.DataFrame:
    if detachment_offset.empty:
        return pd.DataFrame()
    cols = [
        "dataset",
        "rep_id",
        "sensor_code",
        "channel",
        "detachment_onset_absolute_s",
        "analysis_window_end_s",
        "detachment_offset_status",
        "detachment_offset_s",
        "reference_us_level",
        "threshold_10pct_lower",
        "us_level_at_decision",
        "change_points_detected_total",
        "change_points_until_detachment_or_last",
        "pattern_description",
    ]
    out = detachment_offset[[c for c in cols if c in detachment_offset.columns]].copy()
    out["sensor_channel_assignment"] = out["sensor_code"].astype(str).map(SENSOR_CODE_TO_PRIMARY_US)
    out["change_point_count_interpretation"] = out.apply(
        lambda r: (
            f"{int(r['change_points_until_detachment_or_last'])} change points before full detachment offset"
            if r["detachment_offset_status"] == "full_detachment_for_channel"
            else f"{int(r['change_points_until_detachment_or_last'])} change points checked; no full detachment offset"
        ),
        axis=1,
    )
    return out


def get_fonts() -> tuple[ImageFont.ImageFont, ImageFont.ImageFont]:
    try:
        return ImageFont.truetype("arial.ttf", 16), ImageFont.truetype("arial.ttf", 12)
    except Exception:
        return ImageFont.load_default(), ImageFont.load_default()


def plot_series(draw: ImageDraw.ImageDraw, x, y, box, color: str, width: int = 2, ymin=None, ymax=None) -> None:
    x0, y0, x1, y1 = box
    vals = pd.DataFrame({"x": x, "y": y}).dropna()
    if vals.empty:
        return
    xmin, xmax = vals["x"].min(), vals["x"].max()
    ymin = vals["y"].min() if ymin is None else ymin
    ymax = vals["y"].max() if ymax is None else ymax
    if xmax == xmin or ymax == ymin:
        return
    pts = []
    for _, r in vals.iterrows():
        px = x0 + (r["x"] - xmin) / (xmax - xmin) * (x1 - x0)
        py = y1 - (r["y"] - ymin) / (ymax - ymin) * (y1 - y0)
        pts.append((px, py))
    if len(pts) > 1:
        draw.line(pts, fill=color, width=width)


def draw_axes(
    draw: ImageDraw.ImageDraw,
    box: tuple[int, int, int, int],
    x_min: float,
    x_max: float,
    y_min: float,
    y_max: float,
    x_label: str,
    y_label: str,
    font: ImageFont.ImageFont,
    x_ticks: int = 6,
    y_ticks: int = 5,
) -> None:
    x0, y0, x1, y1 = box
    draw.rectangle(box, outline="#444444", width=1)
    if x_max == x_min or y_max == y_min:
        return
    for i in range(x_ticks):
        frac = i / max(x_ticks - 1, 1)
        x = x0 + frac * (x1 - x0)
        val = x_min + frac * (x_max - x_min)
        draw.line((x, y1, x, y1 + 6), fill="#444444", width=1)
        draw.text((x - 18, y1 + 8), f"{val:.0f}", fill="#333333", font=font)
    for i in range(y_ticks):
        frac = i / max(y_ticks - 1, 1)
        y = y1 - frac * (y1 - y0)
        val = y_min + frac * (y_max - y_min)
        draw.line((x0 - 6, y, x0, y), fill="#444444", width=1)
        draw.text((x0 - 58, y - 7), f"{val:.2f}", fill="#333333", font=font)
    draw.text(((x0 + x1) / 2 - 55, y1 + 32), x_label, fill="#111111", font=font)
    # PIL text rotation is awkward; keep the y label near the axis, readable.
    draw.text((x0 - 70, y0 - 22), y_label, fill="#111111", font=font)


def make_lab_zone_us_figures(zones: pd.DataFrame, detachment: pd.DataFrame) -> list[Path]:
    fig_dir = OUTPUT_DIR / "figures_lab"
    fig_dir.mkdir(parents=True, exist_ok=True)
    paths = []
    font, small = get_fonts()
    colors = {"T1": "#cc3311", "T3": "#0077bb", "T4": "#009988", "T5": "#ee7733", "T2": "#aa4499", "T6": "#000000", "Rx1Tx1": "#3344cc", "Rx2Tx2": "#cc3377"}
    for sheet in available_raw_sheets():
        df = read_raw_sheet(sheet)
        if df is None or df.empty:
            continue
        rep_id = norm_id(sheet)
        img = Image.new("RGB", (1450, 860), "white")
        draw = ImageDraw.Draw(img)
        draw.text((40, 18), f"{rep_id}: zones, detachment onset, and full detachment offsets", fill="black", font=font)
        temp_box = (85, 75, 1360, 440)
        us_box = (85, 515, 1360, 790)
        for box, label in [(temp_box, "Temperatures T1/T3/T4/T5, T2 mould, T6 ambient"), (us_box, "Primary ultrasound Rx1Tx1 / Rx2Tx2")]:
            draw.rectangle(box, outline="#444444", width=1)
            draw.text((box[0], box[1] - 22), label, fill="black", font=small)
        tmin, tmax = float(df[TIME_COL].min()), float(df[TIME_COL].max())
        temp_cols = [c for c in [*PRODUCT_TEMPS, "T2", "T6"] if c in df.columns]
        temp_vals = df[temp_cols].stack().dropna()
        temp_ymin, temp_ymax = float(temp_vals.min()) - 0.5, float(temp_vals.max()) + 0.5
        us_vals = df[[c for c in PRIMARY_US if c in df.columns]].stack().dropna()
        us_ymin, us_ymax = float(us_vals.min()) - 0.03, float(us_vals.max()) + 0.03
        rz = zones[zones["raw_sheet"] == sheet].copy()
        shade_colors = ["#e8f2ff", "#eef8ea", "#fff3d6", "#f9e6e8", "#eeeeee"]
        for idx, (_, z) in enumerate(rz.iterrows()):
            if pd.isna(z["start_s"]) or pd.isna(z["end_s"]):
                continue
            x0 = temp_box[0] + (z["start_s"] - tmin) / (tmax - tmin) * (temp_box[2] - temp_box[0])
            x1 = temp_box[0] + (z["end_s"] - tmin) / (tmax - tmin) * (temp_box[2] - temp_box[0])
            fill = shade_colors[idx % len(shade_colors)]
            draw.rectangle((x0, temp_box[1], x1, temp_box[3]), fill=fill)
            draw.rectangle((x0, us_box[1], x1, us_box[3]), fill=fill)
            label = str(z["zone"]).replace("zone_", "Z").replace("_", " ")
            draw.text((x0 + 4, temp_box[1] + 4), label[:28], fill="#333333", font=small)
        for c in temp_cols:
            plot_series(draw, df[TIME_COL], df[c], temp_box, colors.get(c, "#666666"), width=3 if c == "T6" else 2, ymin=temp_ymin, ymax=temp_ymax)
        for ch in PRIMARY_US:
            if ch in df.columns:
                plot_series(draw, df[TIME_COL], df[ch], us_box, colors[ch], width=3, ymin=us_ymin, ymax=us_ymax)
        draw_axes(draw, temp_box, tmin, tmax, temp_ymin, temp_ymax, "Time (s)", "Temperature (C)", small)
        draw_axes(draw, us_box, tmin, tmax, us_ymin, us_ymax, "Time (s)", "Ultrasound (a.u.)", small)
        det = detachment[detachment["rep_id"] == rep_id]
        for _, row in det.iterrows():
            onset = row["detachment_onset_absolute_s"]
            if pd.notna(onset):
                x = us_box[0] + (onset - tmin) / (tmax - tmin) * (us_box[2] - us_box[0])
                draw.line((x, us_box[1], x, us_box[3]), fill="#222222", width=1)
        for _, row in det[det["detachment_offset_status"].eq("full_detachment_for_channel")].iterrows():
            offset = row["detachment_offset_s"]
            if pd.notna(offset):
                x = us_box[0] + (offset - tmin) / (tmax - tmin) * (us_box[2] - us_box[0])
                draw.line((x, us_box[1], x, us_box[3]), fill="#d62728", width=2)
                draw.text((x + 3, us_box[1] + 10), f"{row['sensor_code']} {row['channel']}", fill="#d62728", font=small)
        legend_x, legend_y = 95, 488
        for i, c in enumerate([*PRODUCT_TEMPS, "T2", "T6", *PRIMARY_US]):
            draw.line((legend_x + i * 110, legend_y, legend_x + i * 110 + 22, legend_y), fill=colors.get(c, "#666666"), width=3)
            draw.text((legend_x + i * 110 + 27, legend_y - 7), c, fill="black", font=small)
        draw.text((95, 812), "Black vertical lines = detachment onset; red vertical lines = full detachment offset only.", fill="#333333", font=small)
        out = fig_dir / f"{rep_id}_zones_us_offsets.png"
        img.save(out)
        paths.append(out)
    return paths


def make_lab_rupture_figures(detachment: pd.DataFrame, change_points: pd.DataFrame) -> list[Path]:
    fig_dir = OUTPUT_DIR / "figures_lab"
    fig_dir.mkdir(parents=True, exist_ok=True)
    paths = []
    font, small = get_fonts()
    colors = {"Rx1Tx1": "#3344cc", "Rx2Tx2": "#cc3377"}
    for _, row in detachment.iterrows():
        sheet = "at" + row["rep_id"]
        df = read_raw_sheet(sheet)
        if df is None or df.empty:
            continue
        ch = row["channel"]
        start = float(row["detachment_onset_absolute_s"])
        end = float(row["analysis_window_end_s"])
        seg = df[(df[TIME_COL] >= start) & (df[TIME_COL] <= end)].copy()
        if seg.empty or ch not in seg.columns:
            continue
        img = Image.new("RGB", (1300, 680), "white")
        draw = ImageDraw.Draw(img)
        draw.text((40, 20), f"{row['rep_id']} sensor {row['sensor_code']} {ch}: detachment offset decision", fill="black", font=font)
        box = (80, 95, 1220, 500)
        draw.rectangle(box, outline="#444444", width=1)
        vals = seg[ch].dropna()
        ymin, ymax = float(vals.min()) - 0.03, float(vals.max()) + 0.03
        plot_series(draw, seg[TIME_COL], seg[ch], box, colors.get(ch, "#3344cc"), width=3, ymin=ymin, ymax=ymax)
        for level, label, line_color in [
            (row["reference_us_level"], "pre-deposition reference", "#666666"),
            (row["threshold_10pct_lower"], "10% lower threshold", colors.get(ch, "#3344cc")),
        ]:
            if pd.notna(level) and ymax != ymin:
                ypix = box[3] - (level - ymin) / (ymax - ymin) * (box[3] - box[1])
                draw.line((box[0], ypix, box[2], ypix), fill=line_color, width=1)
                draw.text((box[2] - 280, ypix - 14), f"{ch} {label}", fill=line_color, font=small)
        # detachment onset
        x_on = box[0]
        draw.line((x_on, box[1], x_on, box[3]), fill="#222222", width=2)
        draw.text((x_on + 4, box[1] + 8), "detachment onset", fill="#222222", font=small)
        cps = change_points[
            (change_points["rep_id"] == row["rep_id"])
            & (change_points["sensor_code"].astype(str) == str(row["sensor_code"]))
            & (change_points["channel"] == ch)
        ]
        for _, cp in cps.iterrows():
            tt = cp["change_point_s"]
            if pd.notna(tt):
                x = box[0] + (tt - start) / (end - start) * (box[2] - box[0])
                cp_color = "#2ca02c" if cp["threshold_met"] else "#999999"
                line_top = box[1] if cp["threshold_met"] else box[3] - 70
                draw.line((x, line_top, x, box[3]), fill=cp_color, width=2 if cp["threshold_met"] else 1)
                draw.text((x + 2, box[3] - 18), str(int(cp["change_point_index"])), fill=cp_color, font=small)
        if row["detachment_offset_status"] == "full_detachment_for_channel" and pd.notna(row["detachment_offset_s"]):
            x = box[0] + (row["detachment_offset_s"] - start) / (end - start) * (box[2] - box[0])
            draw.line((x, box[1], x, box[3]), fill="#d62728", width=3)
            draw.text((x + 4, box[1] + 8), "selected full offset", fill="#d62728", font=small)
            line = f"Full offset: {row['detachment_offset_s']:.1f}s | US {row['us_level_at_decision']:.4f} | CPs until offset: {row['change_points_until_detachment_or_last']}"
        else:
            line = f"No full offset. Last diagnostic US {row['us_level_at_decision']:.4f} | total CPs: {row['change_points_detected_total']}"
        draw.text((85, 520), f"Reference {row['reference_us_level']:.4f}; threshold {row['threshold_10pct_lower']:.4f}; pass if US >= threshold", fill="#000000", font=font)
        draw.text((85, 550), line, fill=colors.get(ch, "#3344cc"), font=font)
        draw.text((85, 590), f"Decision: {row['detachment_offset_status']}", fill="#000000", font=font)
        draw.text((760, 590), "Grey = CP; green = passes; red = full offset only", fill="#333333", font=small)
        out = fig_dir / f"{row['rep_id']}_sensor_{row['sensor_code']}_{ch}_rupture_decision.png"
        img.save(out)
        paths.append(out)
    return paths


def concise_readme() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "part": "1",
                "sheet": "1 Summary Factor Effects",
                "what_it_answers": "Can variation in extracted summary parameters be explained by main independent variables?",
                "how_to_read": "Look for large/moderate eta2 effect sizes. This screens chocolate recipe, cooling configuration, cooling temperature, cooling speed, tempering, and trials.",
            },
            {
                "part": "2",
                "sheet": "2 Repeatability CV",
                "what_it_answers": "How repeatable are the extracted summary parameters within each experiment/set?",
                "how_to_read": "CV <5% is high repeatability, 5-15% moderate, >15% should be inspected. Very small means can inflate CV.",
            },
            {
                "part": "3",
                "sheet": "3 Temp Explains Outcomes",
                "what_it_answers": "For only the low-repeatability cases, do product/mould/ambient temperatures explain the variation?",
                "how_to_read": "Only extracted parameters with CV > 20% are shown. For example, if crystallization start is poorly repeatable, zone 2 temperature metrics are checked as possible reasons.",
            },
            {
                "part": "4",
                "sheet": "US Rupture Decisions",
                "what_it_answers": "For each repetition, sensor code, and primary US channel, was full detachment reached and at what offset time?",
                "how_to_read": "A time is reported only if a positive change point reaches the pre-deposition minus 10% threshold. Otherwise the status is partial.",
            },
            {
                "part": "5",
                "sheet": "US Change Points",
                "what_it_answers": "Which positive/upward US change points were detected and did they pass the threshold?",
                "how_to_read": "This is the audit trail behind the detachment offset decision.",
            },
            {
                "part": "6",
                "sheet": "Zone and US Figures / Rupture Decision Figures",
                "what_it_answers": "Visual checks for zones, onset, full-detachment offsets, rupture points, and threshold criteria.",
                "how_to_read": "Red offset lines are only drawn where full detachment was detected.",
            },
            {
                "part": "Zone note",
                "sheet": "Zone and US Figures",
                "what_it_answers": "How lab-device cooling zones are split.",
                "how_to_read": "For non-test lab-device runs, Z2 is cooling simulator until the first rise pattern, Z3 is intermediate cooling simulator until the later strong warming/demoulding transition. a5 anchors this at about 2000 s and 2550 s.",
            },
            {
                "part": "Caution",
                "sheet": "All sheets",
                "what_it_answers": "Statistical meaning",
                "how_to_read": "These are exploratory effect sizes and correlations. With few repetitions, they are pattern evidence and prioritization, not formal proof.",
            },
        ]
    )


def build_revised_tables(tables: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    repeatability = repeatability_cv_summary(tables)
    detachment_offset, detachment_change_points = lab_detachment_analysis(tables)
    return {
        "Read Me": concise_readme(),
        "1 Summary Factor Effects": factor_effects_for_summary(tables),
        "2 Repeatability CV": repeatability,
        "3 Temp Explains Outcomes": targeted_temp_outcome_links(tables, repeatability),
        "US Rupture Decisions": detachment_offset,
        "US Change Points": detachment_change_point_summary(detachment_offset),
    }


def write_df(ws, df: pd.DataFrame) -> None:
    if df.empty:
        ws["A1"] = "No data parsed for this sheet."
        return
    for col_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(1, col_idx, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            if isinstance(value, (np.floating, float)) and pd.notna(value):
                value = float(value)
            elif isinstance(value, (np.integer, int)) and pd.notna(value):
                value = int(value)
            elif isinstance(value, np.ndarray):
                value = None
            elif pd.isna(value):
                value = None
            ws.cell(row_idx, col_idx, value)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, col in enumerate(df.columns, 1):
        samples = [str(col), *[str(v) for v in df[col].head(100).fillna("").tolist()]]
        ws.column_dimensions[get_column_letter(idx)].width = min(max(len(s) for s in samples) + 2, 32)


def build_report() -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out = OUTPUT_DIR / "lab_trials_repeatability_report.xlsx"
    rebuilt_raw = rebuild_lab_raw_workbook_from_csv()
    _, tables = analyze()
    report_tables = build_revised_tables(tables)
    report_tables["Read Me"] = pd.concat(
        [
            report_tables["Read Me"],
            pd.DataFrame(
                [
                    {
                        "part": "Raw",
                        "sheet": "inputs/lab_trials",
                        "what_it_answers": "Which raw workbook was rebuilt from the replacement CSV files?",
                        "how_to_read": str(rebuilt_raw),
                    }
                ]
            ),
        ],
        ignore_index=True,
    )
    fig_dir = OUTPUT_DIR / "figures_lab"
    if fig_dir.exists():
        for png in fig_dir.glob("*.png"):
            png.unlink()
    zones_for_figures = detect_lab_zones(tables)
    rupture_decisions = report_tables["US Rupture Decisions"]
    _, detailed_change_points_for_figures = lab_detachment_analysis(tables)
    zone_figures = make_lab_zone_us_figures(zones_for_figures, rupture_decisions)
    rupture_figures = make_lab_rupture_figures(rupture_decisions, detailed_change_points_for_figures)

    wb = Workbook()
    wb.remove(wb.active)

    for name, df in report_tables.items():
        ws = wb.create_sheet(name[:31])
        write_df(ws, df)
        if name == "1 Summary Factor Effects" and ws.max_row > 1:
            for header in ws[1]:
                if header.value == "eta2_effect_size":
                    col = get_column_letter(header.column)
                    ws.conditional_formatting.add(
                        f"{col}2:{col}{ws.max_row}",
                        ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="F8696B"),
                    )
        if name == "2 Repeatability CV" and ws.max_row > 1:
            for header in ws[1]:
                if header.value == "cv_percent":
                    col = get_column_letter(header.column)
                    ws.conditional_formatting.add(
                        f"{col}2:{col}{ws.max_row}",
                        ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="F8696B"),
                    )
        if name == "3 Temp Explains Outcomes" and ws.max_row > 1:
            for header in ws[1]:
                if header.value in {"pearson_r", "min_curve_correlation"}:
                    col = get_column_letter(header.column)
                    ws.conditional_formatting.add(
                        f"{col}2:{col}{ws.max_row}",
                        ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"),
                    )
                if header.value in {"mean_curve_rmse", "max_curve_rmse", "amplitude_cv_percent"}:
                    col = get_column_letter(header.column)
                    ws.conditional_formatting.add(
                        f"{col}2:{col}{ws.max_row}",
                        ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="F8696B"),
                    )
    ws = wb.create_sheet("Zone and US Figures")
    row = 1
    for path in zone_figures:
        ws.cell(row, 1, path.name)
        img = ExcelImage(str(path))
        img.width = 980
        img.height = 581
        ws.add_image(img, f"A{row + 1}")
        row += 35

    ws = wb.create_sheet("Rupture Decision Figures")
    row = 1
    for path in rupture_figures:
        ws.cell(row, 1, path.name)
        img = ExcelImage(str(path))
        img.width = 980
        img.height = 512
        ws.add_image(img, f"A{row + 1}")
        row += 31
    try:
        wb.save(out)
    except PermissionError:
        out = OUTPUT_DIR / "lab_trials_repeatability_report_sensor_split.xlsx"
        wb.save(out)
    return out


def main() -> None:
    print(build_report())


if __name__ == "__main__":
    main()
