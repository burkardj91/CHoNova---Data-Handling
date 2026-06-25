from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from analyze_aasted_runs_zone_patterns import OUTPUT_DIR, PRODUCT_SENSORS


SOURCE_WORKBOOK = OUTPUT_DIR / "aasted3_old_configuration_comparison_report.xlsx"
SUMMARY_DIR = OUTPUT_DIR / "summary"
RAW_DIR = SUMMARY_DIR / "output_raw_data"
SUMMARY_WORKBOOK = SUMMARY_DIR / "aasted3_reference_based_comparison_summary.xlsx"
ZONE_FINDINGS_WORKBOOK = SUMMARY_DIR / "aasted3_zone_findings.xlsx"
CONTOUR_DATA_WORKBOOK = SUMMARY_DIR / "aasted3_contour_data.xlsx"
RAW_TABLES_WORKBOOK = RAW_DIR / "aasted3_raw_detail_tables.xlsx"


def write_df(ws, df: pd.DataFrame, header_color: str = "1F4E78") -> None:
    if df.empty:
        ws["A1"] = "No data available."
        return
    for col_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(1, col_idx, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=header_color)
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            if pd.isna(value):
                value = None
            ws.cell(row_idx, col_idx, value)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, col in enumerate(df.columns, 1):
        samples = [str(col), *[str(v) for v in df[col].head(100).fillna("").tolist()]]
        ws.column_dimensions[get_column_letter(idx)].width = min(max(len(s) for s in samples) + 2, 38)


def apply_hotspot_color_fills(ws) -> None:
    if ws.max_row <= 1:
        return
    headers = {cell.value: cell.column for cell in ws[1]}
    color_cols = [col for name, col in headers.items() if isinstance(name, str) and name.endswith("_color")]
    for row in range(2, ws.max_row + 1):
        for col in color_cols:
            color = str(ws.cell(row, col).value or "").replace("#", "")
            if len(color) == 6:
                ws.cell(row, col).fill = PatternFill("solid", fgColor=color)
        spread_color_col = headers.get("spread_color")
        spread_severity_col = headers.get("spread_severity")
        if spread_color_col and spread_severity_col:
            color = str(ws.cell(row, spread_color_col).value or "").replace("#", "")
            if len(color) == 6:
                ws.cell(row, spread_severity_col).fill = PatternFill("solid", fgColor=color)


def add_note(ws, row: int, title: str, text: str) -> int:
    ws.cell(row, 1, title)
    ws.cell(row, 2, text)
    ws.cell(row, 1).font = Font(bold=True, color="1F4E78")
    ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 45
    return row + 1


def read_sheet(name: str) -> pd.DataFrame:
    return pd.read_excel(SOURCE_WORKBOOK, sheet_name=name)


def make_readme() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "topic": "Analysis profile",
                "explanation": "Reference-based Aasted-3 comparison: one reference run is compared to one named comparison run using detected process zones, product temperature contours, and mechanical-zone summaries.",
            },
            {
                "topic": "Current comparison",
                "explanation": "Reference = reference_2291-4160. Comparison = old-configuration. Raw copies are stored in inputs/aasted3/.",
            },
            {
                "topic": "Zone definition",
                "explanation": "Reference zones are currently detected from T8, acc z, and gyro y patterns and compared against the reference process map. Conveyor-belt speed can shift physical zone timing, so speed metadata should be provided and used as a scaling/calibration input for future runs.",
            },
            {
                "topic": "Rolling window",
                "explanation": "A rolling window is only a computational smoothing window over nearby seconds. It has no standalone physical meaning; it helps stabilize noisy sensor signals before pattern detection.",
            },
            {
                "topic": "Extra wall time",
                "explanation": "Extra wall time is comparison run duration minus reference run duration. It is not automatically a stop or process fault; physical interpretation only comes after mapping that extra time to process zones.",
            },
            {
                "topic": "Repeatability mode",
                "explanation": "A separate repeatability analysis can be added when multiple repetitions are uploaded. That mode should compare runs within a setup, quantify CV/spread, inspect detachment patterns, and use an optional xxx_summary or parameter_summary input for landmarks and metadata.",
            },
            {
                "topic": "Aasted detachment offset",
                "explanation": "Optional architecture is installed. When a parameter_summary file is placed in inputs/aasted3, the report can mark crystallization onset, detachment onset, and complete/partial ultrasound detachment offsets using positive change points and a pre-deposition reference minus 10% threshold.",
            },
            {
                "topic": "Demoulding subclasses",
                "explanation": "The former broad demoulding zone is split into demoulding_twisting, demoulding_vibration, and final_demoulding. Twisting is identified from the acc-z plateau drop with the acc-x negative-parabola movement; vibration uses elevated IMU variability.",
            },
            {
                "topic": "Mechanical outlier roadmap",
                "explanation": "Mechanical-zone outlier detection should be trained from accumulated reference/repeatability runs. The current output summarizes all available IMU axes by zone; future calibration can define expected ranges by line, mould, conveyor speed, and zone.",
            },
            {
                "topic": "Prompt keywords to provide",
                "explanation": "Please specify: analysis_mode = reference_based_comparison or repeatability_analysis; line_profile = Aasted-3; reference_run; comparison labels; conveyor_belt_speed; raw_data files; optional parameter_summary/xxx_summary file; detachment_threshold_percent if not 10%; and desired customer/report detail level.",
            },
        ]
    )


def make_summary_table() -> pd.DataFrame:
    summary = read_sheet("Summary")
    detected_duration = read_sheet("Detected Duration")
    zone_findings = read_sheet("Zone Findings")
    rows = []
    for metric in ["reference_run", "comparison_run", "comparison_label", "extra_wall_time_s", "dtw_cost"]:
        hit = summary[summary["metric"].eq(metric)]
        if not hit.empty:
            rows.append({"section": "Run metadata", "metric": metric, "value": hit.iloc[0]["value"]})
    if "duration_delta_comparison_minus_reference_s" in detected_duration.columns:
        top = detected_duration.iloc[
            detected_duration["duration_delta_comparison_minus_reference_s"].abs().idxmax()
        ]
        rows.append(
            {
                "section": "Zone duration",
                "metric": "largest absolute zone duration delta",
                "value": f"{top['zone']}: {top['duration_delta_comparison_minus_reference_s']:.1f} s",
            }
        )
    if not zone_findings.empty:
        for _, row in zone_findings.head(8).iterrows():
            rows.append(
                {
                    "section": "Zone finding",
                    "metric": row["zone"],
                    "value": f"{row.get('duration_reading', '')}; {row.get('temperature_reading', '')}; {row.get('mechanical_reading', '')}",
                }
            )
    return pd.DataFrame(rows)


def append_table(ws, df: pd.DataFrame, start_row: int, title: str) -> int:
    ws.cell(start_row, 1, title)
    ws.cell(start_row, 1).font = Font(bold=True, size=13, color="1F4E78")
    start_row += 1
    if df.empty:
        ws.cell(start_row, 1, "No data available.")
        return start_row + 2
    for col_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(start_row, col_idx, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            if pd.isna(value):
                value = None
            ws.cell(row_idx, col_idx, value)
    for idx, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = min(max(len(str(col)) + 2, 14), 34)
    return start_row + len(df) + 3


def build_summary_workbook() -> None:
    SUMMARY_DIR.mkdir(parents=True, exist_ok=True)
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    source_wb = load_workbook(SOURCE_WORKBOOK)

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Read Me")
    write_df(ws, make_readme())

    ws = wb.create_sheet("Summary")
    write_df(ws, make_summary_table())

    ws = wb.create_sheet("Zone Visualization")
    ws["A1"] = "Reference and comparison zone visualizations"
    ws["A1"].font = Font(bold=True, size=15, color="1F4E78")
    images = [
        ("Reference run", OUTPUT_DIR / "reference_2291_4160_temperature_accz_zones.png"),
        ("Comparison run", OUTPUT_DIR / "old_configuration_temperature_accz_zones.png"),
    ]
    row = 3
    for title, path in images:
        ws.cell(row, 1, title)
        ws.cell(row, 1).font = Font(bold=True, color="1F4E78")
        img = XLImage(str(path))
        img.width = 980
        img.height = 537
        ws.add_image(img, f"A{row + 1}")
        row += 32
    row = append_table(ws, read_sheet("Zone Definitions"), row, "Reference Zone Definitions")
    append_table(ws, read_sheet("Detected Zones"), row, "Detected Zones")

    for sheet_name in ["Sensor Coordinates", "Mechanical By Zone", "Hotspot Summary"]:
        ws = wb.create_sheet(sheet_name)
        write_df(ws, read_sheet(sheet_name))
        if sheet_name == "Hotspot Summary":
            apply_hotspot_color_fills(ws)
        if sheet_name == "Mechanical By Zone":
            for header in ws[1]:
                if isinstance(header.value, str) and header.value.startswith("delta_"):
                    col = get_column_letter(header.column)
                    ws.conditional_formatting.add(
                        f"{col}2:{col}{ws.max_row}",
                        ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="num", mid_value=0, mid_color="FFFFFF", end_type="max", end_color="F8696B"),
                    )

    ws = wb.create_sheet("Contour Overview")
    ws["A1"] = "Product temperature contour overview"
    ws["A1"].font = Font(bold=True, size=15, color="1F4E78")
    ws["A2"] = f"Contours use product sensors: {', '.join(PRODUCT_SENSORS)}. Hotspot delta matrix is included below; full contour data is exported separately."
    ws["A2"].alignment = Alignment(wrap_text=True)
    row = 4
    for title, prefix in [("Reference contours", "reference_2291-4160"), ("Comparison contours", "old-configuration")]:
        ws.cell(row, 1, title)
        ws.cell(row, 1).font = Font(bold=True, color="1F4E78")
        row += 1
        images = sorted((OUTPUT_DIR / "contour_images").glob(f"{prefix}_*.png"))
        for idx, path in enumerate(images):
            img = XLImage(str(path))
            img.width = 312
            img.height = 156
            col = "A" if idx % 3 == 0 else "G" if idx % 3 == 1 else "M"
            anchor_row = row + (idx // 3) * 9
            ws.add_image(img, f"{col}{anchor_row}")
        row += max(9, ((len(images) + 2) // 3) * 9) + 3
    append_table(ws, read_sheet("Hotspot Delta Matrix"), row, "Hotspot Delta Matrix")

    wb.save(SUMMARY_WORKBOOK)


def export_detail_workbooks() -> None:
    zone_findings = read_sheet("Zone Findings")
    wb = Workbook()
    ws = wb.active
    ws.title = "Zone Findings"
    write_df(ws, zone_findings)
    wb.save(ZONE_FINDINGS_WORKBOOK)

    wb = Workbook()
    wb.remove(wb.active)
    for sheet in ["Contour Data", "Hotspot Delta Matrix"]:
        ws = wb.create_sheet(sheet)
        write_df(ws, read_sheet(sheet))
    wb.save(CONTOUR_DATA_WORKBOOK)

    wb = Workbook()
    wb.remove(wb.active)
    for sheet in [
        "Detected Duration",
        "Product Delta By Zone",
        "Detected Product By Zone",
        "Detected Product Delta",
        "Hotspot Sensor Data",
        "Aasted Detachment Summary",
        "Aasted US Change Points",
        "Aasted Parameter Landmarks",
        "Alignment Path Sample",
        "Reference Pattern",
        "Comparison Pattern",
    ]:
        ws = wb.create_sheet(sheet[:31])
        write_df(ws, read_sheet(sheet))
    wb.save(RAW_TABLES_WORKBOOK)


def main() -> None:
    build_summary_workbook()
    export_detail_workbooks()
    print(SUMMARY_WORKBOOK)
    print(ZONE_FINDINGS_WORKBOOK)
    print(CONTOUR_DATA_WORKBOOK)
    print(RAW_TABLES_WORKBOOK)


if __name__ == "__main__":
    main()
