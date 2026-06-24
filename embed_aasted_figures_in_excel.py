from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw, ImageFont

from analyze_aasted_runs_zone_patterns import (
    COMPARISON_CSV,
    COMPARISON_RUN_NAME,
    REFERENCE_CSV,
    REFERENCE_RUN_NAME,
    OUTPUT_DIR,
    TEMP_SENSORS,
    build_seconds,
    detect_pattern_zones,
    read_run,
)


SOURCE_WORKBOOK = OUTPUT_DIR / "aasted3_pattern_detected_hotspot_report_v2.xlsx"
OUTPUT_WORKBOOK = OUTPUT_DIR / "aasted3_pattern_detected_hotspot_report_with_ultrasound_figures.xlsx"
TEMP_PLOT_SENSORS = ["T2", "T3", "T4", "T5", "T6", "T7", "T8", "T9"]
ULTRASOUND_PLOT_CHANNELS = ["Rx1Tx1", "Rx2Tx2"]

COLORS = {
    "T1": (127, 29, 29),
    "T2": (220, 38, 38),
    "T3": (37, 99, 235),
    "T4": (22, 163, 74),
    "T5": (249, 115, 22),
    "T6": (124, 58, 237),
    "T7": (8, 145, 178),
    "T8": (17, 24, 39),
    "T9": (161, 98, 7),
    "acc z": (0, 0, 0),
    "Rx1Tx1": (20, 83, 45),
    "Rx2Tx2": (146, 64, 14),
}

ZONE_COLORS = [
    (239, 246, 255),
    (240, 253, 244),
    (255, 247, 237),
    (254, 242, 242),
    (248, 250, 252),
    (236, 254, 255),
    (240, 253, 250),
    (247, 254, 231),
    (253, 242, 248),
    (245, 243, 255),
    (250, 250, 250),
    (241, 245, 249),
]


def font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        r"C:\Windows\Fonts\arialbd.ttf" if bold else r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\calibrib.ttf" if bold else r"C:\Windows\Fonts\calibri.ttf",
    ]
    for path in candidates:
        if Path(path).exists():
            return ImageFont.truetype(path, size=size)
    return ImageFont.load_default()


def nice_ticks(vmin: float, vmax: float, n: int = 6) -> list[float]:
    if not np.isfinite(vmin) or not np.isfinite(vmax) or vmin == vmax:
        return [vmin]
    raw_step = (vmax - vmin) / max(n - 1, 1)
    mag = 10 ** np.floor(np.log10(abs(raw_step)))
    nice = np.array([1, 2, 2.5, 5, 10]) * mag
    step = float(nice[np.argmin(np.abs(nice - raw_step))])
    start = np.floor(vmin / step) * step
    end = np.ceil(vmax / step) * step
    ticks = []
    val = start
    while val <= end + step * 0.5:
        ticks.append(round(float(val), 6))
        val += step
    return ticks


def downsample(df: pd.DataFrame, max_points: int = 1800) -> pd.DataFrame:
    if len(df) <= max_points:
        return df
    return df.iloc[:: int(np.ceil(len(df) / max_points))].copy()


def draw_polyline(draw: ImageDraw.ImageDraw, points: list[tuple[float, float]], color: tuple[int, int, int], width: int = 2) -> None:
    clean = [(float(x), float(y)) for x, y in points if np.isfinite(x) and np.isfinite(y)]
    if len(clean) > 1:
        draw.line(clean, fill=color, width=width, joint="curve")


def paste_vertical_text(
    img: Image.Image,
    text: str,
    xy: tuple[int, int],
    fill: tuple[int, int, int],
    size: int = 15,
    clockwise: bool = False,
) -> None:
    text_font = font(size)
    bbox = ImageDraw.Draw(Image.new("RGBA", (1, 1))).textbbox((0, 0), text, font=text_font)
    label = Image.new("RGBA", (bbox[2] - bbox[0] + 10, bbox[3] - bbox[1] + 10), (255, 255, 255, 0))
    label_draw = ImageDraw.Draw(label)
    label_draw.text((5, 5), text, fill=fill, font=text_font)
    rotated = label.rotate(270 if clockwise else 90, expand=True)
    img.paste(rotated, xy, rotated)


def make_png(run_name: str, csv_path: Path, out_path: Path) -> Path:
    df = read_run(csv_path)
    seconds = build_seconds(df)
    zones = detect_pattern_zones(seconds, run_name)

    temp = df.dropna(subset=TEMP_SENSORS, how="all")[["time", "elapsed_s", *TEMP_SENSORS]].copy()
    temp = downsample(temp, 1800)
    ultrasound = df.dropna(subset=ULTRASOUND_PLOT_CHANNELS, how="all")[["time", "elapsed_s", *ULTRASOUND_PLOT_CHANNELS]].copy()
    ultrasound = downsample(ultrasound, 1800)
    motion = seconds[["elapsed_sec", "accz_mean"]].copy()
    motion["time"] = df["time"].min() + motion["elapsed_sec"]
    motion = downsample(motion, 1800)

    x_min = float(df["time"].min())
    x_max = float(df["time"].max())
    temp_min = float(np.nanmin(temp[TEMP_PLOT_SENSORS].to_numpy()))
    temp_max = float(np.nanmax(temp[TEMP_PLOT_SENSORS].to_numpy()))
    temp_pad = (temp_max - temp_min) * 0.06
    y_temp_min = temp_min - temp_pad
    y_temp_max = temp_max + temp_pad
    acc_min = float(np.nanquantile(motion["accz_mean"], 0.005))
    acc_max = float(np.nanquantile(motion["accz_mean"], 0.995))
    acc_pad = max((acc_max - acc_min) * 0.10, 0.05)
    y_acc_min = acc_min - acc_pad
    y_acc_max = acc_max + acc_pad
    rh_min = float(np.nanmin(temp["T1"].to_numpy()))
    rh_max = float(np.nanmax(temp["T1"].to_numpy()))
    rh_pad = max((rh_max - rh_min) * 0.06, 2.0)
    y_rh_min = rh_min - rh_pad
    y_rh_max = rh_max + rh_pad
    us_values = ultrasound[ULTRASOUND_PLOT_CHANNELS].to_numpy(dtype=float)
    us_min = float(np.nanquantile(us_values, 0.01))
    us_max = float(np.nanquantile(us_values, 0.99))
    us_pad = max((us_max - us_min) * 0.12, 0.02)
    y_us_min = us_min - us_pad
    y_us_max = us_max + us_pad

    width, height = 1680, 920
    left, right, top, bottom = 112, 168, 80, 190
    plot_w = width - left - right
    plot_h = 430
    us_gap = 34
    us_top = top + plot_h + us_gap
    us_h = 210

    def sx(t: float) -> float:
        return left + (t - x_min) / (x_max - x_min) * plot_w

    def sy_temp(v: float) -> float:
        return top + (y_temp_max - v) / (y_temp_max - y_temp_min) * plot_h

    def sy_acc(v: float) -> float:
        return top + (y_acc_max - v) / (y_acc_max - y_acc_min) * plot_h

    def sy_rh(v: float) -> float:
        return top + (y_rh_max - v) / (y_rh_max - y_rh_min) * plot_h

    def sy_us(v: float) -> float:
        return us_top + (y_us_max - v) / (y_us_max - y_us_min) * us_h

    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img, "RGBA")

    draw.text((left, 26), f"{run_name}: temperatures and acc z with detected process zones", fill=(17, 24, 39), font=font(25, True))
    draw.text(
        (left, 58),
        "x-axis uses absolute CSV time. Left axis: temperature T2-T9. Right axis: acc z. Outer right axis: T1 relabeled as humidity/RH.",
        fill=(71, 85, 105),
        font=font(15),
    )

    for idx, row in enumerate(zones.itertuples(index=False)):
        start_abs = x_min + float(row.detected_start_s)
        end_abs = x_min + float(row.detected_end_s)
        x0 = max(left, sx(start_abs))
        x1 = min(left + plot_w, sx(end_abs))
        draw.rectangle((x0, top, x1, top + plot_h), fill=(*ZONE_COLORS[idx % len(ZONE_COLORS)], 230))
        draw.rectangle((x0, us_top, x1, us_top + us_h), fill=(*ZONE_COLORS[idx % len(ZONE_COLORS)], 230))
        if x1 - x0 > 44:
            label = str(row.zone).replace("_", " ")
            # Rotated labels are less reliable in Excel previews, so keep them horizontal and compact.
            draw.text((x0 + 4, top + 8 + (idx % 2) * 16), label[:24], fill=(51, 65, 85), font=font(11))

    for tick in nice_ticks(y_temp_min, y_temp_max, 7):
        y = sy_temp(tick)
        draw.line((left, y, left + plot_w, y), fill=(226, 232, 240), width=1)
        draw.text((left - 56, y - 8), f"{tick:g}", fill=(51, 65, 85), font=font(12))

    for tick in nice_ticks(x_min, x_max, 9):
        x = sx(tick)
        if left <= x <= left + plot_w:
            draw.line((x, top, x, top + plot_h), fill=(241, 245, 249), width=1)
            draw.line((x, us_top, x, us_top + us_h), fill=(241, 245, 249), width=1)
            draw.text((x - 22, us_top + us_h + 18), f"{tick:g}", fill=(51, 65, 85), font=font(12))

    for tick in nice_ticks(y_acc_min, y_acc_max, 6):
        y = sy_acc(tick)
        draw.text((left + plot_w + 12, y - 8), f"{tick:g}", fill=(17, 24, 39), font=font(12))

    rh_axis_x = left + plot_w + 82
    draw.line((rh_axis_x, top, rh_axis_x, top + plot_h), fill=COLORS["T1"], width=2)
    for tick in nice_ticks(y_rh_min, y_rh_max, 6):
        y = sy_rh(tick)
        draw.line((rh_axis_x - 5, y, rh_axis_x + 5, y), fill=COLORS["T1"], width=1)
        draw.text((rh_axis_x + 10, y - 8), f"{tick:g}", fill=COLORS["T1"], font=font(12))

    for tick in nice_ticks(y_us_min, y_us_max, 4):
        y = sy_us(tick)
        draw.line((left, y, left + plot_w, y), fill=(226, 232, 240), width=1)
        draw.text((left - 56, y - 8), f"{tick:g}", fill=(51, 65, 85), font=font(12))

    draw.rectangle((left, top, left + plot_w, top + plot_h), outline=(51, 65, 85), width=2)
    draw.rectangle((left, us_top, left + plot_w, us_top + us_h), outline=(51, 65, 85), width=2)
    draw.text((left + plot_w / 2 - 65, us_top + us_h + 42), "absolute time [s]", fill=(17, 24, 39), font=font(15))
    paste_vertical_text(img, "Temperature T2-T9 [deg C]", (18, int(top + plot_h / 2 - 108)), (17, 24, 39), size=15)
    paste_vertical_text(img, "acc z [g]", (width - 78, int(top + plot_h / 2 - 42)), (17, 24, 39), size=15, clockwise=True)
    paste_vertical_text(img, "RH [%]", (width - 30, int(top + plot_h / 2 - 32)), COLORS["T1"], size=15, clockwise=True)
    paste_vertical_text(img, "Ultrasound (a.u.)", (18, int(us_top + us_h / 2 - 72)), (17, 24, 39), size=15)

    for sensor in TEMP_PLOT_SENSORS:
        points = [(sx(float(t)), sy_temp(float(v))) for t, v in zip(temp["time"], temp[sensor]) if pd.notna(v)]
        draw_polyline(draw, points, COLORS[sensor], width=3 if sensor == "T8" else 2)
    rh_points = [(sx(float(t)), sy_rh(float(v))) for t, v in zip(temp["time"], temp["T1"]) if pd.notna(v)]
    draw_polyline(draw, rh_points, COLORS["T1"], width=2)
    acc_points = [(sx(float(t)), sy_acc(float(v))) for t, v in zip(motion["time"], motion["accz_mean"]) if pd.notna(v)]
    draw_polyline(draw, acc_points, COLORS["acc z"], width=2)
    for channel in ULTRASOUND_PLOT_CHANNELS:
        points = [(sx(float(t)), sy_us(float(v))) for t, v in zip(ultrasound["time"], ultrasound[channel]) if pd.notna(v)]
        draw_polyline(draw, points, COLORS[channel], width=4)

    lx, ly = left, height - 62
    for idx, item in enumerate(["T1 / RH", *TEMP_PLOT_SENSORS, "acc z", *ULTRASOUND_PLOT_CHANNELS]):
        x = lx + (idx % 12) * 118
        y = ly + (idx // 10) * 24
        color_key = "T1" if item == "T1 / RH" else item
        draw.line((x, y, x + 28, y), fill=COLORS[color_key], width=5 if color_key in ULTRASOUND_PLOT_CHANNELS else 4 if color_key in ["T8", "acc z"] else 3)
        draw.text((x + 36, y - 8), item, fill=(17, 24, 39), font=font(13))

    table_x, table_y = left + plot_w - 430, us_top + us_h + 34
    draw.text((table_x, table_y), "Detected zone starts (absolute s)", fill=(17, 24, 39), font=font(13, True))
    for i, row in enumerate(zones.itertuples(index=False)):
        label = str(row.zone).replace("_", " ")
        abs_start = x_min + float(row.detected_start_s)
        if i < 9:
            draw.text((table_x, table_y + 18 + i * 13), f"{abs_start:.0f}: {label}", fill=(51, 65, 85), font=font(10))

    img.save(out_path)
    return out_path


def write_note(ws, text: str) -> None:
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws["A1"].fill = PatternFill("solid", fgColor="D9EAF7")
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 36
    ws.column_dimensions["A"].width = 34


def embed_figures() -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ref_png = OUTPUT_DIR / "reference_2291_4160_temperature_accz_zones.png"
    comparison_png = OUTPUT_DIR / "old_configuration_temperature_accz_zones.png"
    make_png(REFERENCE_RUN_NAME, REFERENCE_CSV, ref_png)
    make_png(COMPARISON_RUN_NAME, COMPARISON_CSV, comparison_png)

    wb = load_workbook(SOURCE_WORKBOOK)
    for sheet_name in ["Figure Reference", "Figure Irregular", "Figure Old Configuration"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    ref_ws = wb.create_sheet("Figure Reference", 0)
    write_note(ref_ws, "Reference run: T2-T9 on temperature axis, T1 as RH, acc z on right axis, and Rx1Tx1/Rx2Tx2 ultrasound traces in lower panel.")
    ref_img = XLImage(str(ref_png))
    ref_img.width = 1120
    ref_img.height = 613
    ref_ws.add_image(ref_img, "A3")

    comp_ws = wb.create_sheet("Figure Old Configuration", 1)
    write_note(comp_ws, "Old-configuration comparison run: T2-T9 on temperature axis, T1 as RH, acc z on right axis, and Rx1Tx1/Rx2Tx2 ultrasound traces in lower panel.")
    comp_img = XLImage(str(comparison_png))
    comp_img.width = 1120
    comp_img.height = 613
    comp_ws.add_image(comp_img, "A3")

    wb.save(OUTPUT_WORKBOOK)
    return OUTPUT_WORKBOOK


def main() -> None:
    print(embed_figures())


if __name__ == "__main__":
    main()
