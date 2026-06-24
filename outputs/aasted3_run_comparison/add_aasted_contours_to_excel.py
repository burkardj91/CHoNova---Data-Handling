from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw, ImageFont

from analyze_aasted_runs_zone_patterns import (
    COMPARISON_CSV,
    COMPARISON_RUN_NAME,
    REFERENCE_CSV,
    REFERENCE_RUN_NAME,
    OUTPUT_DIR,
    TEMP_SENSORS,
    build_seconds,
    detect_pattern_zones,
    read_run,
)


SOURCE_WORKBOOK = OUTPUT_DIR / "aasted3_pattern_detected_hotspot_report_with_ultrasound_figures.xlsx"
OUTPUT_WORKBOOK = OUTPUT_DIR / "aasted3_old_configuration_comparison_report.xlsx"
MOULD_WIDTH = 112.2
MOULD_HEIGHT = 33.3

TEMP_COORDS = {
    "T1": (53.0, 23.0),
    "T2": (70.0, 13.0),
    "T3": (5.0, 5.0),
    "T4": (5.0, 31.0),
    "T5": (108.0, 5.0),
    "T6": (76.0, 24.0),
    "T7": (99.0, 30.0),
    "T8": (93.0, 23.0),
    "T9": (55.0, 10.0),
}
US_COORDS = {
    "Rx1": (60.0, 30.0),
    "Tx1": (76.0, 30.0),
    "Rx2": (60.0, 17.0),
    "Tx2": (76.0, 17.0),
}
PRODUCT_CONTOUR_SENSORS = ["T2", "T3", "T4", "T5", "T7"]


def font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        r"C:\Windows\Fonts\arialbd.ttf" if bold else r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\calibrib.ttf" if bold else r"C:\Windows\Fonts\calibri.ttf",
    ]
    for path in candidates:
        if Path(path).exists():
            return ImageFont.truetype(path, size=size)
    return ImageFont.load_default()


def color_scale(value: float, vmin: float, vmax: float) -> tuple[int, int, int]:
    if not np.isfinite(value):
        return (245, 245, 245)
    t = 0.5 if vmax == vmin else max(0.0, min(1.0, (value - vmin) / (vmax - vmin)))
    stops = [
        (0.00, (49, 54, 149)),
        (0.25, (69, 117, 180)),
        (0.50, (255, 255, 191)),
        (0.75, (253, 174, 97)),
        (1.00, (165, 0, 38)),
    ]
    for (a, ca), (b, cb) in zip(stops, stops[1:]):
        if a <= t <= b:
            local = (t - a) / (b - a)
            return tuple(int(ca[i] + (cb[i] - ca[i]) * local) for i in range(3))
    return stops[-1][1]


def zone_temp_means(df: pd.DataFrame, zones: pd.DataFrame, run_name: str) -> pd.DataFrame:
    temp = df.dropna(subset=TEMP_SENSORS, how="all").copy()
    rows = []
    for row in zones[zones["run"] == run_name].itertuples(index=False):
        subset = temp[(temp["elapsed_s"] >= row.detected_start_s) & (temp["elapsed_s"] < row.detected_end_s)]
        values = subset[TEMP_SENSORS].mean()
        rows.append(
            {
                "run": run_name,
                "zone": row.zone,
                "detected_start_s": row.detected_start_s,
                "detected_end_s": row.detected_end_s,
                "duration_s": row.duration_s,
                **{sensor: float(values[sensor]) for sensor in TEMP_SENSORS},
            }
        )
    return pd.DataFrame(rows)


def idw_grid(values: dict[str, float], nx: int = 260, ny: int = 90) -> np.ndarray:
    xs = np.linspace(0, MOULD_WIDTH, nx)
    ys = np.linspace(0, MOULD_HEIGHT, ny)
    grid = np.zeros((ny, nx), dtype=float)
    coords = np.array([TEMP_COORDS[s] for s in PRODUCT_CONTOUR_SENSORS], dtype=float)
    vals = np.array([values[s] for s in PRODUCT_CONTOUR_SENSORS], dtype=float)
    for yi, y in enumerate(ys):
        dx = xs[None, :] - coords[:, 0:1]
        dy = y - coords[:, 1:2]
        dist2 = dx * dx + dy * dy
        exact = dist2 < 1e-8
        weights = 1.0 / np.maximum(dist2, 1e-6)
        interp = (weights * vals[:, None]).sum(axis=0) / weights.sum(axis=0)
        if exact.any():
            exact_cols = np.where(exact.any(axis=0))[0]
            for col in exact_cols:
                interp[col] = vals[np.where(exact[:, col])[0][0]]
        grid[ny - yi - 1, :] = interp
    return grid


def draw_marker(draw: ImageDraw.ImageDraw, x: float, y: float, label: str, fill: tuple[int, int, int], outline=(255, 255, 255)) -> None:
    r = 5
    draw.ellipse((x - r, y - r, x + r, y + r), fill=fill, outline=outline, width=2)
    draw.text((x + 7, y - 7), label, fill=(17, 24, 39), font=font(9, True))


def local_scale(values: dict[str, float]) -> tuple[float, float]:
    vals = np.array([values[s] for s in PRODUCT_CONTOUR_SENSORS], dtype=float)
    center = float(np.nanmean(vals))
    vmin = float(np.nanmin(vals))
    vmax = float(np.nanmax(vals))
    # Keep every zone readable, but do not exaggerate numerical noise too harshly.
    min_span = 1.0
    if vmax - vmin < min_span:
        vmin = center - min_span / 2
        vmax = center + min_span / 2
    pad = max((vmax - vmin) * 0.08, 0.05)
    return vmin - pad, vmax + pad


def scale_for_values(vals: np.ndarray) -> tuple[float, float]:
    vals = vals[np.isfinite(vals)]
    center = float(np.nanmean(vals))
    vmin = float(np.nanmin(vals))
    vmax = float(np.nanmax(vals))
    min_span = 1.0
    if vmax - vmin < min_span:
        vmin = center - min_span / 2
        vmax = center + min_span / 2
    pad = max((vmax - vmin) * 0.08, 0.05)
    return vmin - pad, vmax + pad


def make_zone_contour(values: dict[str, float], title: str, out_path: Path, vmin: float | None = None, vmax: float | None = None) -> Path:
    if vmin is None or vmax is None:
        vmin, vmax = local_scale(values)
    w, h = 520, 260
    left, top, plot_w, plot_h = 44, 42, 410, 122
    img = Image.new("RGB", (w, h), "white")
    draw = ImageDraw.Draw(img)
    draw.text((16, 12), title, fill=(17, 24, 39), font=font(14, True))

    grid = idw_grid(values)
    heat = Image.new("RGB", (grid.shape[1], grid.shape[0]))
    pix = heat.load()
    for yy in range(grid.shape[0]):
        for xx in range(grid.shape[1]):
            pix[xx, yy] = color_scale(float(grid[yy, xx]), vmin, vmax)
    heat = heat.resize((plot_w, plot_h), Image.Resampling.BILINEAR)
    img.paste(heat, (left, top))

    def sx(x_mm: float) -> float:
        return left + x_mm / MOULD_WIDTH * plot_w

    def sy(y_mm: float) -> float:
        return top + plot_h - y_mm / MOULD_HEIGHT * plot_h

    draw.rectangle((left, top, left + plot_w, top + plot_h), outline=(31, 41, 55), width=2)
    for sensor in PRODUCT_CONTOUR_SENSORS:
        x, y = TEMP_COORDS[sensor]
        draw_marker(draw, sx(x), sy(y), sensor, fill=(255, 255, 255), outline=(17, 24, 39))
        draw.text((sx(x) + 7, sy(y) + 6), f"{values[sensor]:.1f}", fill=(17, 24, 39), font=font(8))

    for label, (x, y) in US_COORDS.items():
        px, py = sx(x), sy(y)
        size = 7
        draw.polygon([(px, py - size), (px - size, py + size), (px + size, py + size)], fill=(0, 0, 0))
        draw.text((px + 8, py - 8), label, fill=(0, 0, 0), font=font(9, True))

    # Color bar.
    bar_x, bar_y, bar_w, bar_h = left, top + plot_h + 28, plot_w, 12
    for i in range(bar_w):
        c = color_scale(vmin + (vmax - vmin) * i / max(bar_w - 1, 1), vmin, vmax)
        draw.line((bar_x + i, bar_y, bar_x + i, bar_y + bar_h), fill=c)
    draw.rectangle((bar_x, bar_y, bar_x + bar_w, bar_y + bar_h), outline=(31, 41, 55), width=1)
    draw.text((bar_x, bar_y + 18), f"{vmin:.1f} C", fill=(17, 24, 39), font=font(10))
    draw.text((bar_x + bar_w - 42, bar_y + 18), f"{vmax:.1f} C", fill=(17, 24, 39), font=font(10))
    draw.text((bar_x + 104, bar_y + 18), "same scale for this zone across both repetitions; IDW from T2, T3, T4, T5, T7", fill=(17, 24, 39), font=font(10))

    mean_temp = np.nanmean([values[s] for s in PRODUCT_CONTOUR_SENSORS])
    draw.text((left, h - 28), f"Zone product mean: {mean_temp:.2f} C | Ultrasound Rx/Tx markers overlaid", fill=(51, 65, 85), font=font(10))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    img.save(out_path)
    return out_path


def write_table(ws, df: pd.DataFrame) -> None:
    for col_idx, col in enumerate(df.columns, 1):
        c = ws.cell(1, col_idx, col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row_idx, col_idx, value)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col) + 2, 24)


def add_contour_sheets() -> Path:
    reference = read_run(REFERENCE_CSV)
    comparison = read_run(COMPARISON_CSV)
    reference_zones = detect_pattern_zones(build_seconds(reference), REFERENCE_RUN_NAME)
    comparison_zones = detect_pattern_zones(build_seconds(comparison), COMPARISON_RUN_NAME)
    zones = pd.concat([reference_zones, comparison_zones], ignore_index=True)
    means = pd.concat(
        [
            zone_temp_means(reference, reference_zones, REFERENCE_RUN_NAME),
            zone_temp_means(comparison, comparison_zones, COMPARISON_RUN_NAME),
        ],
        ignore_index=True,
    )
    zone_scales = {}
    for zone, group in means.groupby("zone"):
        zone_scales[zone] = scale_for_values(group[PRODUCT_CONTOUR_SENSORS].to_numpy(dtype=float).ravel())
    workbook = load_workbook(SOURCE_WORKBOOK)
    for sheet_name in ["Contour Reference", "Contour Irregular", "Contour Old Configuration", "Contour Data", "Sensor Coordinates"]:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    coordinate_rows = []
    for sensor, (x, y) in TEMP_COORDS.items():
        coordinate_rows.append({"type": "temperature/RH", "sensor": sensor, "x_mm": x, "y_mm": y, "used_in_product_contour": sensor in PRODUCT_CONTOUR_SENSORS})
    for sensor, (x, y) in US_COORDS.items():
        coordinate_rows.append({"type": "ultrasound", "sensor": sensor, "x_mm": x, "y_mm": y, "used_in_product_contour": False})

    coords_ws = workbook.create_sheet("Sensor Coordinates", 2)
    write_table(coords_ws, pd.DataFrame(coordinate_rows))

    data_ws = workbook.create_sheet("Contour Data", 3)
    write_table(data_ws, means)

    for run_name, sheet_name in [(REFERENCE_RUN_NAME, "Contour Reference"), (COMPARISON_RUN_NAME, "Contour Old Configuration")]:
        ws = workbook.create_sheet(sheet_name, 2 if run_name.startswith("reference") else 3)
        ws["A1"] = f"{run_name}: mean mould temperature contours by detected zone"
        ws["A1"].font = Font(bold=True, size=15, color="1F4E78")
        ws["A2"] = "Contours use only product mean temperatures per zone: T2, T3, T4, T5, T7. Ultrasound Rx/Tx positions are overlaid as markers."
        ws["A2"].alignment = Alignment(wrap_text=True)
        ws.column_dimensions["A"].width = 18
        run_means = means[means["run"] == run_name].reset_index(drop=True)
        for idx, row in run_means.iterrows():
            image_path = OUTPUT_DIR / "contour_images" / f"{run_name}_{idx+1:02d}_{row['zone']}.png"
            values = {sensor: row[sensor] for sensor in TEMP_SENSORS}
            title = f"{row['zone']} ({row['detected_start_s']:.0f}-{row['detected_end_s']:.0f}s)"
            vmin, vmax = zone_scales[row["zone"]]
            make_zone_contour(values, title, image_path, vmin, vmax)
            xl_img = XLImage(str(image_path))
            xl_img.width = 390
            xl_img.height = 195
            col = "A" if idx % 2 == 0 else "J"
            row_anchor = 4 + (idx // 2) * 12
            ws.add_image(xl_img, f"{col}{row_anchor}")
        for r in range(4, 80):
            ws.row_dimensions[r].height = 18

    workbook.save(OUTPUT_WORKBOOK)
    return OUTPUT_WORKBOOK


def main() -> None:
    print(add_contour_sheets())


if __name__ == "__main__":
    main()
