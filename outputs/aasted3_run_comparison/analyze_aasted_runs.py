from __future__ import annotations

import math
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


WORKSPACE = Path(r"C:\Users\BurkardJohannes\Documents\CHoNova _ Data Understanding")
OUTPUT_DIR = WORKSPACE / "outputs" / "aasted3_run_comparison"
IRREGULAR_CSV = Path(r"C:\Users\BurkardJohannes\Desktop\Temp Files\Aasted 3\260604_aasted3_119-2290s.csv")
REFERENCE_CSV = Path(r"C:\Users\BurkardJohannes\Desktop\Temp Files\Aasted 3\260604_aasted3_2291-4160.csv")

TEMP_SENSORS = [f"T{i}" for i in range(1, 10)]
PRODUCT_SENSORS = ["T2", "T3", "T4", "T5", "T6"]
MECH_SENSORS = ["acc x", "acc y", "acc z", "gyro x", "gyro y", "gyro z"]
KEY_MECH_SENSORS = ["acc z", "gyro y"]
COORDS = {
    "T1": (53.0, 23.0),
    "T2": (70.0, 13.0),
    "T3": (5.0, 5.0),
    "T4": (5.0, 31.0),
    "T5": (108.0, 5.0),
    "T6": (76.0, 24.0),
    "T7": (99.0, 30.0),
    "T8": (93.0, 23.0),
    "T9": (55.0, 10.0),
}


def read_run(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    df["elapsed_s"] = df["time"] - df["time"].min()
    return df


def temp_rows(df: pd.DataFrame) -> pd.DataFrame:
    return df.dropna(subset=TEMP_SENSORS, how="all").copy()


def motion_seconds(df: pd.DataFrame) -> pd.DataFrame:
    motion = df.dropna(subset=KEY_MECH_SENSORS, how="all").copy()
    motion["sec_abs"] = np.floor(motion["time"]).astype(int)
    grouped = motion.groupby("sec_abs").agg(
        n=("time", "count"),
        accx_mean=("acc x", "mean"),
        accy_mean=("acc y", "mean"),
        accz_mean=("acc z", "mean"),
        gyrox_mean=("gyro x", "mean"),
        gyroy_mean=("gyro y", "mean"),
        gyroz_mean=("gyro z", "mean"),
        accx_std=("acc x", "std"),
        accy_std=("acc y", "std"),
        accz_std=("acc z", "std"),
        gyrox_std=("gyro x", "std"),
        gyroy_std=("gyro y", "std"),
        gyroz_std=("gyro z", "std"),
    )
    grouped = grouped.reset_index()
    std_cols = ["accx_std", "accy_std", "accz_std", "gyrox_std", "gyroy_std", "gyroz_std"]
    grouped["motion_std_sum"] = grouped[std_cols].fillna(0).sum(axis=1)
    grouped["roll30_motion_std"] = (
        grouped["motion_std_sum"].rolling(30, min_periods=15, center=True).median()
    )
    grouped["elapsed_s"] = grouped["sec_abs"] - math.floor(df["time"].min())
    return grouped


def contiguous_segments(seconds: pd.Series, mask: pd.Series, min_len: int = 20) -> list[tuple[int, int, int]]:
    segs: list[tuple[int, int, int]] = []
    start = None
    prev = None
    for sec, is_hit in zip(seconds.astype(int), mask.fillna(False).astype(bool)):
        if is_hit and start is None:
            start = sec
        if start is not None and (not is_hit or (prev is not None and sec != prev + 1)):
            end = int(prev)
            if end - int(start) + 1 >= min_len:
                segs.append((int(start), end, end - int(start) + 1))
            start = int(sec) if is_hit else None
        prev = int(sec)
    if start is not None and prev is not None and int(prev) - int(start) + 1 >= min_len:
        segs.append((int(start), int(prev), int(prev) - int(start) + 1))
    return segs


def merge_segments(segs: list[tuple[int, int, int]], max_gap: int = 75) -> list[tuple[int, int, int]]:
    if not segs:
        return []
    merged: list[list[int]] = [[segs[0][0], segs[0][1]]]
    for start, end, _ in segs[1:]:
        if start - merged[-1][1] <= max_gap:
            merged[-1][1] = end
        else:
            merged.append([start, end])
    return [(start, end, end - start + 1) for start, end in merged]


def summarize_product(temp: pd.DataFrame, run_name: str) -> pd.DataFrame:
    rows = []
    for sensor in PRODUCT_SENSORS:
        s = temp[sensor].dropna()
        rows.append(
            {
                "run": run_name,
                "sensor": sensor,
                "x_mm": COORDS[sensor][0],
                "y_mm": COORDS[sensor][1],
                "mean_C": s.mean(),
                "std_C": s.std(),
                "min_C": s.min(),
                "p05_C": s.quantile(0.05),
                "median_C": s.median(),
                "p95_C": s.quantile(0.95),
                "max_C": s.max(),
            }
        )
    return pd.DataFrame(rows)


def aligned_product_comparison(irregular_temp: pd.DataFrame, reference_temp: pd.DataFrame, stop_start: float, stop_end: float) -> pd.DataFrame:
    stop_duration = stop_end - stop_start
    irr = irregular_temp[["time", "elapsed_s", *PRODUCT_SENSORS]].copy()
    ref = reference_temp[["time", "elapsed_s", *PRODUCT_SENSORS]].copy()
    irr["active_elapsed_s"] = irr["elapsed_s"]
    irr.loc[irr["time"] > stop_end, "active_elapsed_s"] -= stop_duration
    irr.loc[(irr["time"] >= stop_start) & (irr["time"] <= stop_end), "active_elapsed_s"] = np.nan

    active = pd.DataFrame({"active_elapsed_s": ref["elapsed_s"].round(0).astype(int)})
    for sensor in PRODUCT_SENSORS:
        valid_irr = irr.dropna(subset=["active_elapsed_s", sensor]).sort_values("active_elapsed_s")
        valid_ref = ref.dropna(subset=[sensor]).sort_values("elapsed_s")
        x = active["active_elapsed_s"].to_numpy(dtype=float)
        active[f"reference_{sensor}"] = np.interp(x, valid_ref["elapsed_s"], valid_ref[sensor])
        active[f"irregular_{sensor}"] = np.interp(x, valid_irr["active_elapsed_s"], valid_irr[sensor])
        active[f"delta_{sensor}"] = active[f"irregular_{sensor}"] - active[f"reference_{sensor}"]
    return active


def product_delta_summary(aligned: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for sensor in PRODUCT_SENSORS:
        delta = aligned[f"delta_{sensor}"].dropna()
        rows.append(
            {
                "sensor": sensor,
                "mean_delta_C": delta.mean(),
                "median_delta_C": delta.median(),
                "mean_abs_delta_C": delta.abs().mean(),
                "p95_abs_delta_C": delta.abs().quantile(0.95),
                "max_abs_delta_C": delta.abs().max(),
            }
        )
    return pd.DataFrame(rows)


def hotspot_table(temp: pd.DataFrame, run_name: str) -> pd.DataFrame:
    rows = []
    baseline = temp[TEMP_SENSORS].median(axis=1)
    product_baseline = temp[PRODUCT_SENSORS].median(axis=1)
    for sensor in TEMP_SENSORS:
        delta_all = temp[sensor] - baseline
        delta_prod = temp[sensor] - product_baseline if sensor in PRODUCT_SENSORS else pd.Series(index=temp.index, dtype=float)
        rows.append(
            {
                "run": run_name,
                "sensor": sensor,
                "x_mm": COORDS[sensor][0],
                "y_mm": COORDS[sensor][1],
                "mean_C": temp[sensor].mean(),
                "mean_vs_all_sensor_median_C": delta_all.mean(),
                "max_vs_all_sensor_median_C": delta_all.max(),
                "pct_time_2C_above_all_median": (delta_all > 2.0).mean(),
                "mean_vs_product_median_C": delta_prod.mean() if sensor in PRODUCT_SENSORS else np.nan,
                "pct_time_1C_above_product_median": (delta_prod > 1.0).mean() if sensor in PRODUCT_SENSORS else np.nan,
            }
        )
    return pd.DataFrame(rows)


def product_spread_series(temp: pd.DataFrame, run_name: str) -> pd.DataFrame:
    out = temp[["time", "elapsed_s", *PRODUCT_SENSORS]].copy()
    out["run"] = run_name
    out["product_mean_C"] = out[PRODUCT_SENSORS].mean(axis=1)
    out["product_min_C"] = out[PRODUCT_SENSORS].min(axis=1)
    out["product_max_C"] = out[PRODUCT_SENSORS].max(axis=1)
    out["product_spread_C"] = out["product_max_C"] - out["product_min_C"]
    out["hottest_product_sensor"] = out[PRODUCT_SENSORS].idxmax(axis=1)
    return out


def write_df(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1, freeze: bool = True) -> None:
    for col_idx, col in enumerate(df.columns, start_col):
        cell = ws.cell(start_row, col_idx, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, val in enumerate(row, start_col):
            if isinstance(val, (np.floating, float)) and not pd.isna(val):
                val = float(val)
            elif isinstance(val, (np.integer, int)) and not pd.isna(val):
                val = int(val)
            elif pd.isna(val):
                val = None
            ws.cell(row_idx, col_idx, val)
    if freeze:
        ws.freeze_panes = ws.cell(start_row + 1, start_col).coordinate
    for col_idx, col in enumerate(df.columns, start_col):
        max_len = min(max(len(str(col)), *(len(str(v)) for v in df[col].head(200).fillna("").tolist())) + 2, 28)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top")
                if isinstance(cell.value, float):
                    cell.number_format = "0.000"
    wb.save(path)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / "aasted3_run_comparison.xlsx"

    irregular = read_run(IRREGULAR_CSV)
    reference = read_run(REFERENCE_CSV)
    irr_temp = temp_rows(irregular)
    ref_temp = temp_rows(reference)
    irr_motion = motion_seconds(irregular)
    ref_motion = motion_seconds(reference)

    ref_threshold = ref_motion["roll30_motion_std"].quantile(0.05)
    raw_stop_segments = contiguous_segments(
        irr_motion["sec_abs"], irr_motion["roll30_motion_std"] < ref_threshold, min_len=20
    )
    merged_stop_segments = merge_segments(raw_stop_segments, max_gap=75)
    main_cluster = max(merged_stop_segments, key=lambda item: item[2])
    core_segments = [seg for seg in raw_stop_segments if seg[0] >= main_cluster[0] and seg[1] <= main_cluster[1]]
    core_segment = max(core_segments, key=lambda item: item[2])

    run_duration_extra = (irregular["time"].max() - irregular["time"].min()) - (
        reference["time"].max() - reference["time"].min()
    )

    # The low-motion cluster is broad; this sub-window is the portion indicated by the signal and
    # approximately accounts for the excess runtime without removing the later restart settling tail.
    stop_start = max(float(main_cluster[0]), float(core_segment[0] - 121))
    stop_end = min(float(core_segment[1]), stop_start + 281.0)

    stop_rows = []
    for label, segs in [
        ("raw low-motion segment", raw_stop_segments),
        ("merged low-motion cluster", merged_stop_segments),
    ]:
        for start, end, duration in segs:
            sub = irr_motion[(irr_motion["sec_abs"] >= start) & (irr_motion["sec_abs"] <= end)]
            stop_rows.append(
                {
                    "type": label,
                    "start_time_s": start,
                    "end_time_s": end,
                    "duration_s": duration,
                    "median_roll30_motion_std": sub["roll30_motion_std"].median(),
                    "median_acc_z_std": sub["accz_std"].median(),
                    "median_gyro_y_std": sub["gyroy_std"].median(),
                    "note": "main stop cluster" if (start, end, duration) == main_cluster else "",
                }
            )
    stop_rows.append(
        {
            "type": "recommended stop window for active-time alignment",
            "start_time_s": stop_start,
            "end_time_s": stop_end,
            "duration_s": stop_end - stop_start,
            "median_roll30_motion_std": irr_motion[
                (irr_motion["sec_abs"] >= stop_start) & (irr_motion["sec_abs"] <= stop_end)
            ]["roll30_motion_std"].median(),
            "median_acc_z_std": irr_motion[
                (irr_motion["sec_abs"] >= stop_start) & (irr_motion["sec_abs"] <= stop_end)
            ]["accz_std"].median(),
            "median_gyro_y_std": irr_motion[
                (irr_motion["sec_abs"] >= stop_start) & (irr_motion["sec_abs"] <= stop_end)
            ]["gyroy_std"].median(),
            "note": "contains the suspected 1415-1696 s stop-like interval",
        }
    )
    stop_df = pd.DataFrame(stop_rows)

    product_stats = pd.concat(
        [summarize_product(irr_temp, "irregular_119-2290s"), summarize_product(ref_temp, "reference_2291-4160")],
        ignore_index=True,
    )
    aligned = aligned_product_comparison(irr_temp, ref_temp, stop_start, stop_end)
    delta_summary = product_delta_summary(aligned)
    hotspots = pd.concat(
        [hotspot_table(irr_temp, "irregular_119-2290s"), hotspot_table(ref_temp, "reference_2291-4160")],
        ignore_index=True,
    )
    spread = pd.concat(
        [product_spread_series(irr_temp, "irregular_119-2290s"), product_spread_series(ref_temp, "reference_2291-4160")],
        ignore_index=True,
    )
    spread_summary = (
        spread.groupby("run")
        .agg(
            product_mean_C=("product_mean_C", "mean"),
            mean_product_spread_C=("product_spread_C", "mean"),
            p95_product_spread_C=("product_spread_C", lambda s: s.quantile(0.95)),
            max_product_spread_C=("product_spread_C", "max"),
        )
        .reset_index()
    )

    motion_summary = []
    for run_name, motion in [("irregular_119-2290s", irr_motion), ("reference_2291-4160", ref_motion)]:
        motion_summary.append(
            {
                "run": run_name,
                "seconds": len(motion),
                "median_motion_std_sum": motion["motion_std_sum"].median(),
                "p05_roll30_motion_std": motion["roll30_motion_std"].quantile(0.05),
                "p50_roll30_motion_std": motion["roll30_motion_std"].quantile(0.50),
                "p95_roll30_motion_std": motion["roll30_motion_std"].quantile(0.95),
                "median_acc_z_std": motion["accz_std"].median(),
                "median_gyro_y_std": motion["gyroy_std"].median(),
            }
        )
    motion_summary = pd.DataFrame(motion_summary)

    summary = pd.DataFrame(
        [
            ["Reference file", REFERENCE_CSV.name],
            ["Irregular file", IRREGULAR_CSV.name],
            ["Reference duration_s", reference["time"].max() - reference["time"].min()],
            ["Irregular duration_s", irregular["time"].max() - irregular["time"].min()],
            ["Irregular extra wall-time_s", run_duration_extra],
            ["Reference-derived low-motion threshold", ref_threshold],
            ["Main low-motion cluster_s", f"{main_cluster[0]} to {main_cluster[1]}"],
            ["Stationary core_s", f"{core_segment[0]} to {core_segment[1]}"],
            ["Recommended stop window_s", f"{stop_start:.0f} to {stop_end:.0f}"],
            ["Recommended stop duration_s", stop_end - stop_start],
            ["Product sensors compared", ", ".join(PRODUCT_SENSORS)],
            ["Mould dimensions_mm", "width 112.2, height 33.3"],
        ],
        columns=["metric", "value"],
    )
    coords = pd.DataFrame(
        [{"sensor": sensor, "x_mm": xy[0], "y_mm": xy[1], "used_as_product": sensor in PRODUCT_SENSORS} for sensor, xy in COORDS.items()]
    )

    # Downsample for readable workbook charts.
    aligned_chart = aligned.iloc[::5].copy()
    motion_chart = irr_motion[["sec_abs", "roll30_motion_std", "accz_std", "gyroy_std"]].iloc[::2].copy()
    ref_motion_chart = ref_motion[["sec_abs", "roll30_motion_std", "accz_std", "gyroy_std"]].iloc[::2].copy()

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    sheets = {
        "Summary": summary,
        "Stop Detection": stop_df,
        "Product Stats": product_stats,
        "Product Delta": delta_summary,
        "Product Spread": spread_summary,
        "Hotspots": hotspots,
        "Sensor Coordinates": coords,
        "Aligned Series": aligned_chart,
        "Irregular Motion": motion_chart,
        "Reference Motion": ref_motion_chart,
    }

    for sheet_name, data in sheets.items():
        ws = wb.create_sheet(sheet_name)
        write_df(ws, data)

    ws = wb["Product Delta"]
    if ws.max_row > 1:
        ws.conditional_formatting.add(
            f"B2:E{ws.max_row}",
            ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="F8696B"),
        )

    ws = wb["Hotspots"]
    if ws.max_row > 1:
        ws.conditional_formatting.add(
            f"E2:J{ws.max_row}",
            ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="F8696B"),
        )

    # Native Excel charts.
    ws = wb["Aligned Series"]
    chart = LineChart()
    chart.title = "Product Temperature: Reference vs Irregular, Active-Time Aligned"
    chart.y_axis.title = "deg C"
    chart.x_axis.title = "active elapsed s"
    chart.height = 9
    chart.width = 22
    data = Reference(ws, min_col=2, max_col=11, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "N2")

    ws = wb["Irregular Motion"]
    chart2 = LineChart()
    chart2.title = "Irregular Run Motion Variability"
    chart2.y_axis.title = "rolling/std metric"
    chart2.x_axis.title = "time s"
    chart2.height = 8
    chart2.width = 18
    data2 = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=ws.max_row)
    cats2 = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    ws.add_chart(chart2, "F2")

    ws = wb["Reference Motion"]
    chart3 = LineChart()
    chart3.title = "Reference Run Motion Variability"
    chart3.y_axis.title = "rolling/std metric"
    chart3.x_axis.title = "time s"
    chart3.height = 8
    chart3.width = 18
    data3 = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=ws.max_row)
    cats3 = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    ws.add_chart(chart3, "F2")

    wb.save(out_path)
    style_workbook(out_path)
    print(out_path)


if __name__ == "__main__":
    main()
