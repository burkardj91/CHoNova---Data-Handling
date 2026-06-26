from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


WORKSPACE = Path(r"C:\Users\BurkardJohannes\Documents\CHoNova _ Data Understanding")
OUTPUT_DIR = WORKSPACE / "outputs" / "aasted3_run_comparison"
COMPARISON_CSV = WORKSPACE / "inputs" / "aasted3" / "17_3_f_m_aa3-25_06_26-2160-4374s.csv"
IRREGULAR_CSV = COMPARISON_CSV  # Backward-compatible import name for helper scripts.
REFERENCE_CSV = WORKSPACE / "inputs" / "aasted3" / "reference_2291_4160.csv"
REFERENCE_RUN_NAME = "reference_2291-4160"
COMPARISON_RUN_NAME = "17_3_f_m_aa3-25_06_26-2160-4374s"
PARAMETER_SUMMARY_PATTERNS = ["*parameter_summary*.xlsx", "*parameter_summary*.csv", "*experimental_summary*.xlsx", "*experimental summary*.xlsx"]
EXPERIMENTAL_SETUP_PATTERNS = ["*experimental_setup*.xlsx", "*experimental setup*.xlsx"]
TEMPERATURE_CORRECTION_CSV = WORKSPACE / "inputs" / "aasted3" / "aasted3_temperature_correction_coefficients.csv"
CORRECTED_RAW_DIR = WORKSPACE / "inputs" / "aasted3" / "temperature_corrected_raw"
US_CHANNELS = ["Rx1Tx1", "Rx2Tx2"]
ALL_US_CHANNELS = ["Rx1Tx1", "Rx1Tx2", "Rx2Tx1", "Rx2Tx2"]
TEMP_CORRECTION_SENSOR = "T7"


def safe_filename(value: str) -> str:
    keep = [ch if ch.isalnum() else "_" for ch in str(value)]
    return "_".join("".join(keep).split("_")).strip("_")


REFERENCE_FIGURE_PREFIX = safe_filename(REFERENCE_RUN_NAME)
COMPARISON_FIGURE_PREFIX = safe_filename(COMPARISON_RUN_NAME)
BASE_REPORT_WORKBOOK = OUTPUT_DIR / "aasted3_pattern_detected_hotspot_report_v2.xlsx"
ULTRASOUND_REPORT_WORKBOOK = OUTPUT_DIR / "aasted3_pattern_detected_hotspot_report_with_ultrasound_figures.xlsx"
FINAL_REPORT_WORKBOOK = OUTPUT_DIR / f"aasted3_{COMPARISON_FIGURE_PREFIX}_comparison_report.xlsx"
RUN_OUTPUT_DIR = OUTPUT_DIR / "runs" / COMPARISON_FIGURE_PREFIX

US_CORRECTION_COEFFICIENTS = {
    "Rx1Tx1": [0.6952085988953896, -0.021822771399044455, 0.0037091753271136572, -0.0002082564521397378, 3.245923482062673e-06],
    "Rx1Tx2": [-0.0692884241972798, 0.010234319131273829, -0.0010987231115144628, 4.696490232034503e-05, -6.960118543105261e-07],
    "Rx2Tx1": [-0.048660102229808155, 0.006857199971676996, -0.0007691346333861185, 3.385840046021632e-05, -5.129374107408218e-07],
    "Rx2Tx2": [0.8917009154106912, 0.037810097379146014, -0.00424490665181043, 0.0001020512497149748, -7.465580911868695e-07],
}

# Product temperature sensors. User correction listed "T2, T4, T4, T5, T7";
# treat the duplicated T4 as T3 to keep five unique product locations.
PRODUCT_SENSORS = ["T2", "T3", "T4", "T5", "T7"]
TEMP_SENSORS = [f"T{i}" for i in range(1, 10)]
FEATURES = ["T8", "accz_mean", "accz_std", "gyroy_mean", "gyroy_std"]
IMU_SENSOR_PREFIXES = {
    "acc x": "accx",
    "acc y": "accy",
    "acc z": "accz",
    "gyro x": "gyrox",
    "gyro y": "gyroy",
    "gyro z": "gyroz",
    "mag x": "magx",
    "mag y": "magy",
    "mag z": "magz",
}
COORDS = {
    "T1": (53.0, 23.0),
    "T2": (70.0, 13.0),
    "T3": (5.0, 5.0),
    "T4": (5.0, 31.0),
    "T5": (108.0, 5.0),
    "T6": (76.0, 24.0),
    "T7": (99.0, 30.0),
    "T8": (93.0, 23.0),
    "T9": (55.0, 10.0),
}

ZONES = [
    (0, 50, "movement_to_conditioning", "Mould movement to conditioning zone; all product sensors warm up."),
    (50, 210, "mould_conditioning", "Conditioning zone; acc z pattern changes before/after and product temperatures drop."),
    (210, 318, "deposition_transfer", "Deposition around 256 s and transfer toward vibration/warming section."),
    (318, 472, "vibration_warming", "Vibration in non-cooled/warmed zone; T8 ambient signal and product temperatures plateau high."),
    (472, 524, "transition_to_cooling", "Transition after vibration before the first T8 cooling minimum near ventilator."),
    (524, 608, "cooling_1", "First cooling section from T8 minimum to next section."),
    (608, 756, "cooling_2", "Second periodic cooling section."),
    (756, 898, "cooling_3", "Third periodic cooling section."),
    (898, 1032, "cooling_4", "Fourth periodic cooling section."),
    (1032, 1638, "less_periodic_cooling_2", "Less periodic cooling section."),
    (1638, 1786, "transition_2", "Transition period before demoulding."),
    (1786, 1802, "demoulding_twisting", "First demoulding subphase: acc z shifts from high plateau to low plateau while acc x shows a negative parabolic movement."),
    (1802, 1850, "demoulding_vibration", "Second demoulding subphase: vibration/shaking visible in acc z, acc x/y, and gyro y."),
    (1850, 99999, "final_demoulding", "Final demoulding/run-out after twisting and vibration."),
]


def correction_polynomial(channel: str, temperature_c: pd.Series | np.ndarray) -> pd.Series:
    coeffs = US_CORRECTION_COEFFICIENTS[channel]
    temperature = pd.Series(temperature_c, dtype=float)
    out = pd.Series(0.0, index=temperature.index)
    for power, coeff in enumerate(coeffs):
        out = out + coeff * (temperature**power)
    return out


def add_temperature_corrected_us(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if TEMP_CORRECTION_SENSOR not in df.columns:
        return df
    temp = df[["time", TEMP_CORRECTION_SENSOR]].dropna().sort_values("time")
    if temp.empty:
        return df
    df[f"{TEMP_CORRECTION_SENSOR}_interp_C"] = np.interp(df["time"], temp["time"], temp[TEMP_CORRECTION_SENSOR])
    for channel in ALL_US_CHANNELS:
        if channel not in df.columns:
            continue
        correction = correction_polynomial(channel, df[f"{TEMP_CORRECTION_SENSOR}_interp_C"])
        df[f"{channel}_temp_correction"] = correction
        df[f"{channel}_tc"] = df[channel] / correction.replace(0, np.nan)
    return df


def read_run(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    df["elapsed_s"] = df["time"] - df["time"].min()
    return add_temperature_corrected_us(df)


def zone_for_elapsed(elapsed: float) -> str:
    for start, end, name, _ in ZONES:
        if start <= elapsed < end:
            return name
    return ZONES[-1][2]


def build_seconds(df: pd.DataFrame) -> pd.DataFrame:
    max_sec = int(np.ceil(df["elapsed_s"].max()))
    seconds = pd.DataFrame({"elapsed_sec": np.arange(0, max_sec + 1)})

    available_imu = [col for col in IMU_SENSOR_PREFIXES if col in df.columns]
    motion = df.dropna(subset=["acc z", "gyro y"], how="all").copy()
    motion["elapsed_sec"] = np.floor(motion["elapsed_s"]).astype(int)
    agg_spec = {"sample_count": ("time", "count")}
    for col in available_imu:
        prefix = IMU_SENSOR_PREFIXES[col]
        agg_spec[f"{prefix}_mean"] = (col, "mean")
        agg_spec[f"{prefix}_std"] = (col, "std")
    motion_sec = motion.groupby("elapsed_sec").agg(**agg_spec).reset_index()
    seconds = seconds.merge(motion_sec, on="elapsed_sec", how="left")

    temp = df.dropna(subset=["T8"], how="all")[["elapsed_s", "T8"]].sort_values("elapsed_s")
    seconds["T8"] = np.interp(seconds["elapsed_sec"], temp["elapsed_s"], temp["T8"])
    for col in [c for c in seconds.columns if c.endswith("_mean")]:
        seconds[col] = seconds[col].interpolate(limit_direction="both")
    for col in [c for c in seconds.columns if c.endswith("_std")]:
        seconds[col] = seconds[col].fillna(seconds[col].median())
    seconds["zone_reference"] = seconds["elapsed_sec"].map(zone_for_elapsed)
    seconds["T8_smooth"] = seconds["T8"].rolling(15, center=True, min_periods=1).median()
    seconds["gyroy_std_roll"] = seconds["gyroy_std"].rolling(12, center=True, min_periods=3).median()
    seconds["accz_std_roll"] = seconds["accz_std"].rolling(12, center=True, min_periods=3).median()
    return seconds


def _segments_from_mask(seconds: pd.Series, mask: pd.Series, min_len: int = 5) -> list[tuple[int, int, int]]:
    segs = []
    start = None
    prev = None
    for sec, hit in zip(seconds.astype(int), mask.fillna(False).astype(bool)):
        if hit and start is None:
            start = int(sec)
        if start is not None and (not hit or (prev is not None and int(sec) != prev + 1)):
            end = int(prev)
            if end - start + 1 >= min_len:
                segs.append((start, end, end - start + 1))
            start = int(sec) if hit else None
        prev = int(sec)
    if start is not None and prev is not None and int(prev) - start + 1 >= min_len:
        segs.append((start, int(prev), int(prev) - start + 1))
    return segs


def _t8_valleys(seconds: pd.DataFrame, start_after: int, min_distance: int = 80) -> list[tuple[int, float]]:
    s = seconds[seconds["elapsed_sec"] >= start_after].copy()
    vals = s["T8_smooth"].to_numpy()
    secs = s["elapsed_sec"].to_numpy()
    raw = []
    for i in range(8, len(vals) - 8):
        window = vals[i - 8 : i + 9]
        if vals[i] <= np.nanmin(window) + 1e-9 and vals[i] < vals[i - 8] and vals[i] < vals[i + 8]:
            raw.append((int(secs[i]), float(vals[i])))
    grouped = []
    for sec, val in raw:
        if not grouped or sec - grouped[-1][-1][0] > 3:
            grouped.append([(sec, val)])
        else:
            grouped[-1].append((sec, val))
    candidates = []
    for group in grouped:
        candidates.append(min(group, key=lambda item: item[1]))
    selected = []
    for sec, val in candidates:
        if not selected or sec - selected[-1][0] >= min_distance:
            selected.append((sec, val))
        elif val < selected[-1][1]:
            selected[-1] = (sec, val)
    return selected


def detect_pattern_zones(seconds: pd.DataFrame, run_name: str) -> pd.DataFrame:
    high = seconds["gyroy_std_roll"] > 1.0
    high_segments = _segments_from_mask(seconds["elapsed_sec"], high, min_len=5)
    vib_candidates = [seg for seg in high_segments if 250 <= seg[0] <= 650]
    vibration = max(vib_candidates, key=lambda item: item[2]) if vib_candidates else (320, 450, 130)
    dep_candidates = [seg for seg in high_segments if 150 <= seg[0] < vibration[0]]
    deposition_event = dep_candidates[-1] if dep_candidates else (max(0, vibration[0] - 75), vibration[0] - 50, 25)

    before_vib = seconds[(seconds["elapsed_sec"] >= 40) & (seconds["elapsed_sec"] < vibration[0])]
    high_t8_idx = before_vib["T8_smooth"].idxmax() if not before_vib.empty else seconds.index[0]
    high_t8_time = int(seconds.loc[high_t8_idx, "elapsed_sec"])
    deposition_transfer_start = min(deposition_event[0] - 30, high_t8_time + 20)
    deposition_transfer_start = max(60, int(deposition_transfer_start))

    early = seconds[(seconds["elapsed_sec"] >= 10) & (seconds["elapsed_sec"] < deposition_transfer_start)]
    baseline = float(seconds.loc[seconds["elapsed_sec"] <= 10, "T8_smooth"].median())
    warm_cross = early[early["T8_smooth"] >= baseline + 1.0]
    early_imu = early[early["gyroy_std_roll"] > 0.5]
    movement_end_options = []
    if not warm_cross.empty:
        movement_end_options.append(int(warm_cross["elapsed_sec"].iloc[0]))
    if not early_imu.empty:
        movement_end_options.append(int(early_imu["elapsed_sec"].iloc[0]))
    movement_end = min(movement_end_options) if movement_end_options else max(20, deposition_transfer_start - 160)

    valleys = _t8_valleys(seconds, start_after=vibration[1], min_distance=80)
    # First five valleys define transition-to-cooling plus four periodic cooling sections.
    while len(valleys) < 5:
        last = valleys[-1][0] if valleys else vibration[1] + 60
        valleys.append((last + 120, np.nan))
    cooling_start, c2, c3, c4, less_periodic_start = [v[0] for v in valleys[:5]]

    # Use the last genuinely cold T8 valley before the final rise. Small late wiggles
    # around 13-14 C are transition behavior, not the start of transition_2.
    low_valley_limit = float(seconds["T8_smooth"].quantile(0.20) + 1.0)
    late_valleys = [v for v in valleys if less_periodic_start < v[0] and v[1] <= low_valley_limit]
    transition2_start = late_valleys[-1][0] if late_valleys else less_periodic_start

    # Demoulding starts at the sharp acc-z regime change after the last T8-minimum
    # transition, not merely when T8 begins rising.
    accz_med = seconds["accz_mean"].rolling(30, center=True, min_periods=5).median()
    stable = seconds[
        (seconds["elapsed_sec"] >= transition2_start)
        & (seconds["elapsed_sec"] <= min(transition2_start + 70, int(seconds["elapsed_sec"].max())))
    ]
    baseline = float(stable["accz_mean"].median()) if not stable.empty else float(seconds["accz_mean"].median())
    search = seconds[seconds["elapsed_sec"] >= transition2_start + 20].copy()
    search["accz_med"] = accz_med.loc[search.index]
    search["accz_shift"] = (search["accz_med"] - baseline).abs()
    candidates = search[
        (search["accz_shift"] > 0.35)
        | ((search["accz_mean"] - baseline).abs() > 0.55)
        | (search["accz_std_roll"] > 0.20)
    ]
    if not candidates.empty:
        demoulding_start = int(candidates["elapsed_sec"].iloc[0])
    else:
        late_rise = seconds[(seconds["elapsed_sec"] > transition2_start) & (seconds["T8_smooth"] >= 14.0)]
        demoulding_start = int(late_rise["elapsed_sec"].iloc[0]) if not late_rise.empty else int(seconds["elapsed_sec"].max() - 90)
    run_end = int(seconds["elapsed_sec"].max())

    demould_search = seconds[
        (seconds["elapsed_sec"] >= demoulding_start)
        & (seconds["elapsed_sec"] <= min(run_end, demoulding_start + 120))
    ].copy()
    pre_demould = seconds[
        (seconds["elapsed_sec"] >= max(0, demoulding_start - 45))
        & (seconds["elapsed_sec"] < demoulding_start)
    ].copy()
    pre_accx = float(pre_demould["accx_mean"].median()) if "accx_mean" in pre_demould and not pre_demould.empty else 0.0
    pre_accz = float(pre_demould["accz_mean"].median()) if "accz_mean" in pre_demould and not pre_demould.empty else baseline
    if not demould_search.empty:
        demould_search["accx_med"] = demould_search["accx_mean"].rolling(5, center=True, min_periods=2).median()
        demould_search["accz_med"] = demould_search["accz_mean"].rolling(5, center=True, min_periods=2).median()
        demould_search["motion_std"] = (
            demould_search[["accx_std", "accy_std", "accz_std", "gyroy_std"]]
            .fillna(0)
            .abs()
            .sum(axis=1)
        )
        low_accz = demould_search["accz_med"] < (pre_accz - 0.8)
        accx_returned = (demould_search["accx_med"] - pre_accx).abs() < 0.18
        twist_hits = demould_search[
            (demould_search["elapsed_sec"] >= demoulding_start + 6)
            & low_accz
            & accx_returned
        ]
        if not twist_hits.empty:
            twisting_end = int(twist_hits["elapsed_sec"].iloc[0])
        else:
            twisting_end = min(run_end, demoulding_start + 16)

        vib_candidate = demould_search[demould_search["elapsed_sec"] >= twisting_end].copy()
        vib_baseline = float(pre_demould[["accx_std", "accy_std", "accz_std"]].sum(axis=1).median()) if not pre_demould.empty else 0.10
        vib_mask = (
            (vib_candidate["accz_std_roll"] > max(0.15, float(seconds["accz_std_roll"].quantile(0.80))))
            | (vib_candidate["motion_std"] > max(0.45, vib_baseline * 4.0))
            | (vib_candidate["gyroy_std"] > 2.0)
        )
        vib_segments = _segments_from_mask(vib_candidate["elapsed_sec"], vib_mask, min_len=5)
        vib_segments = [seg for seg in vib_segments if seg[0] <= twisting_end + 55]
        if vib_segments:
            vibration_start = twisting_end
            vibration_end = min(run_end, vib_segments[-1][1] + 1)
        else:
            vibration_start = twisting_end
            vibration_end = min(run_end, twisting_end + 48)
    else:
        twisting_end = min(run_end, demoulding_start + 16)
        vibration_start = twisting_end
        vibration_end = min(run_end, twisting_end + 48)

    twisting_end = max(demoulding_start, min(int(twisting_end), run_end))
    vibration_start = max(twisting_end, min(int(vibration_start), run_end))
    vibration_end = max(vibration_start, min(int(vibration_end), run_end))

    boundaries = [
        (0, movement_end, "movement_to_conditioning", "first warm-up/IMU-change landmark", "medium" if movement_end_options else "low"),
        (movement_end, deposition_transfer_start, "mould_conditioning", "T8 high/conditioning phase before deposition event", "medium"),
        (deposition_transfer_start, vibration[0], "deposition_transfer", "pre-vibration deposition/transfer IMU event", "medium"),
        (vibration[0], vibration[1], "vibration_warming", "longest sustained high gyro-y variability segment", "high"),
        (vibration[1], cooling_start, "transition_to_cooling", "post-vibration transition until first T8 valley", "high"),
        (cooling_start, c2, "cooling_1", "T8 valley-to-valley cooling cycle", "high"),
        (c2, c3, "cooling_2", "T8 valley-to-valley cooling cycle", "high"),
        (c3, c4, "cooling_3", "T8 valley-to-valley cooling cycle", "high"),
        (c4, less_periodic_start, "cooling_4", "T8 valley-to-valley cooling cycle", "high"),
        (less_periodic_start, transition2_start, "less_periodic_cooling_2", "after periodic valleys until the last cold T8 minimum", "medium"),
        (transition2_start, demoulding_start, "transition_2", "last cold T8 minimum until sharp acc-z regime change", "medium"),
        (demoulding_start, twisting_end, "demoulding_twisting", "acc-z plateau drop with acc-x negative parabolic movement", "medium"),
        (vibration_start, vibration_end, "demoulding_vibration", "high IMU variability during demoulding vibration/shaking", "medium"),
        (vibration_end, run_end, "final_demoulding", "after twisting/vibration subphases", "medium"),
    ]

    rows = []
    previous_end = 0
    for start, end, zone, basis, confidence in boundaries:
        start = max(int(start), previous_end)
        end = max(start, min(int(end), run_end))
        rows.append(
            {
                "run": run_name,
                "zone": zone,
                "detected_start_s": start,
                "detected_end_s": end,
                "duration_s": end - start,
                "detection_basis": basis,
                "confidence": confidence,
            }
        )
        previous_end = end
    return pd.DataFrame(rows)


def dtw_map(irregular: pd.DataFrame, reference: pd.DataFrame) -> tuple[pd.DataFrame, float]:
    x = irregular[FEATURES].to_numpy(dtype=float)
    y = reference[FEATURES].to_numpy(dtype=float)
    combined = np.vstack([x, y])
    mean = combined.mean(axis=0)
    std = combined.std(axis=0)
    std[std == 0] = 1
    x = (x - mean) / std
    y = (y - mean) / std

    n, m = len(x), len(y)
    prev = np.full(m + 1, np.inf)
    curr = np.full(m + 1, np.inf)
    prev[0] = 0.0
    direction = np.zeros((n + 1, m + 1), dtype=np.uint8)
    for i in range(1, n + 1):
        curr[0] = np.inf
        xi = x[i - 1]
        for j in range(1, m + 1):
            cost = float(np.linalg.norm(xi - y[j - 1]))
            choices = (prev[j - 1], prev[j], curr[j - 1])
            best = int(np.argmin(choices))
            curr[j] = cost + choices[best]
            direction[i, j] = best
        prev, curr = curr, prev

    i, j = n, m
    pairs = []
    while i > 0 and j > 0:
        pairs.append((i - 1, j - 1))
        step = direction[i, j]
        if step == 0:
            i -= 1
            j -= 1
        elif step == 1:
            i -= 1
        else:
            j -= 1
    pairs.reverse()

    pair_df = pd.DataFrame(pairs, columns=["irregular_idx", "reference_idx"])
    map_df = (
        pair_df.groupby("irregular_idx")["reference_idx"]
        .median()
        .round()
        .astype(int)
        .reset_index()
    )
    map_df["irregular_elapsed_s"] = map_df["irregular_idx"]
    map_df["reference_elapsed_s"] = map_df["reference_idx"]
    map_df["reference_zone"] = map_df["reference_elapsed_s"].map(zone_for_elapsed)
    return map_df, float(prev[m])


def temp_with_zones(df: pd.DataFrame, run_name: str, map_df: pd.DataFrame | None = None) -> pd.DataFrame:
    temp = df.dropna(subset=TEMP_SENSORS, how="all").copy()
    temp["run"] = run_name
    temp["elapsed_sec"] = np.floor(temp["elapsed_s"]).astype(int)
    if map_df is None:
        temp["reference_elapsed_s"] = temp["elapsed_s"]
        temp["zone"] = temp["elapsed_s"].map(zone_for_elapsed)
    else:
        mapper = map_df.set_index("irregular_elapsed_s")
        temp["reference_elapsed_s"] = temp["elapsed_sec"].map(mapper["reference_elapsed_s"])
        temp["zone"] = temp["elapsed_sec"].map(mapper["reference_zone"])
    temp["product_mean_C"] = temp[PRODUCT_SENSORS].mean(axis=1)
    temp["product_spread_C"] = temp[PRODUCT_SENSORS].max(axis=1) - temp[PRODUCT_SENSORS].min(axis=1)
    temp["hottest_product_sensor"] = temp[PRODUCT_SENSORS].idxmax(axis=1)
    return temp


def temp_with_detected_zones(df: pd.DataFrame, run_name: str, detected: pd.DataFrame) -> pd.DataFrame:
    temp = df.dropna(subset=TEMP_SENSORS, how="all").copy()
    temp["run"] = run_name
    temp["zone"] = None
    for row in detected[detected["run"] == run_name].itertuples(index=False):
        mask = (temp["elapsed_s"] >= row.detected_start_s) & (temp["elapsed_s"] < row.detected_end_s)
        temp.loc[mask, "zone"] = row.zone
    temp["zone"] = temp["zone"].fillna(detected[detected["run"] == run_name]["zone"].iloc[-1])
    temp["product_mean_C"] = temp[PRODUCT_SENSORS].mean(axis=1)
    temp["product_spread_C"] = temp[PRODUCT_SENSORS].max(axis=1) - temp[PRODUCT_SENSORS].min(axis=1)
    temp["hottest_product_sensor"] = temp[PRODUCT_SENSORS].idxmax(axis=1)
    return temp


def product_zone_summary(temp: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for (run, zone), group in temp.groupby(["run", "zone"], sort=False):
        for sensor in PRODUCT_SENSORS:
            s = group[sensor].dropna()
            rows.append(
                {
                    "run": run,
                    "zone": zone,
                    "sensor": sensor,
                    "mean_C": s.mean(),
                    "min_C": s.min(),
                    "max_C": s.max(),
                    "std_C": s.std(),
                    "n_temp_points": len(s),
                }
            )
        rows.append(
            {
                "run": run,
                "zone": zone,
                "sensor": "PRODUCT_SPREAD",
                "mean_C": group["product_spread_C"].mean(),
                "min_C": group["product_spread_C"].min(),
                "max_C": group["product_spread_C"].max(),
                "std_C": group["product_spread_C"].std(),
                "n_temp_points": len(group),
            }
        )
    return pd.DataFrame(rows)


def product_zone_delta(summary: pd.DataFrame) -> pd.DataFrame:
    pivot = summary.pivot_table(index=["zone", "sensor"], columns="run", values="mean_C").reset_index()
    pivot.columns.name = None
    if COMPARISON_RUN_NAME in pivot.columns and REFERENCE_RUN_NAME in pivot.columns:
        pivot["mean_delta_comparison_minus_reference_C"] = (
            pivot[COMPARISON_RUN_NAME] - pivot[REFERENCE_RUN_NAME]
        )
    return pivot


def duration_by_zone(reference_seconds: pd.DataFrame, map_df: pd.DataFrame) -> pd.DataFrame:
    ref_counts = reference_seconds.groupby("zone_reference").size().rename("reference_duration_s")
    comp_counts = map_df.groupby("reference_zone").size().rename("comparison_pattern_duration_s")
    out = pd.concat([ref_counts, comp_counts], axis=1).fillna(0).reset_index().rename(columns={"index": "zone"})
    out["extra_duration_s"] = out["comparison_pattern_duration_s"] - out["reference_duration_s"]
    order = {name: i for i, (_, _, name, _) in enumerate(ZONES)}
    out["order"] = out["zone"].map(order)
    return out.sort_values("order").drop(columns=["order"])


def mechanical_zone_summary(seconds: pd.DataFrame, run: str, zone_col: str) -> pd.DataFrame:
    rows = []
    for zone, group in seconds.groupby(zone_col, sort=False):
        row = {
            "run": run,
            "zone": zone,
            "T8_mean_C": group["T8"].mean(),
            "T8_min_C": group["T8"].min(),
            "T8_max_C": group["T8"].max(),
        }
        for prefix in IMU_SENSOR_PREFIXES.values():
            mean_col = f"{prefix}_mean"
            std_col = f"{prefix}_std"
            if mean_col in group.columns:
                row[f"{prefix}_mean"] = group[mean_col].mean()
                row[f"{prefix}_min"] = group[mean_col].min()
                row[f"{prefix}_max"] = group[mean_col].max()
            if std_col in group.columns:
                row[f"{prefix}_std_median"] = group[std_col].median()
                row[f"{prefix}_std_max"] = group[std_col].max()
        rows.append(row)
    return pd.DataFrame(rows)


def mechanical_delta(reference_seconds: pd.DataFrame, irregular_seconds: pd.DataFrame, map_df: pd.DataFrame) -> pd.DataFrame:
    comp = irregular_seconds.copy()
    mapper = map_df.set_index("irregular_elapsed_s")
    comp["reference_zone"] = comp["elapsed_sec"].map(mapper["reference_zone"])
    summary = pd.concat(
        [
            mechanical_zone_summary(reference_seconds, REFERENCE_RUN_NAME, "zone_reference"),
            mechanical_zone_summary(comp, COMPARISON_RUN_NAME, "reference_zone"),
        ],
        ignore_index=True,
    )
    metrics = [
        col
        for col in summary.columns
        if col not in {"run", "zone"} and pd.api.types.is_numeric_dtype(summary[col])
    ]
    pivot = summary.pivot_table(index="zone", columns="run", values=metrics)
    rows = []
    for zone in pivot.index:
        row = {"zone": zone}
        notable = []
        for metric in metrics:
            ref = pivot.loc[zone, (metric, REFERENCE_RUN_NAME)] if (metric, REFERENCE_RUN_NAME) in pivot else np.nan
            comp_val = pivot.loc[zone, (metric, COMPARISON_RUN_NAME)] if (metric, COMPARISON_RUN_NAME) in pivot else np.nan
            row[f"reference_{metric}"] = ref
            row[f"comparison_{metric}"] = comp_val
            row[f"delta_{metric}"] = comp_val - ref
            if pd.notna(ref) and pd.notna(comp_val):
                if metric.startswith("T8") and abs(comp_val - ref) >= 1.0:
                    notable.append(metric)
                if metric.startswith(("accz", "gyroy")) and abs(comp_val - ref) >= max(0.05, abs(ref) * 0.35):
                    notable.append(metric)
        row["notable_differences"] = ", ".join(notable)
        rows.append(row)
    order = {name: i for i, (_, _, name, _) in enumerate(ZONES)}
    return pd.DataFrame(rows).assign(order=lambda d: d["zone"].map(order)).sort_values("order").drop(columns=["order"])


def hotspot_by_zone(temp: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for (run, zone), group in temp.groupby(["run", "zone"], sort=False):
        product_median = group[PRODUCT_SENSORS].median(axis=1)
        for sensor in PRODUCT_SENSORS:
            delta = group[sensor] - product_median
            rows.append(
                {
                    "run": run,
                    "zone": zone,
                    "sensor": sensor,
                    "x_mm": COORDS[sensor][0],
                    "y_mm": COORDS[sensor][1],
                    "mean_vs_product_median_C": delta.mean(),
                    "pct_time_1C_above_product_median": (delta > 1.0).mean(),
                    "pct_time_1C_below_product_median": (delta < -1.0).mean(),
                }
            )
    return pd.DataFrame(rows)


def clearer_hotspots(temp: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    rows = []
    matrix_rows = []
    for (run, zone), group in temp.groupby(["run", "zone"], sort=False):
        means = group[PRODUCT_SENSORS].mean()
        zone_mean = means.mean()
        hottest = means.idxmax()
        coolest = means.idxmin()
        spread = means[hottest] - means[coolest]
        rows.append(
            {
                "run": run,
                "zone": zone,
                "zone_product_mean_C": zone_mean,
                "hottest_sensor": hottest,
                "hottest_mean_C": means[hottest],
                "coolest_sensor": coolest,
                "coolest_mean_C": means[coolest],
                "spread_hottest_minus_coolest_C": spread,
                "spread_severity": "high" if spread >= 1.5 else "moderate" if spread >= 0.75 else "low",
                "spread_color": "#F8696B" if spread >= 1.5 else "#FFEB84" if spread >= 0.75 else "#63BE7B",
                "plain_language": f"{hottest} is warmest and {coolest} is coolest in this zone.",
            }
        )
        for sensor in PRODUCT_SENSORS:
            sensor_delta = means[sensor] - zone_mean
            sensor_class = "hotspot" if sensor_delta >= 0.5 else "coolspot" if sensor_delta <= -0.5 else "neutral"
            rows[-1][f"{sensor}_mean_C"] = means[sensor]
            rows[-1][f"{sensor}_vs_zone_mean_C"] = sensor_delta
            rows[-1][f"{sensor}_thermal_class"] = sensor_class
            rows[-1][f"{sensor}_color"] = "#F8696B" if sensor_class == "hotspot" else "#4472C4" if sensor_class == "coolspot" else "#FFFFFF"
            matrix_rows.append(
                {
                    "run": run,
                    "zone": zone,
                    "sensor": sensor,
                    "x_mm": COORDS[sensor][0],
                    "y_mm": COORDS[sensor][1],
                    "mean_C": means[sensor],
                    "delta_vs_zone_product_mean_C": sensor_delta,
                    "thermal_class": sensor_class,
                    "color_code": "#F8696B" if sensor_class == "hotspot" else "#4472C4" if sensor_class == "coolspot" else "#FFFFFF",
                }
            )
    hotspot_summary = pd.DataFrame(rows)
    hotspot_long = pd.DataFrame(matrix_rows)
    delta_matrix = hotspot_long.pivot_table(
        index=["run", "zone"], columns="sensor", values="delta_vs_zone_product_mean_C"
    ).reset_index()
    delta_matrix.columns.name = None
    return hotspot_summary, hotspot_long, delta_matrix


def change_scores(t: np.ndarray, y: np.ndarray, window_points: int) -> np.ndarray:
    scores = np.full(len(y), np.nan)
    for i in range(window_points, len(y) - window_points):
        left = y[i - window_points : i]
        right = y[i : i + window_points]
        denom = np.nanstd(np.r_[left, right])
        if not np.isfinite(denom) or denom <= 1e-9:
            denom = 1e-9
        scores[i] = (np.nanmean(right) - np.nanmean(left)) / denom
    return scores


def detect_us_change_points(seg: pd.DataFrame, channel: str) -> list[dict[str, float]]:
    valid = seg[["elapsed_s", channel]].dropna().sort_values("elapsed_s")
    if len(valid) < 12:
        return []
    t = valid["elapsed_s"].to_numpy(float)
    y = valid[channel].rolling(5, center=True, min_periods=2).median().to_numpy(float)
    dt = np.nanmedian(np.diff(t))
    window_points = max(4, int(round(45.0 / max(dt, 1.0))))
    scores = change_scores(t, y, window_points)
    positive = scores[np.isfinite(scores) & (scores > 0)]
    threshold = np.nanpercentile(positive, 80) if len(positive) else np.inf
    candidates = []
    for idx in np.argsort(np.nan_to_num(scores, nan=-np.inf))[::-1]:
        if not np.isfinite(scores[idx]) or scores[idx] <= 0 or scores[idx] < max(0.8, threshold):
            continue
        tt = float(t[idx])
        if any(abs(tt - c["change_point_s"]) < 60.0 for c in candidates):
            continue
        before = float(np.nanmedian(y[max(0, idx - window_points) : idx]))
        after = float(np.nanmedian(y[idx : min(len(y), idx + window_points)]))
        delta = after - before
        if delta < 0.025:
            continue
        candidates.append(
            {
                "change_point_s": tt,
                "change_score": float(scores[idx]),
                "local_before_median": before,
                "local_after_median": after,
                "local_delta_after_minus_before": delta,
            }
        )
        if len(candidates) >= 30:
            break
    return sorted(candidates, key=lambda c: c["change_point_s"])


def median_window(df: pd.DataFrame, channel: str, start: float, end: float) -> float:
    w = df[(df["elapsed_s"] >= start) & (df["elapsed_s"] <= end)][channel].dropna()
    if w.empty:
        idx = (df["elapsed_s"] - end).abs().idxmin()
        return float(df.loc[idx, channel])
    return float(w.median())


def us_level_near(seg: pd.DataFrame, channel: str, t0: float, width: float = 5.0) -> float:
    w = seg[(seg["elapsed_s"] >= t0) & (seg["elapsed_s"] <= t0 + width)][channel].dropna()
    return np.nan if w.empty else float(w.median())


def find_parameter_summary_file() -> Path | None:
    input_dir = WORKSPACE / "inputs" / "aasted3"
    for pattern in PARAMETER_SUMMARY_PATTERNS:
        hits = sorted(input_dir.glob(pattern))
        if hits:
            return hits[0]
    return None


def find_experimental_setup_file() -> Path | None:
    input_dir = WORKSPACE / "inputs" / "aasted3"
    for pattern in EXPERIMENTAL_SETUP_PATTERNS:
        hits = sorted(input_dir.glob(pattern))
        if hits:
            return hits[0]
    return None


def _normalized_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [
        str(c).strip().lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "")
        for c in out.columns
    ]
    return out


def load_experimental_setup() -> pd.DataFrame:
    path = find_experimental_setup_file()
    if path is None:
        return pd.DataFrame()
    sheets = pd.read_excel(path, sheet_name=None)
    selected = None
    for df in sheets.values():
        normalized = _normalized_columns(df)
        cols = list(normalized.columns)
        if _find_column(cols, ["quality"]) and _find_column(cols, ["remarks"]):
            selected = normalized
            break
    if selected is None:
        selected = _normalized_columns(next(iter(sheets.values())))
    raw = selected
    raw = _normalized_columns(raw).dropna(how="all")
    columns = list(raw.columns)
    run_col = _find_column(columns, ["experimental", "name"])
    if run_col is None:
        return pd.DataFrame()
    chocolate_col = "chocolate" if "chocolate" in columns else _find_column(columns, ["chocolate"])
    rename_map = {
        run_col: "protocol_run_name",
        _find_column(columns, ["chocolate", "test"]): "protocol_chocolate_test_or_empty_flag",
        _find_column(columns, ["runs", "within"]): "runs_within_trial",
        _find_column(columns, ["date"]): "date",
        _find_column(columns, ["duration"]): "protocol_duration_s",
        chocolate_col: "chocolate",
        _find_column(columns, ["tempering"]): "tempering",
        _find_column(columns, ["remarks"]): "remarks",
        _find_column(columns, ["quality"]): "quality_0_not_good_1_good",
    }
    rename_map = {k: v for k, v in rename_map.items() if k is not None}
    out = raw.rename(columns=rename_map)[list(rename_map.values())].copy()
    out["run"] = out["protocol_run_name"].map(run_lookup_name)
    out["source_file"] = path.name
    return out


def _first_existing(row: pd.Series, candidates: list[str]) -> object:
    for col in candidates:
        if col in row and pd.notna(row[col]):
            return row[col]
    return np.nan


def _find_column(columns: list[str], must_include: list[str], any_include: list[str] | None = None) -> str | None:
    for col in columns:
        text = str(col).lower()
        if all(token in text for token in must_include) and (
            any_include is None or any(token in text for token in any_include)
        ):
            return col
    return None


def load_parameter_landmarks() -> pd.DataFrame:
    path = find_parameter_summary_file()
    if path is None:
        return pd.DataFrame(
            columns=[
                "run",
                "deposition_s",
                "crystallization_onset_s",
                "detachment_onset_s",
                "channel",
                "source_file",
                "input_status",
            ]
        )
    if path.suffix.lower() == ".csv":
        raw = pd.read_csv(path)
    else:
        raw = pd.concat(pd.read_excel(path, sheet_name=None).values(), ignore_index=True)
    raw = _normalized_columns(raw).dropna(how="all")
    columns = list(raw.columns)
    run_col = _find_column(columns, ["experimental", "name"]) or _find_column(columns, ["run"]) or _find_column(columns, ["sample"])
    deposition_col = _find_column(columns, ["deposition"])
    wide_channel_specs = {
        "Rx1Tx1": {
            "crystallization": _find_column(columns, ["crystallization", "start"], ["t1r1", "r1t1", "rx1tx1"]),
            "detachment": _find_column(columns, ["detachment", "onset"], ["t1r1", "r1t1", "rx1tx1"]),
        },
        "Rx2Tx2": {
            "crystallization": _find_column(columns, ["crystallization", "start"], ["t2r2", "r2t2", "rx2tx2"]),
            "detachment": _find_column(columns, ["detachment", "onset"], ["t2r2", "r2t2", "rx2tx2"]),
        },
    }
    rows = []
    if run_col and deposition_col and any(spec["detachment"] for spec in wide_channel_specs.values()):
        for _, row in raw.iterrows():
            run = row.get(run_col)
            deposition = row.get(deposition_col)
            if pd.isna(run) or pd.isna(deposition):
                continue
            for channel, spec in wide_channel_specs.items():
                detachment_col = spec["detachment"]
                if detachment_col is None or pd.isna(row.get(detachment_col)):
                    continue
                crystallization_col = spec["crystallization"]
                rows.append(
                    {
                        "run": str(run),
                        "deposition_s": float(row[deposition_col]),
                        "crystallization_onset_s": float(row[crystallization_col]) if crystallization_col and pd.notna(row.get(crystallization_col)) else np.nan,
                        "detachment_onset_s": float(row[detachment_col]),
                        "channel": channel,
                        "source_file": path.name,
                        "input_status": "loaded_wide_aasted_summary",
                    }
                )
        return pd.DataFrame(rows)

    for _, row in raw.iterrows():
        run = _first_existing(row, ["run", "run_name", "sample", "sample_id", "file", "filename", "comparison_label"])
        deposition = _first_existing(row, ["deposition_s", "deposition", "deposition_time_s", "deposition_absolute_s"])
        crystallization = _first_existing(
            row,
            [
                "crystallization_onset_s",
                "cryst_onset_s",
                "cryst._onset_s",
                "crystallization_start_s",
                "crystallization_onset_absolute_s",
            ],
        )
        detachment = _first_existing(
            row,
            [
                "detachment_onset_s",
                "detachment_onset_absolute_s",
                "detachment_start_s",
                "detachment_onset",
            ],
        )
        channel = _first_existing(row, ["channel", "us_channel", "sensor_channel"])
        if pd.isna(run) or pd.isna(detachment) or pd.isna(deposition):
            continue
        channels = [str(channel)] if pd.notna(channel) and str(channel) in US_CHANNELS else US_CHANNELS
        for ch in channels:
            rows.append(
                {
                    "run": str(run),
                    "deposition_s": float(deposition),
                    "crystallization_onset_s": float(crystallization) if pd.notna(crystallization) else np.nan,
                    "detachment_onset_s": float(detachment),
                    "channel": ch,
                    "source_file": path.name,
                    "input_status": "loaded",
                }
            )
    return pd.DataFrame(rows)


def run_lookup_name(name: str) -> str:
    text = str(name).lower()
    if "ref" in text or "2291" in text:
        return REFERENCE_RUN_NAME
    if "2160-4374" in text or "2160_4374" in text:
        return COMPARISON_RUN_NAME
    if "old" in text or "configuration" in text or "split" in text:
        return COMPARISON_RUN_NAME
    return str(name)


def zone_at_time(zones: pd.DataFrame, run_name: str, elapsed_s: float) -> str:
    if pd.isna(elapsed_s):
        return ""
    hit = zones[
        (zones["run"] == run_name)
        & (zones["detected_start_s"] <= float(elapsed_s))
        & (zones["detected_end_s"] >= float(elapsed_s))
    ]
    return "" if hit.empty else str(hit.iloc[0]["zone"])


def aasted_detachment_analysis(
    runs: dict[str, pd.DataFrame],
    detected_zones: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    landmarks = load_parameter_landmarks()
    if landmarks.empty:
        note = pd.DataFrame(
            [
                {
                    "status": "waiting_for_parameter_summary",
                    "expected_location": str(WORKSPACE / "inputs" / "aasted3"),
                    "expected_filename": "any file containing parameter_summary in the name",
                    "required_columns": "run/sample/file, deposition_s, detachment_onset_s; optional crystallization_onset_s and channel",
                    "method": "Once present, positive upward ultrasound change points are tested against the pre-deposition reference minus 10%.",
                }
            ]
        )
        return note, pd.DataFrame(), landmarks

    cp_rows = []
    summary_rows = []
    for _, landmark in landmarks.iterrows():
        run_name = run_lookup_name(landmark["run"])
        if run_name not in runs:
            continue
        df = runs[run_name]
        channel = landmark["channel"]
        signal_col = f"{channel}_tc" if f"{channel}_tc" in df.columns else channel
        if signal_col not in df.columns:
            continue
        demould = detected_zones[(detected_zones["run"] == run_name) & (detected_zones["zone"] == "demoulding_twisting")]
        demould_start = float(demould.iloc[0]["detected_start_s"]) if not demould.empty else np.nan
        window_end = float(df["elapsed_s"].max())
        onset = float(landmark["detachment_onset_s"])
        seg = df[(df["elapsed_s"] >= onset) & (df["elapsed_s"] <= window_end)].copy()
        if seg.empty:
            continue
        reference = median_window(df, signal_col, float(landmark["deposition_s"]) - 5.0, float(landmark["deposition_s"]))
        threshold = reference - 0.10 * abs(reference)
        cps = detect_us_change_points(seg, signal_col)
        selected = None
        count_until = 0
        first_cp = cps[0] if cps else None
        for idx, cp in enumerate(cps, 1):
            level = us_level_near(seg, signal_col, cp["change_point_s"], 5.0)
            met = bool(pd.notna(level) and level >= threshold)
            if selected is None:
                count_until = idx
            cp_zone = zone_at_time(detected_zones, run_name, cp["change_point_s"])
            cp_rows.append(
                {
                    "run": run_name,
                    "channel": channel,
                    "signal_column": signal_col,
                    "change_point_index": idx,
                    **cp,
                    "detachment_onset_s": onset,
                    "change_point_after_detachment_onset_s": cp["change_point_s"] - onset,
                    "change_point_zone": cp_zone,
                    "analysis_domain": "detachment_onset_to_run_end",
                    "analysis_window_start_s": onset,
                    "analysis_window_end_s": window_end,
                    "deposition_reference_window_s": f"{float(landmark['deposition_s']) - 5.0:.1f}-{float(landmark['deposition_s']):.1f}",
                    "reference_us_level": reference,
                    "threshold_10pct_lower": threshold,
                    "us_level_at_change_point": level,
                    "threshold_met": met,
                    "note": "first passing change point used as detachment offset" if met and selected is None else ("later passing change point" if met else ""),
                }
            )
            if met and selected is None:
                selected = {**cp, "level": level, "idx": idx}
        if selected is not None:
            offset = selected["change_point_s"]
            status = "complete_detachment_for_channel"
            selected_level = selected["level"]
            count_until = selected["idx"]
        elif cps:
            offset = np.nan
            status = "partial_detachment_no_change_point_met_threshold"
            selected_level = us_level_near(seg, signal_col, cps[-1]["change_point_s"], 5.0)
            count_until = len(cps)
        else:
            offset = np.nan
            status = "no_positive_change_point_detected"
            selected_level = np.nan
            count_until = 0
        offset_zone = zone_at_time(detected_zones, run_name, offset)
        offset_zone_row = detected_zones[(detected_zones["run"] == run_name) & (detected_zones["zone"] == offset_zone)]
        offset_zone_start = float(offset_zone_row.iloc[0]["detected_start_s"]) if not offset_zone_row.empty else np.nan
        offset_zone_end = float(offset_zone_row.iloc[0]["detected_end_s"]) if not offset_zone_row.empty else np.nan
        complete_before_demoulding = bool(pd.notna(offset) and pd.notna(demould_start) and offset < demould_start)
        summary_rows.append(
            {
                "run": run_name,
                "channel": channel,
                "signal_column": signal_col,
                "signal_basis": "T7 temperature-normalized ultrasound; channel_tc = raw / polynomial(T7)",
                "deposition_s": float(landmark["deposition_s"]),
                "crystallization_onset_s": landmark["crystallization_onset_s"],
                "detachment_onset_s": onset,
                "analysis_domain": "detachment_onset_to_run_end",
                "analysis_window_start_s": onset,
                "analysis_window_end_s": window_end,
                "first_change_point_s": first_cp["change_point_s"] if first_cp is not None else np.nan,
                "first_change_point_after_onset_s": first_cp["change_point_s"] - onset if first_cp is not None else np.nan,
                "first_change_point_zone": zone_at_time(detected_zones, run_name, first_cp["change_point_s"]) if first_cp is not None else "",
                "detachment_offset_s": offset,
                "detachment_offset_status": status,
                "detachment_offset_zone": offset_zone,
                "detachment_offset_zone_start_s": offset_zone_start,
                "detachment_offset_zone_end_s": offset_zone_end,
                "demoulding_twisting_start_s": demould_start,
                "complete_detachment_before_demoulding": complete_before_demoulding,
                "detachment_onset_to_offset_s": offset - onset if pd.notna(offset) else np.nan,
                "detachment_offset_minus_deposition_s": offset - float(landmark["deposition_s"]) if pd.notna(offset) else np.nan,
                "detachment_onset_minus_deposition_s": onset - float(landmark["deposition_s"]),
                "crystallization_onset_minus_deposition_s": landmark["crystallization_onset_s"] - float(landmark["deposition_s"]) if pd.notna(landmark["crystallization_onset_s"]) else np.nan,
                "reference_us_level": reference,
                "threshold_10pct_lower": threshold,
                "us_level_at_decision": selected_level,
                "change_points_detected_total": len(cps),
                "change_points_until_detachment_or_last": count_until,
                "pattern_description": (
                    f"{len(cps)} positive upward change point(s) from detachment onset to run end. "
                    f"{'First complete offset at ' + f'{offset:.1f}s' if pd.notna(offset) else 'No change point reached the pre-deposition -10% threshold'}."
                ),
                "method": "positive/upward local mean-shift change points on T7-corrected ultrasound; complete detachment if US at change point >= 10%-below-pre-deposition-reference threshold",
                "source_file": landmark["source_file"],
            }
        )
    return pd.DataFrame(summary_rows), pd.DataFrame(cp_rows), landmarks


def viscosity_ratio_summary(runs: dict[str, pd.DataFrame], landmarks: pd.DataFrame) -> pd.DataFrame:
    if landmarks.empty:
        return pd.DataFrame()
    rows = []
    for run_name, df in runs.items():
        run_landmarks = landmarks[landmarks["run"].map(run_lookup_name).eq(run_name)].copy()
        if run_landmarks.empty:
            continue
        deposition = float(pd.to_numeric(run_landmarks["deposition_s"], errors="coerce").dropna().iloc[0])
        ratios = []
        for channel in US_CHANNELS:
            signal_col = f"{channel}_tc" if f"{channel}_tc" in df.columns else channel
            if signal_col not in df.columns:
                continue
            pre = median_window(df, signal_col, deposition - 5.0, deposition)
            post = median_window(df, signal_col, deposition + 50.0, deposition + 55.0)
            ratio = post / pre if pd.notna(pre) and abs(pre) > 1e-12 else np.nan
            ratios.append(ratio)
            rows.append(
                {
                    "run": run_name,
                    "channel": channel,
                    "signal_column": signal_col,
                    "deposition_s": deposition,
                    "pre_deposition_window_s": f"{deposition - 5.0:.1f}-{deposition:.1f}",
                    "post_deposition_window_s": f"{deposition + 50.0:.1f}-{deposition + 55.0:.1f}",
                    "pre_deposition_us_level": pre,
                    "post_50s_us_level": post,
                    "viscosity_ratio_channel": ratio,
                    "viscosity_ratio_mean_rx1rx2": np.nan,
                    "method": "median(T-corrected US 50-55 s after deposition) / median(T-corrected US 5 s before deposition)",
                }
            )
        mean_ratio = float(np.nanmean(ratios)) if ratios else np.nan
        for row in rows:
            if row["run"] == run_name:
                row["viscosity_ratio_mean_rx1rx2"] = mean_ratio
    return pd.DataFrame(rows)


def detachment_homogeneity_summary(detachment_summary: pd.DataFrame) -> pd.DataFrame:
    if detachment_summary.empty or "detachment_offset_s" not in detachment_summary.columns:
        return pd.DataFrame()
    rows = []
    for run, group in detachment_summary.groupby("run", dropna=False):
        offsets = pd.to_numeric(group["detachment_offset_s"], errors="coerce").dropna()
        onsets = pd.to_numeric(group["detachment_onset_s"], errors="coerce").dropna()
        durations = pd.to_numeric(group["detachment_onset_to_offset_s"], errors="coerce").dropna()
        cps = pd.to_numeric(group["change_points_until_detachment_or_last"], errors="coerce").dropna()
        complete = group[group["detachment_offset_status"].eq("complete_detachment_for_channel")]
        rows.append(
            {
                "run": run,
                "channels_evaluated": int(group["channel"].nunique()),
                "channels_complete": int(complete["channel"].nunique()),
                "complete_channels": ", ".join(complete["channel"].astype(str).tolist()),
                "both_primary_channels_complete": bool(set(US_CHANNELS).issubset(set(complete["channel"].astype(str)))),
                "all_complete_before_demoulding": bool((complete["complete_detachment_before_demoulding"] == True).all()) if not complete.empty else False,
                "detachment_onset_spread_s": float(onsets.max() - onsets.min()) if len(onsets) > 1 else 0.0,
                "detachment_offset_spread_s": float(offsets.max() - offsets.min()) if len(offsets) > 1 else np.nan,
                "detachment_duration_spread_s": float(durations.max() - durations.min()) if len(durations) > 1 else np.nan,
                "mean_change_points_until_offset_or_last": float(cps.mean()) if not cps.empty else np.nan,
                "homogeneity_reading": (
                    "homogeneous complete detachment"
                    if len(offsets) >= 2 and (offsets.max() - offsets.min()) <= 30
                    else "heterogeneous or partial detachment"
                    if len(offsets) >= 1
                    else "no complete offset detected"
                ),
            }
        )
    return pd.DataFrame(rows)


def quality_comparison_summary(
    setup: pd.DataFrame,
    landmarks: pd.DataFrame,
    viscosity: pd.DataFrame,
    detachment: pd.DataFrame,
    detected_product_delta: pd.DataFrame,
    hotspot_summary: pd.DataFrame,
) -> pd.DataFrame:
    rows = []
    if setup.empty:
        return pd.DataFrame()
    setup_idx = setup.drop_duplicates("run").set_index("run")
    runs = [REFERENCE_RUN_NAME, COMPARISON_RUN_NAME]
    for run in runs:
        meta = setup_idx.loc[run].to_dict() if run in setup_idx.index else {}
        run_landmarks = landmarks[landmarks["run"].map(run_lookup_name).eq(run)] if not landmarks.empty else pd.DataFrame()
        run_visc = viscosity[viscosity["run"].eq(run)] if not viscosity.empty else pd.DataFrame()
        run_det = detachment[detachment["run"].eq(run)] if not detachment.empty else pd.DataFrame()
        mean_product_delta = 0.0 if run == REFERENCE_RUN_NAME else (
            detected_product_delta.loc[
                detected_product_delta["sensor"].isin(PRODUCT_SENSORS),
                "mean_delta_comparison_minus_reference_C",
            ].mean()
            if "mean_delta_comparison_minus_reference_C" in detected_product_delta.columns
            else np.nan
        )
        rows.append(
            {
                "run": run,
                "quality_0_not_good_1_good": meta.get("quality_0_not_good_1_good", np.nan),
                "remarks": meta.get("remarks", ""),
                "chocolate": meta.get("chocolate", ""),
                "protocol_duration_s": meta.get("protocol_duration_s", np.nan),
                "viscosity_ratio_mean_rx1rx2": run_visc["viscosity_ratio_mean_rx1rx2"].dropna().iloc[0] if not run_visc.empty and run_visc["viscosity_ratio_mean_rx1rx2"].notna().any() else np.nan,
                "crystallization_onset_minus_deposition_mean_s": (run_landmarks["crystallization_onset_s"] - run_landmarks["deposition_s"]).mean() if not run_landmarks.empty else np.nan,
                "detachment_onset_minus_deposition_mean_s": (run_landmarks["detachment_onset_s"] - run_landmarks["deposition_s"]).mean() if not run_landmarks.empty else np.nan,
                "detachment_offset_minus_deposition_mean_s": run_det["detachment_offset_minus_deposition_s"].mean() if not run_det.empty else np.nan,
                "complete_detachment_channels": ", ".join(run_det.loc[run_det["detachment_offset_status"].eq("complete_detachment_for_channel"), "channel"].astype(str).tolist()) if not run_det.empty else "",
                "complete_before_demoulding_channels": ", ".join(run_det.loc[run_det["complete_detachment_before_demoulding"].eq(True), "channel"].astype(str).tolist()) if not run_det.empty else "",
                "mean_change_points_until_offset_or_last": run_det["change_points_until_detachment_or_last"].mean() if not run_det.empty else np.nan,
                "mean_product_delta_vs_reference_C": mean_product_delta,
                "mean_hotspot_spread_C": hotspot_summary.loc[hotspot_summary["run"].eq(run), "spread_hottest_minus_coolest_C"].mean() if not hotspot_summary.empty else np.nan,
            }
        )
    out = pd.DataFrame(rows)
    if len(out) == 2:
        ref = out[out["run"].eq(REFERENCE_RUN_NAME)].iloc[0]
        comp = out[out["run"].eq(COMPARISON_RUN_NAME)].iloc[0]
        delta_rows = []
        for metric in [
            "viscosity_ratio_mean_rx1rx2",
            "crystallization_onset_minus_deposition_mean_s",
            "detachment_onset_minus_deposition_mean_s",
            "detachment_offset_minus_deposition_mean_s",
            "mean_change_points_until_offset_or_last",
            "mean_hotspot_spread_C",
        ]:
            delta_rows.append(
                {
                    "run": "comparison_minus_reference",
                    "quality_0_not_good_1_good": "",
                    "remarks": f"{metric}: {comp.get(metric, np.nan) - ref.get(metric, np.nan):.3f}" if pd.notna(comp.get(metric, np.nan)) and pd.notna(ref.get(metric, np.nan)) else f"{metric}: n/a",
                    "chocolate": "",
                    "protocol_duration_s": np.nan,
                    "viscosity_ratio_mean_rx1rx2": np.nan,
                    "crystallization_onset_minus_deposition_mean_s": np.nan,
                    "detachment_onset_minus_deposition_mean_s": np.nan,
                    "detachment_offset_minus_deposition_mean_s": np.nan,
                    "complete_detachment_channels": "",
                    "complete_before_demoulding_channels": "",
                    "mean_change_points_until_offset_or_last": np.nan,
                    "mean_product_delta_vs_reference_C": np.nan,
                    "mean_hotspot_spread_C": np.nan,
                }
            )
        out = pd.concat([out, pd.DataFrame(delta_rows)], ignore_index=True)
    return out


def export_temperature_corrected_raw_inputs() -> pd.DataFrame:
    CORRECTED_RAW_DIR.mkdir(parents=True, exist_ok=True)
    rows = []
    input_dir = WORKSPACE / "inputs" / "aasted3"
    for csv_path in sorted(input_dir.glob("*.csv")):
        if csv_path.name == TEMPERATURE_CORRECTION_CSV.name:
            continue
        try:
            df = read_run(csv_path)
        except Exception as exc:
            rows.append({"source_file": csv_path.name, "status": f"skipped: {exc}", "output_file": ""})
            continue
        output_path = CORRECTED_RAW_DIR / f"{csv_path.stem}_temperature_corrected.csv"
        df.to_csv(output_path, index=False)
        rows.append(
            {
                "source_file": csv_path.name,
                "status": "exported",
                "output_file": str(output_path.relative_to(WORKSPACE)),
                "rows": int(df.shape[0]),
                "temperature_sensor": TEMP_CORRECTION_SENSOR,
                "corrected_channels": ", ".join([f"{ch}_tc" for ch in ALL_US_CHANNELS if f"{ch}_tc" in df.columns]),
            }
        )
    return pd.DataFrame(rows)


def zone_findings(duration: pd.DataFrame, mech_delta: pd.DataFrame, product_delta: pd.DataFrame) -> pd.DataFrame:
    rows = []
    descriptions = {name: desc for _, _, name, desc in ZONES}
    duration_lookup = duration.set_index("zone")
    mech_lookup = mech_delta.set_index("zone")
    for _, _, zone, _ in ZONES:
        if zone not in duration_lookup.index:
            continue
        prod_zone = product_delta[(product_delta["zone"] == zone) & (product_delta["sensor"].isin(PRODUCT_SENSORS))]
        spread_zone = product_delta[(product_delta["zone"] == zone) & (product_delta["sensor"] == "PRODUCT_SPREAD")]
        mean_prod_delta = prod_zone["mean_delta_comparison_minus_reference_C"].mean()
        max_abs_sensor = prod_zone.loc[
            prod_zone["mean_delta_comparison_minus_reference_C"].abs().idxmax(), "sensor"
        ] if not prod_zone.empty else ""
        max_abs_delta = prod_zone["mean_delta_comparison_minus_reference_C"].abs().max() if not prod_zone.empty else np.nan
        spread_delta = (
            spread_zone["mean_delta_comparison_minus_reference_C"].iloc[0]
            if not spread_zone.empty
            else np.nan
        )
        notable = mech_lookup.loc[zone, "notable_differences"] if zone in mech_lookup.index else ""
        extra = duration_lookup.loc[zone, "extra_duration_s"]
        rows.append(
            {
                "zone": zone,
                "process_interpretation": descriptions[zone],
                "reference_duration_s": duration_lookup.loc[zone, "reference_duration_s"],
                "comparison_pattern_duration_s": duration_lookup.loc[zone, "comparison_pattern_duration_s"],
                "extra_duration_s": extra,
                "duration_reading": (
                    "main stretched section"
                    if extra >= 60
                    else "slightly longer"
                    if extra > 5
                    else "similar length"
                    if extra >= -5
                    else "shorter after alignment"
                ),
                "mechanical_reading": notable if notable else "no large zone-level IMU difference by this summary",
                "mean_product_delta_C": mean_prod_delta,
                "largest_sensor_delta": f"{max_abs_sensor} ({max_abs_delta:.2f} C abs mean delta)" if max_abs_sensor else "",
                "product_spread_delta_C": spread_delta,
                "temperature_reading": (
                    "comparison warmer than reference"
                    if mean_prod_delta >= 0.5
                    else "comparison cooler than reference"
                    if mean_prod_delta <= -0.5
                    else "similar mean product temperature"
                ),
            }
        )
    return pd.DataFrame(rows)


def write_df(ws, df: pd.DataFrame) -> None:
    for col_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(1, col_idx, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            if isinstance(val, (float, np.floating)) and not pd.isna(val):
                val = float(val)
            elif isinstance(val, (int, np.integer)) and not pd.isna(val):
                val = int(val)
            elif pd.isna(val):
                val = None
            ws.cell(row_idx, col_idx, val)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, col in enumerate(df.columns, 1):
        samples = [str(col), *[str(v) for v in df[col].head(200).fillna("").tolist()]]
        ws.column_dimensions[get_column_letter(idx)].width = min(max(len(s) for s in samples) + 2, 34)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                if isinstance(cell.value, float):
                    cell.number_format = "0.000"
    wb.save(path)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = BASE_REPORT_WORKBOOK

    comparison = read_run(COMPARISON_CSV)
    reference = read_run(REFERENCE_CSV)
    comparison_seconds = build_seconds(comparison)
    reference_seconds = build_seconds(reference)
    map_df, dtw_cost = dtw_map(comparison_seconds, reference_seconds)
    detected_reference = detect_pattern_zones(reference_seconds, REFERENCE_RUN_NAME)
    detected_comparison = detect_pattern_zones(comparison_seconds, COMPARISON_RUN_NAME)
    detected_zones = pd.concat([detected_reference, detected_comparison], ignore_index=True)
    detachment_summary, detachment_change_points, parameter_landmarks = aasted_detachment_analysis(
        {REFERENCE_RUN_NAME: reference, COMPARISON_RUN_NAME: comparison},
        detected_zones,
    )
    experimental_setup = load_experimental_setup()
    viscosity_summary = viscosity_ratio_summary(
        {REFERENCE_RUN_NAME: reference, COMPARISON_RUN_NAME: comparison},
        parameter_landmarks,
    )
    detachment_homogeneity = detachment_homogeneity_summary(detachment_summary)
    corrected_raw_inventory = export_temperature_corrected_raw_inputs()
    temperature_correction = pd.read_csv(TEMPERATURE_CORRECTION_CSV) if TEMPERATURE_CORRECTION_CSV.exists() else pd.DataFrame()

    comparison_temp = temp_with_zones(comparison, COMPARISON_RUN_NAME, map_df)
    reference_temp = temp_with_zones(reference, REFERENCE_RUN_NAME)
    temp_all = pd.concat([reference_temp, comparison_temp], ignore_index=True)
    detected_reference_temp = temp_with_detected_zones(reference, REFERENCE_RUN_NAME, detected_reference)
    detected_comparison_temp = temp_with_detected_zones(comparison, COMPARISON_RUN_NAME, detected_comparison)
    detected_temp_all = pd.concat([detected_reference_temp, detected_comparison_temp], ignore_index=True)

    duration = duration_by_zone(reference_seconds, map_df)
    mech_delta = mechanical_delta(reference_seconds, comparison_seconds, map_df)
    product_summary = product_zone_summary(temp_all)
    product_delta = product_zone_delta(product_summary)
    hotspots = hotspot_by_zone(temp_all)
    findings = zone_findings(duration, mech_delta, product_delta)
    detected_product_summary = product_zone_summary(detected_temp_all)
    detected_product_delta = product_zone_delta(detected_product_summary)
    hotspot_summary, hotspot_long, hotspot_delta_matrix = clearer_hotspots(detected_temp_all)
    quality_comparison = quality_comparison_summary(
        experimental_setup,
        parameter_landmarks,
        viscosity_summary,
        detachment_summary,
        detected_product_delta,
        hotspot_summary,
    )

    detected_duration = detected_zones.pivot_table(index="zone", columns="run", values="duration_s").reset_index()
    detected_duration.columns.name = None
    if COMPARISON_RUN_NAME in detected_duration.columns and REFERENCE_RUN_NAME in detected_duration.columns:
        detected_duration["duration_delta_comparison_minus_reference_s"] = (
            detected_duration[COMPARISON_RUN_NAME] - detected_duration[REFERENCE_RUN_NAME]
        )
    order = {name: i for i, (_, _, name, _) in enumerate(ZONES)}
    detected_duration["order"] = detected_duration["zone"].map(order)
    detected_duration = detected_duration.sort_values("order").drop(columns=["order"])

    zone_defs = pd.DataFrame(
        [
            {
                "reference_start_s": start,
                "reference_end_s": end if end < 99999 else "run_end",
                "zone": name,
                "interpretation": description,
            }
            for start, end, name, description in ZONES
        ]
    )

    extra_wall = (comparison["elapsed_s"].max() - reference["elapsed_s"].max())
    summary = pd.DataFrame(
        [
            ["reference_run", REFERENCE_CSV.name],
            ["comparison_run", COMPARISON_CSV.name],
            ["comparison_label", COMPARISON_RUN_NAME],
            ["method", "Pattern-based DTW alignment using T8, acc z mean/std, gyro y mean/std per second; ultrasound analyses use T7-temperature-corrected US columns."],
            ["new_in_this_version", "Adds T7 polynomial correction for ultrasound, viscosity ratio, Aasted parameter-summary landmarks, and detachment offset/quality comparison."],
            ["aasted_detachment_architecture", "Reads aa3_trials experimental summary/parameter_summary from inputs/aasted3 and populates crystallization onset, detachment onset, and T-corrected ultrasound detachment offsets."],
            ["ultrasound_temperature_correction", "Corrected channel definition: channel_tc = raw channel / polynomial(T7). Coefficients are stored in inputs/aasted3/aasted3_temperature_correction_coefficients.csv."],
            ["extra_wall_time_s", extra_wall],
            ["meaning_of_extra_wall_time", "Comparison run elapsed duration minus reference elapsed duration; not itself a stop window."],
            ["dtw_cost", dtw_cost],
            ["largest_zone_extra_s", duration.loc[duration["extra_duration_s"].idxmax(), "zone"]],
            ["largest_zone_extra_duration_s", duration["extra_duration_s"].max()],
            ["product_sensors", ", ".join(PRODUCT_SENSORS)],
            ["ambient/location_marker", "T8"],
        ],
        columns=["metric", "value"],
    )

    readme = pd.DataFrame(
        [
            ["rolling_window", "A moving local summary, e.g. median over nearby 30 seconds. Useful for smoothing but poor for distinguishing expected quiet zones from true process stops."],
            ["extra_wall_time", "The comparison file duration minus the reference duration. This is only the duration difference between files."],
            ["low_motion_threshold", "The prior report took the 5th percentile of reference low-motion behavior as a cut-off. That can mislabel normal quiet zones as stops."],
            ["pattern_alignment", "This report aligns the comparison run to the reference process pattern using T8, acc z, and gyro y features, then asks which zones absorbed extra time."],
            ["zone_extra_duration", "For each reference process zone, this is comparison pattern duration minus reference duration after alignment."],
            ["temperature_comparison", "Product temperatures are compared inside the aligned process zones, not only against absolute wall-clock time."],
            ["product_hotspot_definition", "For the remade hotspot sheets, a product hotspot means a product sensor (T2, T3, T4, T5, T7) is warmer than the product-sensor average within the same detected zone."],
            ["pattern_detected_zones", "Experimental per-run segmentation from T8 valleys, sustained gyro-y variability, acc-z variability, and late T8 rise. Early zones are lower confidence than vibration/cooling cycles."],
            ["demoulding_subphases", "The former broad demoulding zone is split into demoulding_twisting, demoulding_vibration, and final_demoulding from acc-z/acc-x/gyro-y patterns."],
            ["aasted_detachment_offset", "If parameter_summary/experimental summary is available, ultrasound detachment offset uses positive/upward change points in T-corrected US and a pre-deposition reference minus 10% threshold, analogous to the lab-trial logic. Search window is detachment onset to run end."],
            ["viscosity_ratio", "For Rx1Tx1 and Rx2Tx2: median T-corrected US 50-55 s after deposition divided by median T-corrected US in the 5 s before deposition; the report also stores the two-channel mean."],
        ],
        columns=["term", "plain_language_explanation"],
    )

    path_sample = map_df.iloc[:: max(1, len(map_df) // 500)].copy()
    path_sample["comparison_zone_assigned"] = path_sample["reference_zone"]

    wb = Workbook()
    wb.remove(wb.active)
    sheets = {
        "Read Me": readme,
        "Summary": summary,
        "Zone Definitions": zone_defs,
        "Zone Findings": findings,
        "Detected Zones": detected_zones,
        "Detected Duration": detected_duration,
        "Zone Duration": duration,
        "Mechanical By Zone": mech_delta,
        "Product Temp By Zone": product_summary,
        "Product Delta By Zone": product_delta,
        "Old Hotspots By Zone": hotspots,
        "Hotspot Summary": hotspot_summary,
        "Hotspot Sensor Data": hotspot_long,
        "Hotspot Delta Matrix": hotspot_delta_matrix,
        "Temperature Correction": temperature_correction,
        "Corrected Raw Inventory": corrected_raw_inventory,
        "Experimental Setup": experimental_setup,
        "Viscosity Ratio": viscosity_summary,
        "Aasted Detachment Summary": detachment_summary,
        "Aasted US Change Points": detachment_change_points,
        "Aasted Detachment Homogeneity": detachment_homogeneity,
        "Aasted Parameter Landmarks": parameter_landmarks,
        "Quality Comparison": quality_comparison,
        "Detected Product By Zone": detected_product_summary,
        "Detected Product Delta": detected_product_delta,
        "Alignment Path Sample": path_sample,
    }
    for name, df in sheets.items():
        ws = wb.create_sheet(name)
        write_df(ws, df)

    ws = wb["Zone Duration"]
    if ws.max_row > 1:
        ws.conditional_formatting.add(
            f"D2:D{ws.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="63BE7B",
                mid_type="num",
                mid_value=0,
                mid_color="FFFFFF",
                end_type="max",
                end_color="F8696B",
            ),
        )
        chart = BarChart()
        chart.title = "Extra Duration by Process Zone"
        chart.y_axis.title = "seconds"
        chart.x_axis.title = "zone"
        data = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18
        ws.add_chart(chart, "F2")

    ws = wb["Detected Duration"]
    if ws.max_row > 1 and ws.max_column >= 4:
        ws.conditional_formatting.add(
            f"D2:D{ws.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="63BE7B",
                mid_type="num",
                mid_value=0,
                mid_color="FFFFFF",
                end_type="max",
                end_color="F8696B",
            ),
        )
        chart = BarChart()
        chart.title = "Detected Zone Duration Delta"
        chart.y_axis.title = "seconds"
        chart.x_axis.title = "zone"
        data = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18
        ws.add_chart(chart, "F2")

    ws = wb["Product Delta By Zone"]
    if ws.max_row > 1 and ws.max_column >= 5:
        ws.conditional_formatting.add(
            f"E2:E{ws.max_row}",
            ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="num", mid_value=0, mid_color="FFFFFF", end_type="max", end_color="F8696B"),
        )

    ws = wb["Hotspot Delta Matrix"]
    if ws.max_row > 1 and ws.max_column >= 7:
        ws.conditional_formatting.add(
            f"C2:G{ws.max_row}",
            ColorScaleRule(start_type="min", start_color="4472C4", mid_type="num", mid_value=0, mid_color="FFFFFF", end_type="max", end_color="F8696B"),
        )

    ws = wb["Hotspot Summary"]
    if ws.max_row > 1:
        spread_col = None
        spread_color_col = None
        sensor_color_cols = []
        for cell in ws[1]:
            if cell.value == "spread_hottest_minus_coolest_C":
                spread_col = cell.column
            if cell.value == "spread_color":
                spread_color_col = cell.column
            if isinstance(cell.value, str) and cell.value.endswith("_color"):
                sensor_color_cols.append(cell.column)
        if spread_col:
            col = get_column_letter(spread_col)
            ws.conditional_formatting.add(
                f"{col}2:{col}{ws.max_row}",
                ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="F8696B"),
            )
            chart = BarChart()
            chart.title = "Product Hotspot Spread by Detected Zone"
            chart.y_axis.title = "deg C"
            chart.x_axis.title = "run/zone row"
            data = Reference(ws, min_col=spread_col, min_row=1, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.height = 8
            chart.width = 18
            ws.add_chart(chart, "U2")
        for row in range(2, ws.max_row + 1):
            if spread_color_col:
                color = str(ws.cell(row, spread_color_col).value or "").replace("#", "")
                if len(color) == 6:
                    for col_idx in ["spread_severity", "spread_color"]:
                        header_lookup = {c.value: c.column for c in ws[1]}
                        if col_idx in header_lookup:
                            ws.cell(row, header_lookup[col_idx]).fill = PatternFill("solid", fgColor=color)
            for col in sensor_color_cols:
                color = str(ws.cell(row, col).value or "").replace("#", "")
                if len(color) == 6 and color != "FFFFFF":
                    ws.cell(row, col).fill = PatternFill("solid", fgColor=color)

    # Add compact trajectory sheet for charting the alignment.
    chart_df = reference_seconds[["elapsed_sec", "T8", "accz_mean", "gyroy_mean", "zone_reference"]].iloc[::2].copy()
    ws = wb.create_sheet("Reference Pattern")
    write_df(ws, chart_df)
    chart = LineChart()
    chart.title = "Reference Pattern: T8, acc z, gyro y"
    chart.y_axis.title = "scaled raw values"
    chart.x_axis.title = "elapsed s"
    data = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 20
    ws.add_chart(chart, "G2")

    chart_df = comparison_seconds[["elapsed_sec", "T8", "accz_mean", "gyroy_mean"]].iloc[::2].copy()
    ws = wb.create_sheet("Comparison Pattern")
    write_df(ws, chart_df)
    chart = LineChart()
    chart.title = "Comparison Pattern: T8, acc z, gyro y"
    chart.y_axis.title = "scaled raw values"
    chart.x_axis.title = "elapsed s"
    data = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 20
    ws.add_chart(chart, "F2")

    wb.save(out_path)
    style_workbook(out_path)
    print(out_path)


if __name__ == "__main__":
    main()
